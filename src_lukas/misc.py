import yaml
import re
import pandas as pd
import seaborn as sns
import numpy as np
import openpyxl
import string
from sklearn.metrics import r2_score, roc_curve, precision_recall_curve
from sklearn.model_selection import GroupKFold
from collections import OrderedDict
from sklearn.model_selection import GridSearchCV
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from sklearn.linear_model import Ridge, Lasso, ElasticNet, LogisticRegression, LinearRegression
from sklearn.linear_model import MultiTaskLasso, MultiTaskElasticNet, LogisticRegressionCV
from sklearn.linear_model import HuberRegressor
from sklearn.experimental import enable_iterative_imputer
from sklearn.impute import IterativeImputer
import openpyxl
from MTLR import MultiTaskLogisticRegression, MultiTaskRocAucScorerWithNulls, MultiTaskStratifiedKFoldWithNulls
from sklearn.ensemble import RandomForestRegressor, RandomForestClassifier, GradientBoostingClassifier, \
    GradientBoostingRegressor
from sklearn.pipeline import Pipeline
import os
import inspect
# import gseapy
import subprocess
import datetime
from bcg_utils import _encode_coef
from pandas.api.types import is_numeric_dtype
from bcg_utils import *
from bcg_colors import *

RANDOM_STATE = 0
DATA = '../data'  # '/scratch/lab_bsf/projects/BSA_0322_BCG'
METADATA = '../metadata'  # '/home/dbarreca/src/BCG_final/metadata'
PROJECT = 'BCG'
ML_RESULTS_ROOT = 'results/ML_final'
LOLA_DB_DIR = os.path.join('resources', 'LOLA')

MARKERS = {'CD14': 'Monocytes (CD14+)', 'FCGR3A': 'Neutrophils (CD16+)', 'CD3G': 'T cells (CD3+)',
           'CD4': 'Helper T cells (CD4+)', 'CD8A': 'Killer T cells (CD8+)',
           'CD19': 'B cells (CD19+)', 'NCAM1': 'NK cells (CD56+)'}

NO_DIFF_COLS = ['DONOR:SEX', 'DONOR:AGE', 'DONOR:BMI', 'DONOR:WEIGHT', 'DONOR:HEIGHT',
                'DONOR:oralContraceptivesIncludingMen', 'DONOR:IC_alcoholInLast24h',
                'DONOR:IC_TIME_REAL', 'DONOR:IC_MONTH_REAL', 'DONOR:IC_DAYS_FROM_0101',
                'DONOR:IC_DATE_2PI_SIN', 'DONOR:IC_DATE_2PI_COS']

VISIT_TO_DAYS = dict(V2='14', V3='90')

DONOR_EXTRA_VISIT_COLUMNS_PREFIXES = ['alcoholInLast24h', 'alcohol_unitsLast24h', 'sleepHoursLastNight',
                                      'sport_yesNo24hBeforeAppointment', 'sport_hoursIn24hBeforeAppointment',
                                      'stress_dayBefore']

GSEA_WEIGHT_SCHEMES = {0: 'classic', 1: 'weighted', 1.5: 'weighted_1.5', 2: 'weighted_p2'}

PEAK_FEATURE_ORDER = {'characterization': ['TSS', 'TSS_proximal', 'gene_body', 'distal', 'intergenic'],
                      REGULATORY_ANNOT_COL: [PROMOTER, PROMOTER_FLANK, ENHANCER, TF_BINDING,
                                      CTCF_BINDING, OPEN_CHROMATIN, UNASSIGNED],
                      FEATURE_TYPE_COL: [PROTEIN_CODING, LNCRNA, IGTR],
                      GENOMIC_LOCATION_COL: [TSS_PROXIMAL_LABEL, GENE_BODY_LABEL, DISTAL_LABEL, INTERGENIC_LABEL]}

BLOOD = [
    'PBMC_PERC:MONO', 'PBMC_PERC:MONO/INTERMEDIATE', 'PBMC_PERC:MONO/NON_CLASSICAL',
    'PBMC_PERC:T/CD8', 'PBMC_PERC:T/CD4', 'PBMC_PERC:T/CD4/TREG',
    'PBMC_PERC:B', 'PBMC_PERC:NK', 'PBMC_PERC:NKT',
    'PBMC_PERC:BASO', 'PBMC_PERC:NEUTRO'
]

WHOLE_BLOOD = [
    'WB_PER_ML:MONO', 'WB_PER_ML:MONO/INTERMEDIATE', 'WB_PER_ML:MONO/NON_CLASSICAL',
    'WB_PER_ML:T/CD8', 'WB_PER_ML:T/CD4', 'WB_PER_ML:T/CD4/TREG',
    'WB_PER_ML:B', 'WB_PER_ML:NK', 'WB_PER_ML:NKT',
    'WB_PER_ML:NEUTRO'
]

CYTO_MODEL_COLS = ['SAMPLE:VISIT', 'SAMPLE:DONOR', 'SAMPLE:VISIT_TIME_REAL'] + BLOOD
CM_MODEL_COLS = ['SAMPLE:VISIT', 'SAMPLE:DONOR', 'SAMPLE:VISIT_TIME_REAL'] + WHOLE_BLOOD

NAN_BOOLS_ANNOT_COLS = [
    'QC:PASS', 'LAB:WRONG_BATCH', 'DONOR:IC_EVENING',
    'SAMPLE:sport_yesNo24hBeforeAppointment', 'SAMPLE:alcoholInLast24h',
    'DONOR:IC_sport_yesNo24hBeforeAppointment', 'DONOR:IC_alcoholInLast24h',
    'DONOR:scarPresent_V3', 'DONOR:scarSize_v3_binary', 'DONOR:sizeOfVaccinationSpot_v2_binary', 'DONOR:currentSmoker',
    'DONOR:motherBcg', 'DONOR:fatherBcg', 'DONOR:parentBcg', 'DONOR:motherPox', 'DONOR:fatherPox',
    'DONOR:oralContraceptives', 'DONOR:oralContraceptivesIncludingMen',
    'DONOR:postMenopausal_binary', 'DONOR:postMenopausalIncludingMen_binary',
    'DONOR:currentMenopausal', 'DONOR:currentMenopausalIncludingMen',
    'DONOR:IC_stress_dayBefore_binary', 'DONOR:exerciseDuringDay_binary', 'DONOR:happiness_binary',
    'DONOR:mentalHealth_3weeksBefore_binary', 'DONOR:postMenopausalIncludingMen_binary', 'DONOR:postMenopausal_binary',
    'DONOR:qualityTeeth_binary', 'DONOR:sleepQuality_lastTwoWeeks_v1_binary', 'DONOR:stress_2weeksBefore_binary',
    'DONOR:vaginalBirth_binary', 'SAMPLE:stress_dayBefore_binary'
]

NAN_STRING_ANNOT_COLS = ['DONOR:SNP_OUTLIER', 'SAMPLE:EXCLUSION']


def safe_R_name(s, as_list=True):
    if isinstance(s, Iterable) and not isinstance(s, string_types):
        _map = map(lambda x: safe_R_name(x), s)
        return list(_map) if as_list else _map
    else:
        return s.replace(':', '.').replace('/', '.').replace('-', '.')


GENE_SETS_FIXES = {
    ' pathway$': '',
    ' Pathway$': '',
    ' PATHWAY$': '',
}

LOLA_CELLTYPES_FIXES = {
    ' cell$': '',
    ' cells$': '',
    ' tissue$': '',
    'Leukaemia': 'Leukemia',
    'leukaemia': 'leukemia',
    'acute lymphoblastic leukemia': 'ALL',
    'Umbilical Cord Blood Stem and Progenitor Cells': 'UCB HSPC',
    'Embryonic Stem Cell': 'ESC',
    'Hematopoietic Stem and Progenitor Cells': 'HSPC',
    'Acute Myeloid Leukemia': 'AML'
}

GENCODE_RESOURCE_DIR = os.path.join("resources", "gencode")
GENCODE_HUMAN_GTF_FN = os.path.join(GENCODE_RESOURCE_DIR, "gencode.v31.basic.annotation.gtf")
GENCODE_MOUSE_GTF_FN = os.path.join(GENCODE_RESOURCE_DIR, "gencode.vM22.basic.annotation.gtf")
GENCODE_GTF_FN = {'hs': GENCODE_HUMAN_GTF_FN, 'mm': GENCODE_MOUSE_GTF_FN}

UROPA_CONFIG_FN = os.path.join(METADATA, 'uropa_gencode_config.json')
ENSEMBL_MOUSE_ORTHOLOGS_FN = os.path.join(METADATA, 'ens_human_to_mouse_r97.tsv')
HUMAN_ENSEMBL_IDS_FN = os.path.join(METADATA, 'human_ens_to_gene_r97.tsv')
MOUSE_ENSEMBL_IDS_FN = os.path.join(METADATA, 'mouse_ens_to_gene_r97.tsv')

SAMPLE_ANNOT_FN = os.path.join(METADATA, 'complete_metadata{which}.csv')
STIMULI_NAMES_FN = os.path.join(METADATA, 'StimulusToPrettyNames.xlsx')
IMPUTED_CYTO_FN = os.path.join(METADATA, 'imputed_cytokines.csv')

DONOR_DATA_FN = None
DONOR_BMI_FN = None
CASTOR_METADATA_FN = None
DONOR_EXTRA_DATA_FN = None
EXCLUSION_FN = None
GENETIC_OUTLIERS_FN = None
SYSMEX_DATA_FN = None
FACS_DATA_FN = None

PEAK_ANNOT_ALL_FN = os.path.join(DATA, 'DE', 'peaks_filtered_PBMC.csv.gz')
COUNT_DATA_FN = os.path.join(DATA, 'filtered_normalized', 'BCG.{celltype}.counts.{data}.csv.gz')
NORM_COUNT_DATA_FN = os.path.join(DATA, 'DE', 'normalized_log2_CPM_{celltype}.csv.gz')
COMBAT_COUNT_DATA_FN = os.path.join(DATA, 'DE', 'batch_corrected_normalized_log2_CPM_{celltype}.csv.gz')
COMBAT_COUNT_CELLTYPES_DATA_FN = os.path.join(DATA, 'DE', 'batch_corrected_per_celltype_normalized_log2_CPM_{celltype}.csv.gz')
RAW_COUNT_DATA_FN = os.path.join(DATA, 'DE', 'quantification_filtered_{celltype}.csv.gz')
DESEQ_DATA_FN = os.path.join('results', 'DESeq2', 'deseq2_{celltype}.{model}.{data}_counts.csv.gz')

SELECTED_RAW_CYTO_FN = None
RAW_CYTO_FN = None
CORR_CYTO_FN = None
CYTO_QUAL_FN = None
CORR_CIRC_MEDIATORS_FN = None

ROB_BLOOD_FN = None
CORONA_PERIOD_ANNOT_FN = None
CIRCADIAN_DATA_FN = None

ML_RESULTS_FN = os.path.join(ML_RESULTS_ROOT, 'BCG.ML.results.hdf')
SEASON_CORRECTED_CYTO = os.path.join('results', 'LR', 'PBMC.final.CYTO.BLOOD', 'CYTO.corrected.fixed_factor_interactions.LMM.DONOR:IC_DATE_2PI_SIN__DONOR:IC_DATE_2PI_COS.csv')

FIX_LIMMA_COEFS = {
    'sin___DATE_2PI': 'sin__SAMPLE.VISIT_DATE_2PI',
    'cos___DATE_2PI': 'cos__SAMPLE.VISIT_DATE_2PI',
    'X_DATE_2PI': 'SAMPLE.VISIT_DATE_2PI',
    'V2.sin___DATE_2PI': 'sin__SAMPLE.VISIT_DATE_2PI.V2',
    'V3.sin___DATE_2PI': 'sin__SAMPLE.VISIT_DATE_2PI.V3',
    'V2.cos___DATE_2PI': 'cos__SAMPLE.VISIT_DATE_2PI.V2',
    'V3.cos___DATE_2PI': 'cos__SAMPLE.VISIT_DATE_2PI.V3',
    'V2.DONOR.IC_TIME_REAL': 'DONOR.IC_TIME_REAL.V2',
    'V3.DONOR.IC_TIME_REAL': 'DONOR.IC_TIME_REAL.V3'
}

ESTIMATORS = {
    'RFC': (RandomForestClassifier, dict(random_state=None), ['n_estimators']),
    'RFR': (RandomForestRegressor, dict(random_state=None), ['n_estimators']),
    'LR_L1': (LogisticRegression, dict(penalty='l1', solver='liblinear', random_state=None, max_iter=1000), ['C']),
    'Lasso': (Lasso, dict(random_state=None), ['alpha']),
    'LR_L2': (LogisticRegression, dict(penalty='l2', solver='liblinear', random_state=None, max_iter=1000), ['C']),
    'LR': (LogisticRegression, dict(penalty='none', solver='lbfgs', random_state=None, max_iter=1000), []),
    'BaggingClassifier': (BaggingClassifier, dict(base_estimator=LogisticRegressionCV(penalty='l2', solver='saga'), random_state=None), ['n_estimators']),
    'RidgeClassifierCV': (RidgeClassifierCV, dict(store_cv_values=True), []),
    'Ridge': (Ridge, dict(random_state=None), ['alpha']),
    'LR_Enet': (LogisticRegression, dict(penalty='elasticnet', solver='saga', random_state=None, max_iter=1000), ['C', 'l1_ratio']),
    'SGDC_Enet': (SGDClassifier, dict(loss='hinge', penalty='elasticnet', random_state=None), ['alpha', 'l1_ratio']),
    'Enet': (ElasticNet, dict(random_state=None), ['alpha', 'l1_ratio']),
    'HuberRegressor': (HuberRegressor, dict(), ['epsilon', 'alpha']),
    'GBC': (GradientBoostingClassifier, dict(random_state=None), ['n_estimators', 'learning_rate']),
    'GBR': (GradientBoostingRegressor, dict(random_state=None), ['n_estimators', 'learning_rate']),
    'SVM': (SVC, dict(kernel='linear', random_state=None), ['C']),
    'KernelRidge': (KernelRidge, dict(kernel='poly'), ['alpha', 'degree']),
    'MultiTask_Lasso': (MultiTaskLasso, dict(random_state=None), ['alpha']),
    'MultiTask_Enet': (MultiTaskElasticNet, dict(random_state=None), ['alpha', 'l1_ratio']),
    'MultiTask_LR_L21': (MultiTaskLogisticRegression, dict(random_state=None), ['C']),
    'MultiTask_LR_Enet': (MultiTaskLogisticRegression, dict(random_state=None), ['C', 'l1_ratio'])
}


def _no_diff_cols(index):
    return index.str.contains('^DONOR:|^thm\.|^IC_[A-Z_]+:')


def visit_diff(df, visit, base_visit='V1', no_diff_cols=_no_diff_cols, diff_op='-'):
    """
    lfc_atac_V2_df = visit_diff(atac_df, visit='V2', base_visit='V1')
    lfc_cyto_V2_df = visit_diff(cyto_df, visit='V2', base_visit='V1')
    """
    return df_diff(df=df, diff_col='SAMPLE:VISIT', new=visit, base=base_visit,
                   no_diff_cols=no_diff_cols, diff_op=diff_op)


def df_diff(df, diff_col, new, base, no_diff_cols=_no_diff_cols, diff_op='-', verbose=False):
    assert diff_op in ['-', '/']
    assert all([name in df.index.names for name in ['SAMPLE:DONOR', diff_col]])
    no_diff_col_mask = df.columns.isin(no_diff_cols) if is_iterable(no_diff_cols) else no_diff_cols(df.columns)
    drop_levels = [n for n in df.index.names if n != 'SAMPLE:DONOR']
    new_df = df.loc[df.index.get_level_values(diff_col) == new, ~no_diff_col_mask].reset_index(
        level=drop_levels, drop=True)
    base_df = df.loc[df.index.get_level_values(diff_col) == base, ~no_diff_col_mask].reset_index(
        level=drop_levels, drop=True)
    no_diff_df = df.loc[df.index.get_level_values(diff_col) == base, no_diff_col_mask].reset_index(
        level=drop_levels, drop=True)
    if verbose:
        print('diff_op:', diff_op)
        print('new_df:', new_df.columns)
        print('base_df:', base_df.columns)
        print('no_diff_df:', no_diff_df.columns)
    assert new_df.index.is_unique
    assert base_df.index.is_unique
    assert no_diff_df.index.is_unique
    diff_df = (new_df - base_df) if diff_op == '-' else (new_df / base_df)
    return pd.concat([diff_df, no_diff_df.reindex(diff_df.index)], axis=1)


def select_visits(df, visits, no_diff_cols=NO_DIFF_COLS):
    if all([v in ['V1', 'V2', 'V3'] for v in visits]):
        df = df.loc[df.index.get_level_values('SAMPLE:VISIT').isin(visits)]
    else:
        assert len(visits) == 1
        visit, base_visit = visits[0].split('-')
        df = visit_diff(df, visit=visit, base_visit=base_visit, no_diff_cols=no_diff_cols)
    return df


def get_data_name_from_config(config):
    if 'celltype' in config:
        return config['name'].format(model=config['model'], celltype=config['celltype'], X=config['X_name'],
                                     X_visits='_'.join(config['X_visits']), Y=config['Y'].lstrip('^').rstrip('$'),
                                     Y_visits='_'.join(config['Y_visits']))
    else:
        return config['name'].format(model=config['model'], X=config['X_name'],
                                     X_visits='_'.join(config['X_visits']), Y=config['Y'].lstrip('^').rstrip('$'),
                                     Y_visits='_'.join(config['Y_visits']))


def get_filename_from_config(config, results_dir='', prefix='{}.'.format(PROJECT)):
    fn_prefix = '{prefix}{bin}{scale}{estimator}{features}{pathways}{DE_filter}{pca}{global_PCA}{global_KPCA}{global_UMAP}{poly}{select}{seed}{cv}{merge_folds}'.format(
        prefix=prefix,
        bin='bin{:g}_{:g}.'.format(config['binarize']['negative'][1]*100, config['binarize']['positive'][0]*100)
            if not config['binarize'].get('thresholds') and config['binarize'] and config['binarize'] != dict(negative=[0.0, 0.25], positive=[0.75, 1.0]) else
            'bin{:g}_{:g}.'.format(config['binarize']['negative'], config['binarize']['positive'])
            if config['binarize'].get('thresholds') else '',
        scale='scaled.' if config['scale'] else '',
        estimator=config['estimator'],
        features='.{}{}'.format(config['features']['type'], config['features']['value']) if config['features'] else '',
        pathways='.{}{}'.format(config['pathways']['type'],
                                '_{}'.format(
                                    config['pathways']['value']) if config['pathways']['value'] not in [
                                    config['pathways']['type'], 'all_pathways'] else '') \
            if config['pathways'] else '',
        DE_filter='.{}_{}'.format(config['DE_filter']['type'], config['DE_filter']['value']) if config[
            'DE_filter'] else '',
        pca='.pca{}'.format(config['pca']) if config['pca'] else '',
        global_PCA='.global_PCA{}'.format(config['global_PCA']) if 'global_PCA' in config and config['global_PCA'] else '',
        global_KPCA='.global_KPCA{}'.format(config['global_KPCA']) if 'global_KPCA' in config and config['global_KPCA'] else '',
        global_UMAP='.global_UMAP{}'.format(config['global_UMAP']) if 'global_UMAP' in config and config['global_UMAP'] else '',
        poly='.poly{}'.format(config['poly']) if 'poly' in config and config['poly'] else '',
        select='.{}{}'.format(config['select']['type'], config['select']['value']) if config['select'] else '',
        seed='.seed{}'.format(config['seed']),
        cv='' if config['test_splits'] == 10 and config['cv_splits'] == 10 else '.test{}cv{}'.format(config['test_splits'], config['cv_splits']),
        merge_folds='.merge_folds' if 'merge_folds' in config and config['merge_folds'] else ''
    )
    return os.path.join(results_dir, '{}.{}.{}'.format(fn_prefix, '{data}', '{ext}'))


def get_model_for_visits(model_name, visits):
    if ':' not in model_name:
        model = model_name
    else:
        model_name_split = model_name.split(':')
        assert len(model_name_split) == 2
        if len(visits) == 1 and '-' in visits[0]:
            model = model_name_split[1]
            assert model == 'raw'
        else:
            model = model_name_split[0]
            assert model in ['combat_batch', 'correct_batch']
    return model


def get_numeric_targets(df, regex, specials, include_binary=True, verbose=True):
    Y_df = df.loc[:, df.columns.str.contains(regex)].copy()
    for label in specials:
        if label in Y_df.columns:
            Y_df[label] = specials[label](Y_df[label])
    return force_df_to_numeric(Y_df, include_binary=include_binary, verbose=verbose)


def get_basic_binary_specials():
    specials = {
        'DONOR:IC_TIME_8-9_11-19': lambda s: binarize_categorical(
            s, binary_map={1: ['MOR'], 0: ['ARVO']}),
        'DONOR:IC_TIME_8-9_18-19': lambda s: binarize_categorical(
            s, binary_map={1: ['MOR'], 0: ['EVE']}),
        'DONOR:IC_TIME_CAT': lambda s: binarize_categorical(
            s, binary_map={1: ['MOR8_9'], 0: ['MOR11_13', 'EVE']}),
        'DONOR:SEX': lambda s: binarize_categorical(
            s, binary_map={0: ['F'], 1: ['M']}),
        'SAMPLE:VISIT': lambda s: binarize_categorical(
            s, binary_map={0: ['V1'], 1: ['V2', 'V3']}),
        'SAMPLE:TISSUE': lambda s: binarize_categorical(
            s, binary_map={0: ['PBMC'], 1: ['nkcell', 'monocyte', 'cd8t']})
    }

    for score in ['thm.adaptive_MTB_7d_V3_FC1.2_responder', 'thm.heterologous_nonspecific_7d_V3_FC1.2_responder',
        'thm.innate_nonspecific_24h_wo_LAC_IL10_IL1ra_V3_FC1.2_responder']:
        specials[score] = lambda s: binarize_categorical(s, binary_map={0: ['N'], 1: ['R']})

    return specials


def get_numeric_targets_for_config(config, results_dir, prefix=PROJECT):
    model = get_model_for_visits(config['model'], config['Y_visits'])
    data_fn = os.path.join(results_dir, model, '{}.{}.hdf'.format(prefix, model))
    sample_annot_df = pd.read_hdf(data_fn, key='{}/sample_annot'.format(config['celltype']))
    Y_df = get_numeric_targets(sample_annot_df,
                               regex=config['Y'],
                               specials=get_basic_binary_specials(),
                               verbose=False)
    return Y_df.columns.values

    if config['binarize'] is not None:
        _flu_season = lambda x: 0 if 6 <= x < 10 else (1 if x < 4 or x >= 11 else np.nan)
        _specials['SAMPLE:VISIT_MONTH_REAL'] = lambda s: s.apply(_flu_season)
        _specials['DONOR:IC_MONTH_REAL'] = lambda s: s.apply(_flu_season)


def binarize_categorical(s, binary_map):
    assert isinstance(s, pd.Series)
    s = s.copy()
    _binarized = []
    for binary_value in binary_map:
        s[s.isin(binary_map[binary_value])] = binary_value
        _binarized.extend(binary_map[binary_value])
    s[~s.isin(binary_map.keys())] = np.nan
    return s


def evaluate_target(target, config, results_dir, alt_scorer={'neg_mean_squared_error': 'r2'}):
    fn = get_filename_from_config(config, results_dir)
    preds_fn = fn.format(data='{}.predictions'.format(target), ext='npz')
    grids_fn = fn.format(data='{}.models'.format(target), ext='pickle.gz')
    if os.path.exists(preds_fn) and os.path.exists(preds_fn):
        with gzip.open(grids_fn) as f:
            grids = joblib.load(f)

        predictions = np.load(preds_fn, allow_pickle=True)
        y_true, y_pred = predictions['y_true'], predictions['y_pred']
        train_scores, test_scores = predictions['train_scores'], predictions['test_scores']

        # TEST THIS
        for k in range(len(y_pred)):
            assert len(y_pred[k]) == len(y_true[k])

        samples = predictions['samples']
        non_imputed_samples = predictions['non_imputed_samples']
        if non_imputed_samples:
            _y_true, _y_pred = [], []
            for k in range(len(samples)):
                mask = np.isin(non_imputed_samples[k], samples[k])
                print('Removing {} imputed samples'.format(mask.shape[0] - mask.sum()))
                _y_true.append(y_true[k][mask])
                _y_pred.append(y_pred[k][mask])
            _y_true = np.asarray(_y_true)
            _y_pred = np.asarray(_y_pred)

            # if everything works, only these two lines are necessary:
            y_true = np.asarray([y_true[k][np.isin(non_imputed_samples[k], samples[k])] for k in range(len(grids))])
            y_pred = np.asarray([y_pred[k][np.isin(non_imputed_samples[k], samples[k])] for k in range(len(grids))])

            assert np.array_equal(_y_true, y_true)
            assert np.array_equal(_y_pred, y_pred)
            for k in range(len(y_pred)):
                assert len(y_pred[k]) == len(y_true[k])
                print('len(y_true)', len(y_true), 'without imputed')
        ##

        if config['binarize'] and not config['binarize'].get('thresholds'):
            _thresholds = get_binarize_thresholds(pd.Series(np.concatenate(y_true)),
                                                  negative=config['binarize']['negative'],
                                                  positive=config['binarize']['positive']).drop([0, 1]).values
        else:
            assert not config['scoring'].startswith('quantile_')
            _thresholds = None

        score = get_scorer_func(
            get_scorer(config['scoring']) if config['scoring'] not in ['quantile_roc_auc', 'quantile_average_precision', 'partial_roc_auc'] else \
                make_scorer(score_func=roc_auc_score25, greater_is_better=True, needs_threshold=True) if config['scoring'] == 'partial_roc_auc' else \
                BinarizedRegressionScorer(average_precision_score if config['scoring'] == 'quantile_average_precision' else roc_auc_score, sign=1, binarized=_thresholds),
            binarized=_thresholds
        )

        _alt_scoring = alt_scorer.get(config['scoring'], config['scoring'])
        alt_score = get_scorer_func(
            get_scorer(_alt_scoring) if _alt_scoring not in ['quantile_roc_auc', 'quantile_average_precision', 'partial_roc_auc'] else \
                make_scorer(score_func=roc_auc_score25, greater_is_better=True, needs_threshold=True) if _alt_scoring == 'partial_roc_auc' else \
                BinarizedRegressionScorer(average_precision_score if _alt_scoring == 'quantile_average_precision' else roc_auc_score, sign=1, binarized=_thresholds),
            binarized=_thresholds
        )
        params = {p[len(EST_STEP) + 2:]: [grids[k].best_params_[p] for k in range(len(grids))] for p in
                  grids[0].best_params_}
        _best_model = grids[0].best_estimator_.named_steps.estimator if isinstance(grids[0].best_estimator_, Pipeline) else grids[0].best_estimator_
        if hasattr(_best_model, 'coef_'):
            coefs = [grids[k].best_estimator_.named_steps.estimator.coef_ for k in range(len(grids))]
        else:
            coefs = None
        _classifier = is_classifier(_best_model)
        cv_scores = np.asarray([grids[k].best_score_ for k in range(len(grids))])

        if config['test_splits'] != -1:
            _test_scores = np.asarray([score(y_true[k], y_pred[k]) for k in range(len(grids))])
            if non_imputed_samples:
                print('{} {:.3f} | {} without imputed {:.3f}'.format(config['scoring'], np.mean(test_scores),
                                                                     config['scoring'], np.mean(_test_scores)))
            else:
                np.testing.assert_array_equal(_test_scores, test_scores)
            test_scores = _test_scores
            alt_test_scores = np.asarray([alt_score(y_true[k], y_pred[k]) for k in range(len(grids))])

        y_true, y_pred = np.concatenate(y_true), np.concatenate(y_pred)
        union_test_score = score(y_true, y_pred)
        alt_union_test_score = alt_score(y_true, y_pred)

        if config['test_splits'] == -1:
            test_scores = [union_test_score]
            alt_test_scores = [alt_union_test_score]

        results = {}
        results['model'] = config['model']
        results['merge_folds'] = config.get('merge_folds', False)
        results['poly'] = config.get('poly', None)
        results['seed'] = config['seed']
        results['downsampling'] = config['downsampling']
        if config['binarize']:
            if config['binarize'].get('thresholds'):
                results['binarize'] = 'FC:{}_{}'.format(np.power(2, config['binarize']['negative']), np.power(2, config['binarize']['positive']))
            else:
                lower_n, upper_n = config['binarize']['negative']
                lower_p, upper_p = config['binarize']['positive']
                results['binarize'] = '{}-{}_{}-{}'.format(lower_n, upper_n, lower_p, upper_p)
        else:
            results['binarize'] = np.nan
        results['scale'] = config['scale']
        results['pca'] = np.nan if config['pca'] is None else config['pca']
        results['global_PCA'] = np.nan if 'global_PCA' not in config or config['global_PCA'] is None else config['global_PCA']
        results['global_KPCA'] = np.nan if 'global_KPCA' not in config or config['global_KPCA'] is None else config['global_KPCA']
        results['global_UMAP'] = np.nan if 'global_UMAP' not in config or config['global_UMAP'] is None else config['global_UMAP']
        results['select'] = np.nan if config['select'] is None else '{}_{}'.format(config['select']['type'],
                                                                                   config['select']['value'])
        results['features'] = np.nan if config['features'] is None else '{}_{}'.format(config['features']['type'],
                                                                                       config['features']['value'])
        results['DE_filter'] = np.nan if config['DE_filter'] is None else '{}_{}'.format(config['DE_filter']['type'],
                                                                                         config['DE_filter']['value'])
        results['X_name'] = config['X_name']
        results['X'] = config['X']
        results['y'] = target
        results['X_visits'] = '_'.join(config['X_visits'])
        results['y_visits'] = '_'.join(config['Y_visits'])
        results['estimator'] = config['estimator']
        results['scoring'] = config['scoring']
        results['alt_scoring'] = _alt_scoring
        results['params'] = params
        results['non_zero_coefs'] = [(_coefs != 0).sum() for _coefs in coefs] if coefs else np.nan
        results['coefs'] = [_coefs.flatten().shape[0] for _coefs in coefs] if coefs else np.nan
        results['test_size'] = y_true.shape[0]
        results['class_ratio'] = ((y_true == 1).sum() / y_true.shape[0]) if _classifier else np.nan
        results['mean_cv_score'], results['std_cv_score'] = np.mean(cv_scores), np.std(cv_scores)
        results['mean_test_score'], results['std_test_score'] = np.mean(test_scores), np.std(test_scores)
        results['alt_mean_test_score'], results['alt_std_test_score'] = np.mean(alt_test_scores), np.std(alt_test_scores)
        results['mean_train_score'], results['std_train_score'] = np.mean(train_scores), np.std(train_scores)
        results['union_test_score'] = union_test_score
        results['alt_union_test_score'] = alt_union_test_score
        results['FPR'], results['TPR'], _ = roc_curve(y_true, y_pred) if _classifier else (np.nan, np.nan, np.nan)
        results['precision'], results['recall'], _ = precision_recall_curve(y_true, y_pred) if _classifier else (np.nan, np.nan, np.nan)
        results['preds_fn'] = preds_fn
        results['grids_fn'] = grids_fn
    else:
        results = None
    return results


def deprecated_correct_cytokines(cyto_df, covariates, do_not_correct=None, interactions=None, verbose=True):
    print('see bcg _ utils . fit _ linear _ model instead')
    raise ValueError
    # This does not work because if do_not_correct contains SAMPLE:VISIT,
    # it will also NOT correct SAMPLE:VISIT_TIME_REAL because of the shared prefix
    covariates = np.asarray(covariates)
    # some covariates might be a numpy function of the index level, e.g. np.sin() or np.cos()
    function_covariates = np.asarray([c.count('__') for c in covariates]) == 1
    design_df = pd.get_dummies(cyto_df.index.to_frame().loc[:, covariates[~function_covariates]], drop_first=True)
    for cov in covariates[function_covariates]:
        function, covariate = cov.split('__')
        design_df[cov] = getattr(np, function)(cyto_df.index.get_level_values(covariate))

    if interactions:
        design_cols_freeze = design_df.columns.copy()
        for cov1, cov2 in interactions:
            for _cov1 in design_cols_freeze[design_cols_freeze.str.contains(cov1)]:
                for _cov2 in design_cols_freeze[design_cols_freeze.str.contains(cov2)]:
                    raise ValueError(
                        'The problem is with the kind of interactions: IC_DATE/IC_TIME but then it needs to be removed and VISIT_DATE/VISIT_TIME inserted for the non-interaction effect')
                    design_df['{}:{}'.format(_cov1, _cov2)] = design_df[_cov1] * design_df[_cov2]

    if verbose:
        print('Model matrix:', design_df.columns.values)

    corrected_df = pd.DataFrame(columns=cyto_df.index)
    coefs_df = pd.DataFrame()
    for cyto in cyto_df.columns:
        lm = LinearRegression(fit_intercept=True)
        not_nan = ~cyto_df[cyto].isnull()
        X = design_df.loc[not_nan]
        y = cyto_df.loc[not_nan][cyto]
        assert X.index.equals(y.index)
        lm.fit(X, y)
        coefs_df = coefs_df.append(pd.Series(lm.coef_, index=design_df.columns, name=cyto))
        if do_not_correct:
            cov_mask = np.asarray([not any([col.startswith(n) for n in do_not_correct]) for col in design_df.columns])
        else:
            cov_mask = np.ones_like(X.columns, dtype=bool)
        if verbose:
            print('Correcting:', X.columns[cov_mask].values)
        corrected_y = y - np.dot(X.loc[:, cov_mask], lm.coef_[cov_mask])
        corrected_df = corrected_df.append(corrected_y)
    return corrected_df.T, design_df, coefs_df


def barplot(df, score_name, ylabel=None, title=None, order=None, fig_fn=None):
    ax = sns.barplot(data=df, x='comparison', y=score_name, hue='y', order=order)
    ax.legend(bbox_to_anchor=(1, 1))
    ax.axhline(0.5, c='0.15')
    ax.set_xticklabels(ax.get_xticklabels(), rotation=90)
    ax.set_xlabel('$X${}$y$ visits'.format(ARROW))
    if ylabel:
        ax.set_ylabel(ylabel)
    ax.set_ylim((0.4, 0.8))
    ax.set_title(title)
    sns.despine()
    if fig_fn:
        savefig(fig_fn)
    return ax


def correct_date(date, century, sep='.'):
    if not pd.isnull(date):
        day, month, year = date.split(sep)
        date = sep.join([day, month, str(int(year) + century)])
    return date


def rename_feature_selection(df, visit=None, base_visit=None):
    df = df.copy()
    df.loc[df['features'] == 'characterization_TSS', 'features'] = '_TSS'
    df.loc[df['select'] == 'mutual_info_classif_1000', 'select'] = '_fsMI'
    if visit and base_visit:
        df.loc[df['DE_filter'] == 'pval_{}_{}_1000'.format(visit, base_visit), 'DE_filter'] = '_pval'
        df.loc[df['DE_filter'] == 'max_rank_{}_{}_1000'.format(visit, base_visit), 'DE_filter'] = '_mrank'
    return df


def add_month_real(df, label='SAMPLE:VISIT_MONTH_REAL', date_label='SAMPLE:VISIT_DATE_REAL'):
    df = df.copy()
    df[label] = df[date_label] * 12 + 1
    df.loc[df[label] >= 13, label] -= 12
    assert not ((df[label] < 1).any() or (df[label] >= 13).any())
    return df


def load_atac_V1(fn, atac=True, sample_annot=True, peak_annot=False):
    assert np.sum([atac, sample_annot, peak_annot]) != 0
    _res = []
    if atac:
        atac_df = pd.read_hdf(fn, key='atac')
        atac_df = atac_df.loc[atac_df.index.get_level_values('SAMPLE:VISIT') == 'V1']
        atac_df.index = atac_df.index.get_level_values('SAMPLE:DONOR')
        assert atac_df.index.is_unique
        _res.append(atac_df)
    if sample_annot:
        sample_annot_df = pd.read_hdf(fn, key='sample_annot')
        sample_annot_df = sample_annot_df.loc[sample_annot_df['SAMPLE:VISIT'] == 'V1']
        sample_annot_df.index = sample_annot_df['SAMPLE:DONOR'].values
        assert sample_annot_df.index.is_unique
        assert not atac or sample_annot_df.index.equals(atac_df.index)
        _res.append(sample_annot_df)
    if peak_annot:
        peak_annot_df = pd.read_hdf(fn, key='peak_annot')
        assert not atac or peak_annot_df.index.equals(atac_df.columns)
        _res.append(peak_annot_df)
    return tuple(_res) if len(_res) > 1 else _res[0]


def correct_with_limma(counts_df, celltype, model, coefs='auto', do_not_correct=None, remove_intercept=False,
                       verbose=True):
    if counts_df is None:
        counts_df = pd.read_csv(de_fn(celltype=celltype, model=model, data='counts'), index_col=0)
    design_df = pd.read_csv(de_fn(celltype=celltype, model=model, data='design'), index_col=0)
    design_df = design_df.rename(FIX_LIMMA_COEFS, axis=1)
    assert design_df.index.is_unique and design_df.columns.is_unique
    assert design_df.index.equals(counts_df.columns)
    assert do_not_correct is None or coefs == 'auto'

    if coefs == 'auto':
        if do_not_correct is None:
            do_not_correct = ['V2', 'V3']
        coefs = set(design_df.columns).difference(['X__Intercept'] + do_not_correct)

    coef_df = pd.read_csv(de_fn(celltype=celltype, model=model, data='coefficients'), index_col=0)
    coef_df = coef_df.rename(FIX_LIMMA_COEFS, axis=1)
    assert coef_df.columns.equals(design_df.columns)
    assert coef_df.index.is_unique and coef_df.columns.is_unique
    assert coef_df.index.equals(counts_df.index)

    if remove_intercept:
        # cov_mask = design_df.columns == 'X.Intercept.'
        cov_mask = design_df.columns == 'X__Intercept'
    else:
        cov_mask = np.zeros_like(design_df.columns, dtype=bool)
    for c in coefs:
        cov_mask |= design_df.columns.str.startswith(c)

    confounding_effects_df = coef_df.loc[:, cov_mask].dot(design_df.loc[:, cov_mask].T)
    assert confounding_effects_df.index.equals(counts_df.index)
    assert confounding_effects_df.columns.equals(counts_df.columns)

    if verbose:
        print('Removing effects of:', design_df.columns[cov_mask].values)
        print('Not removing:', design_df.columns[~cov_mask].values)
    return counts_df - confounding_effects_df


def get_corrected_counts(celltype, model, do_not_correct=['V2', 'V3'], remove_singletons=False, plot=True,
                         figsize=(5, 4.5), fig_fn=None):
    counts_df = pd.read_csv(de_fn(celltype=celltype, model=model, data='counts'), index_col=0)
    if plot:
        fig, axs = plt.subplots(1, 2, figsize=(2 * figsize[0], figsize[1]))
        fig.subplots_adjust(wspace=0.4)
        ax = plot_sequenced_samples(counts_df, n_samples=min(counts_df.shape[1], 100), samples_as_rows=False, ax=axs[0])
        ax.set_title('Normalized'.format(celltype, model))

    counts_df = correct_with_limma(counts_df=counts_df, celltype=celltype, model=model, do_not_correct=do_not_correct)
    if plot:
        ax = plot_sequenced_samples(counts_df, n_samples=min(counts_df.shape[1], 100), samples_as_rows=False, ax=axs[1])
        ax.set_title('Normalized and corrected'.format(celltype, model))
        if fig_fn:
            savefig(fig_fn)

    counts_df = counts_df.T
    if remove_singletons:
        counts_df.index = counts_df.index.str.split('_', expand=True)
        counts_df.index.names = ['donor', 'visit', 'celltype']
        donor_counts = counts_df.groupby('donor').count().iloc[:, 0]
        donors = donor_counts.loc[donor_counts > 1].index.values
        counts_df = counts_df.loc[counts_df.index.get_level_values('donor').isin(donors)]
        counts_df.index = ['_'.join(ix) for ix in counts_df.index]

    return (counts_df, axs) if plot else counts_df


def _fmt_ext(ext):
    return '{}{}'.format('.' if len(ext) != 0 and ext[0] != '.' else '', ext)


def de_fn(celltype, model, data='results_p5', extras=None, ext='csv', gzipped=True,
          subdir=None, results_dir='results', legacy=False):
    fn = os.path.join(results_dir, 'DE', '{celltype}.{model}' if not legacy else '', '{subdir}',
                      'de_{data}_{celltype}.{model}{extras}{ext}{gz}')
    return fn.format(celltype=celltype, model=model, data=data, extras='{}'.format(extras) if extras else '',
                     ext=_fmt_ext(ext), gz='.gz' if gzipped else '', subdir=subdir if subdir else '')


def lr_fn(celltype, model, target, data='results', extras=None, ext='csv', gzipped=False,
          subdir=None, results_dir='results'):
    fn = os.path.join(results_dir, 'LR', '{celltype}.{model}', '{subdir}',
                      'LR_{data}_{target}{extras}{ext}{gz}')
    return fn.format(celltype=celltype, model=model, target=target, data=data, extras='.{}'.format(extras) if extras else '',
                     ext=_fmt_ext(ext), gz='.gz' if gzipped else '', subdir=subdir if subdir else '')


def lr_fn_suffix(visits=None, visit_interaction=None, visit_interaction_and_correct_evening=None,
                 visit_interaction_and_correct_season=None, fold_changes=None, LMM=False, scale=False,
                 include_season=False, include_cyto=False, include_CM=False, include_blood_perc=False, include_blood_per_ml=False, include_scores=False,
                 remove_evening=False, correct_evening=False, correct_season=False):
    return '{visits}{visit_interaction}{steady_state}{fold_changes}{LMM}{scale}{remove_evening}{correct_evening}{correct_season}{include_season}{include_cyto}{include_CM}{include_blood_perc}{include_blood_per_ml}{include_scores}'.format(
        visits='{}.'.format('_'.join(sorted(visits))) if visits else '',
        visit_interaction='visit_interaction' if visit_interaction or visit_interaction_and_correct_evening or visit_interaction_and_correct_season else '',
        steady_state='steady_state' if not visit_interaction and not visit_interaction_and_correct_evening and not visit_interaction_and_correct_season and not fold_changes else '',
        fold_changes='fold_changes_{}'.format(fold_changes) if fold_changes else '',
        LMM='.LMM' if LMM else '',
        scale='.scaled' if scale else '',
        remove_evening='.remove_evening' if remove_evening else '',
        correct_evening='.correct_evening' if correct_evening or visit_interaction_and_correct_evening else '',
        correct_season='.correct_season' if correct_season or visit_interaction_and_correct_season else '',
        include_season='.DONOR:IC_DATE_2PI_SIN__DONOR:IC_DATE_2PI_COS' if include_season else '',
        include_cyto='.include_cyto' if include_cyto else '',
        include_CM='.include_CM' if include_CM else '',
        include_blood_perc='.include_blood_perc' if include_blood_perc else '',
        include_blood_per_ml='.include_blood_per_ml' if include_blood_per_ml else '',
        include_scores='.include_scores' if include_scores else '',
    )


def summary_lr_fn(celltype, model, Y_name, data='results', visits=None, visit_interaction=None, visit_interaction_and_correct_evening=None, visit_interaction_and_correct_season=None, fold_changes=None, LMM=False,
                  scale=False, include_season=False, include_cyto=False, include_CM=False, include_blood_perc=False, include_blood_per_ml=False, include_scores=False,
                  remove_evening=False, correct_evening=False, correct_season=False, extras=None, results_dir='results'):
    suffix = lr_fn_suffix(visits=visits, visit_interaction=visit_interaction,
                          visit_interaction_and_correct_evening=visit_interaction_and_correct_evening,
                          visit_interaction_and_correct_season=visit_interaction_and_correct_season,
                          fold_changes=fold_changes, LMM=LMM, scale=scale, remove_evening=remove_evening,
                          include_season=include_season, include_cyto=include_cyto, include_CM=include_CM, include_blood_perc=include_blood_perc, include_blood_per_ml=include_blood_per_ml, include_scores=include_scores,
                          correct_evening=correct_evening, correct_season=correct_season)
    return lr_fn(celltype=celltype, model=model, target='{}.{}'.format(Y_name, suffix),
                 data=data, extras=extras, results_dir=results_dir)


def _enrichment_fn(method, celltype, model, data, coef, effect_size_filter, rank_metric, top_n, direction, db, regions,
                   swap_to_non_resp=False, random_seed=None, ext=None, gzipped=False, subdir=None, results_dir='results'):
    fn = os.path.join(results_dir, 'DE', '{celltype}.{model}', '{subdir}',
            '{method}_{data}{coef}{effect_size_filter}{rank_metric}{top_n}{direction}{db}{regions}{random}{ext}{gz}')
    if swap_to_non_resp:
        if coef.endswith('_V3_FC1.2_R'):
            direction *= -1
        elif coef.endswith('_R'):
            raise ValueError('Unexpected: cannot determine if to switch responder phenotypeor not')

    return fn.format(method=method, data=data, celltype=celltype, model=model,
                     coef='.{}'.format(coef) if coef else '',
                     top_n='.{}{}'.format('top_' if top_n >= 1 else 'FDR_', top_n) if top_n else '',
                     effect_size_filter='.ESF_{}'.format(effect_size_filter) if effect_size_filter else '',
                     rank_metric='.{}'.format(rank_metric) if rank_metric else '',
                     direction='.{}'.format(
                         'down' if direction == -1 else 'up' if direction == 1 else direction) if direction else '',
                     db='.{}'.format(db) if db else '',
                     regions='.{}'.format(regions) if regions else '',
                     random='.random{}'.format(random_seed) if random_seed is not None else '',
                     ext=_fmt_ext(ext) if ext else '', gz='.gz' if gzipped else '', subdir=subdir if subdir else '')


def homer_fn(celltype, model, data, coef, effect_size_filter, rank_metric, top_n, direction, regions=None,
             swap_to_non_resp=False, random_seed=None, ext='auto', gzipped=False, subdir='Homer', results_dir='results'):
    assert data in ['{data}', 'results', 'top_regions', 'universe']
    return _enrichment_fn(method='homer', data=data, celltype=celltype, model=model, coef=coef, top_n=top_n,
                          rank_metric=rank_metric, effect_size_filter=effect_size_filter, direction=direction,
                          db=None, regions=regions, swap_to_non_resp=swap_to_non_resp, random_seed=random_seed,
                          gzipped=gzipped, subdir=subdir,
                          ext=('bed' if data in ['top_regions', 'universe'] else 'tsv') if ext == 'auto' else ext,
                          results_dir=results_dir)


def lola_fn(celltype, model, data, coef, direction, rank_metric, top_n, effect_size_filter=None, regions=None, db=None,
            swap_to_non_resp=False, random_seed=None, ext='auto', gzipped=False, subdir='LOLA', results_dir='results'):
    assert data in ['{data}', 'results', 'top_regions', 'universe'] or data.startswith('barplot')
    return _enrichment_fn(method='lola', data=data, celltype=celltype, model=model, coef=coef, top_n=top_n,
                          rank_metric=rank_metric, effect_size_filter=effect_size_filter, direction=direction,
                          db=db, regions=regions, swap_to_non_resp=swap_to_non_resp,
                          random_seed=random_seed, gzipped=gzipped, subdir=subdir,
                          ext=('bed' if data in ['top_regions', 'universe'] else 'pdf' if data.startswith(
                              'barplot') else 'tsv') if ext == 'auto' else ext, results_dir=results_dir)


def enrichr_fn(celltype, model, data, coef, effect_size_filter, rank_metric, top_n, direction, db, regions, ext='auto',
               swap_to_non_resp=False, gzipped=False, subdir='Enrichr', results_dir='results'):
    assert data in ['{data}', 'results', 'top_genes', 'top_regions', 'true_top_regions'] or data.startswith('barplot')
    return _enrichment_fn(method='enrichr', data=data, celltype=celltype, model=model, coef=coef, top_n=top_n,
                          rank_metric=rank_metric,
                          effect_size_filter=effect_size_filter, direction=direction, db=db, regions=regions,
                          ext=('txt' if data in ['top_genes', 'top_regions'] else 'pdf' if data.startswith(
                              'barplot') else 'tsv') if ext == 'auto' else ext,
                          swap_to_non_resp=swap_to_non_resp, gzipped=gzipped, subdir=subdir, results_dir=results_dir)


def gsea_fn(celltype, model, data, coef, effect_size_filter, rank_metric, db, regions, ext='auto',
            swap_to_non_resp=False, gzipped=False, subdir='GSEA', results_dir='results', method='gsea'):
    assert method in ['gsea', 'gseapy']
    assert data in ['{data}', 'results', 'ranked_genes', 'ranked_regions'] or data.startswith('barplot')
    return _enrichment_fn(method=method, data=data, celltype=celltype, model=model, coef=coef, top_n=None,
                          rank_metric=rank_metric,
                          effect_size_filter=effect_size_filter, direction=None, db=db, regions=regions,
                          ext=('rnk' if data in ['ranked_genes', 'ranked_regions'] else 'pdf' if data.startswith(
                              'barplot') else 'tsv') if ext == 'auto' else ext,
                          swap_to_non_resp=swap_to_non_resp, gzipped=gzipped, subdir=subdir, results_dir=results_dir)


def make_test_samples_de_config(celltype, model, remove_samples, donor_score_col, suffix, annot_fn):
    _annot_df = get_sample_annot()
    assert (donor_score_col if not suffix else '{}{}'.format(donor_score_col, suffix)) in _annot_df.columns, donor_score_col

    config = {}
    config['results_dir'] = ML_RESULTS_ROOT
    config['debug'] = False

    config['celltype'] = celltype
    config['model'] = model
    config['batch_corrected'] = False
    config['dream'] = False

    config['annot_fn'] = annot_fn
    config['peak_fn'] = '../data/DE/peaks_filtered_PBMC.csv.gz'
    config['counts_fn'] = '../data/DE/quantification_filtered_{celltype}.csv.gz'.format(celltype=config['celltype'])

    if suffix == 'score':
        _score = safe_R_name(donor_score_col)
        _interaction = '{score}*SAMPLE.VISIT'.format(score=_score)
        _contrasts = [[_score, 'V2', 'V3', '{}.V2'.format(_score), '{}.V3'.format(_score)]]
    else:
        _score = '{}{}'.format(safe_R_name(donor_score_col), safe_R_name(suffix))
        _interaction = '{score} + {score}:SAMPLE.VISIT'.format(score=_score)
        _contrasts = [['{}{}N'.format(safe_R_name(donor_score_col), safe_R_name(suffix.replace('_responder', '_'))),
                       '{}{}N.V2'.format(safe_R_name(donor_score_col), safe_R_name(suffix.replace('_responder', '_'))),
                       '{}{}N.V3'.format(safe_R_name(donor_score_col), safe_R_name(suffix.replace('_responder', '_'))),
                       '{}{}R.V2'.format(safe_R_name(donor_score_col), safe_R_name(suffix.replace('_responder', '_'))),
                       '{}{}R.V3'.format(safe_R_name(donor_score_col), safe_R_name(suffix.replace('_responder', '_')))
                       ]]

    _blood = safe_R_name(BLOOD)
    config['columns'] = ['SAMPLE.DONOR', 'SAMPLE.VISIT', 'LAB.BATCH', 'DONOR.SEX', 'DONOR.AGE',
                         'RUN.TSS_ENRICHMENT', 'SAMPLE.VISIT_TIME_REAL'] + \
    (_blood if config['celltype'] == 'PBMC' else []) + \
    [_score]

    config['design'] = '~ {interaction} + {covariates}'.format(
        interaction=_interaction,
        covariates=' + '.join(
            [c for c in config['columns'] if c not in [_score, 'SAMPLE.VISIT', 'SAMPLE.DONOR']]))

    config['contrasts'] = _contrasts
    assert is_iterable(config['contrasts'][0])

    # e.g., "DONOR.IC_DATE_REAL" # this will create "SPLINES" variables - use it in the design
    config['splines'] = None
    config['splines_df'] = None

    config['block_on_donor'] = True
    config['remove_samples'] = remove_samples
    config['remove_wrong_batch'] = True
    config['remove_exclusions'] = True
    config['remove_300BCG315'] = False
    config['remove_evening'] = False
    config['complete_design'] = False
    config['useful_samples'] = False
    config['save_non_log_CPM'] = False
    config['do_not_correct'] = ['V2', 'V3']

    config['volcano_fdr'] = [0.05, 0.1, 0.2, 0.25]

    make_dir(ML_RESULTS_ROOT, 'DE', '{}.{}'.format(config['celltype'], config['model']))
    with open(de_fn(config['celltype'], config['model'], data='config', ext='yml', gzipped=False, results_dir=ML_RESULTS_ROOT), 'w') as f:
        yaml.dump(config, f, default_flow_style=False, sort_keys=False)


def get_gsea_ranked_genes(library, celltype, model, regions, coef, n_perms=5000, results_dir='results'):
    print('WARNING: this assumes GSEA was run on genes (not on regions)')
    ranked_genes_df = pd.read_csv(
        gsea_fn(library=library, celltype=celltype, model=model, regions=regions, coef=coef, n_perms=n_perms,
                ranked_genes=True, results_dir=results_dir),
        index_col=0, sep='\t')
    ranked_genes_df = ranked_genes_df.rename({'genes': 'region'}, axis=1)
    return ranked_genes_df


def get_ledge_regions(ledge_genes, ranked_genes_df):
    print('WARNING: this assumes GSEA was run on genes (not on regions)')
    ledge_regions = []
    for genes in ledge_genes:
        if isinstance(genes, str):
            genes = eval(genes)
        ledge_regions.append(ranked_genes_df.loc[genes, 'region'].tolist())
    return ledge_regions


def signed_log10_pvals(de_df, coef=None, neg_log10=True):
    s = de_df[pval_col(coef)].copy()
    if neg_log10:
        s = -np.log10(s)
        s = s.where(s >= 0, other=0)
    s.loc[de_df[coef_col(coef)] < 0] *= -1
    return s.rename('signed.log10.p.value.{}'.format(coef))


def adjusted_pvals(de_df, coef=None, method='fdr_bh'):
    assert isinstance(de_df, (pd.DataFrame, pd.Series))
    if isinstance(de_df, pd.DataFrame):
        assert coef is not None
        pvals = de_df[pval_col(coef)]
    elif isinstance(de_df, pd.Series):
        assert coef is None
        pvals = de_df
    pvals = pvals.loc[~pvals.isnull()]
    if pvals.shape[0] != 0:
        padj = multipletests(pvals, method=method)[1]
    else:
        padj = []
    return pd.Series(padj, index=pvals.index, name=padj_col(coef) if coef else 'padj')


def read_signed_log10_pvals(fn, coef):
    de_df = pd.read_csv(fn, usecols=[pval_col(coef), coef_col(coef)])
    return signed_log10_pvals(de_df, coef)


def read_adjusted_pvals(fn, coef, method='fdr_bh'):
    de_df = pd.read_csv(fn, usecols=[pval_col(coef)])
    return adjusted_pvals(de_df, coef, method)


def get_peak_annot(fn=None):
    if fn is None:
        fn = PEAK_ANNOT_ALL_FN
    print('Using this peak annotation:', fn)
    peaks_info_df = pd.read_csv(fn, index_col=0)
    assert peaks_info_df.index.is_unique
    return peaks_info_df


def simplify_peak_annot(df, location_col=GENOMIC_LOCATION_COL, feature_col=FEATURE_TYPE_COL, regulatory_col=REGULATORY_ANNOT_COL, allow_nulls=False):

    if location_col is not None:
        assert location_col not in df.columns
        df[location_col] = simplify_genomic_locations(df, tss_proximal_label=TSS_PROXIMAL_LABEL, gene_body_label=GENE_BODY_LABEL, distal_label=DISTAL_LABEL, intergenic_label=INTERGENIC_LABEL)
        assert allow_nulls or not df[location_col].isnull().any(), df[location_col].isnull().sum()
        assert df.loc[~df[location_col].isnull(), location_col].isin([TSS_PROXIMAL_LABEL, GENE_BODY_LABEL, DISTAL_LABEL, INTERGENIC_LABEL]).all(), set(df[location_col])

    if feature_col is not None:
        assert feature_col not in df.columns
        df[feature_col] = np.nan
        df.loc[:, feature_col] = df.loc[:, 'feat_type'].str.replace('^transcript:', '').str.replace('^gene:', '')
        df.loc[df[feature_col].isin(['IG_C_gene', 'IG_D_gene', 'IG_J_gene', 'IG_V_gene', 'TR_C_gene', 'TR_D_gene', 'TR_J_gene', 'TR_V_gene']), feature_col] = IGTR
        df.loc[df[feature_col] == 'protein_coding', feature_col] = PROTEIN_CODING
        df.loc[df[feature_col] == 'lncRNA', feature_col] = LNCRNA
        assert allow_nulls or not df[feature_col].isnull().any(), df[feature_col].isnull().sum()
        assert df.loc[~df[feature_col].isnull(), feature_col].isin([IGTR, PROTEIN_CODING, LNCRNA]).all(), set(df[feature_col])

    if regulatory_col is not None:
        df[regulatory_col] = df['reg_feature'].map(RENAME_REG_BUILD)
        # assert regulatory_col not in df.columns
        # df[regulatory_col] = np.nan
        # df.loc[:, regulatory_col] = df.loc[:, 'reg_feature'].str.replace('_region$', '').str.replace(
        #     '_site$', '').str.replace('^reg_NONE$', 'unassigned').str.replace('^p', 'P').str.replace('^e', 'E').str.replace(
        #     '^o', 'O').str.replace('^u', 'U').str.replace('_flanking', '-flanking').str.replace('_', ' ')
        # ['Promoter', 'Promoter-flanking', 'Enhancer', 'TF binding', 'CTCF binding', 'Open chromatin', 'Unassigned']
        assert allow_nulls or not df[regulatory_col].isnull().any(), df[regulatory_col].isnull().sum()
        assert df.loc[~df[regulatory_col].isnull(), regulatory_col].isin(RENAME_REG_BUILD.values()).all(), set(df[regulatory_col])

    return df


def plot_enrichment(df, direction_col, comparison_col, value_col, significance_col, significance_thr,
                    force_directions=True, title=None, xlabel=None, ylabel=None, height=1,
                    sep_line=True, sep_line_kws=None, figsize=None):
    ax = None
    for direction in [-1, 1]:
        if ax is None:
            _, ax = plt.subplots(1, 1, figsize=figsize)
            if title is not None:
                ax.set_title(title)
            if xlabel is not None:
                ax.set_xlabel(xlabel)
            if ylabel is not None:
                ax.set_ylabel(ylabel)
        else:
            ax = ax.twinx()
        data = df.loc[df.index.get_level_values(direction_col) == DIRECTIONS[direction]]
        colors = [(RED if direction > 0 else BLUE) if p <= significance_thr else LIGHT_GRAY for p in data[significance_col]]
        plt.barh(data.index.get_level_values(comparison_col), data[value_col] * (direction if force_directions else 1),
                 color=colors, height=height)

    ax.tick_params(axis='y', right=False, labelright=False)
    xlim = np.max(np.abs(ax.get_xlim()))
    ax.set_xlim((-xlim, xlim))
    if sep_line:
        if sep_line_kws is None:
            sep_line_kws = {}
        ax.axvline(0, **sep_line_kws)
    sns.despine()
    return ax


def format_enrichment_for_plotting(df, main_col=REGION_SET_COL, main_col_direction='rs_association',
                                   col=COMPARISON_COL, extra_col=ASSOC_COL,
                                   sort_lambda=lambda x: 0 if 'd0' in x else 1 if '14' in x else 2,
                                   ascending=False):
    df = df.reset_index()
    assert len(set(df[main_col])) == 1
    enr_name = df[main_col].iloc[0]
    df['sort'] = df[col].apply(sort_lambda)
    sort_cols = [main_col_direction, 'sort']
    if extra_col is not None:
        sort_cols += [extra_col]
    df = df.sort_values(sort_cols, ascending=ascending)
    if extra_col is not None:
        df[col] = df[col] + ' -- ' + df[extra_col]
    df = df.set_index([main_col_direction, col])
    return df, enr_name


def make_bed(regions, fn, peak_annot, extras=['gene_name', 'characterization'], include_region_id=True):
    if isinstance(peak_annot, str):
        peak_annot = get_peak_annot(fn=peak_annot)
    bed = peak_annot.loc[regions, ['chr', 'start', 'end'] + extras]
    if include_region_id:
        bed['region'] = bed.index.values
    if fn:
        bed.to_csv(fn, index=False, header=False, sep='\t')
    return bed


def get_raw_counts(celltype):
    fn = RAW_COUNT_DATA_FN.format(celltype=celltype)
    return pd.read_csv(fn, index_col=0)


def get_norm_counts(celltype):
    fn = NORM_COUNT_DATA_FN.format(celltype=celltype)
    return pd.read_csv(fn, index_col=0)


def get_batch_corrected_counts(celltype):
    if celltype == 'PBMC':
        fn = COMBAT_COUNT_DATA_FN.format(celltype=celltype)
    else:
        fn = COMBAT_COUNT_CELLTYPES_DATA_FN.format(celltype=celltype)
    return pd.read_csv(fn, index_col=0)


def get_counts_fold_changes(celltype, visit):
    fn = COUNT_DATA_FN.format(celltype=celltype, data='norm.log2_fold_change.V1_{}'.format(visit))
    return pd.read_csv(fn, index_col=0)


def get_sample_annot_fn(which='.corrected'):
    return SAMPLE_ANNOT_FN.format(which=which)


def get_sample_annot(fn=SAMPLE_ANNOT_FN.format(which='.corrected'), which=None):
    df = pd.read_csv(fn if not which else SAMPLE_ANNOT_FN.format(which=which), index_col=0,
                     dtype={**{c: 'boolean' for c in NAN_BOOLS_ANNOT_COLS},
                            **{c: str for c in NAN_STRING_ANNOT_COLS}})
    _rename = {'CYTO:LPS.100ng_7d_PBMC_IFNg_good': 'CYTO:MTB_7d_PBMC_IFNg_good',
               'CYTO:LPS.100ng_7d_PBMC_IL.17_good': 'CYTO:MTB_7d_PBMC_IL.17_good'}
    df = df.rename(_rename, axis=1)

    # df['DONOR:IC_TIME_EARLY_VS_LATE'] = df['DONOR:IC_TIME_CAT']
    # df.loc[~df['DONOR:IC_TIME_EARLY_VS_LATE'].isin(['MOR8_9', 'MOR11_13']), 'DONOR:IC_TIME_EARLY_VS_LATE'] = np.nan
    # df.loc[df['DONOR:IC_TIME_EARLY_VS_LATE'] == 'MOR8_9', 'DONOR:IC_TIME_EARLY_VS_LATE'] = 0.0
    # df.loc[df['DONOR:IC_TIME_EARLY_VS_LATE'] == 'MOR11_13', 'DONOR:IC_TIME_EARLY_VS_LATE'] = 1.0
    # df['DONOR:IC_TIME_EARLY_VS_LATE'] = df['DONOR:IC_TIME_EARLY_VS_LATE'].astype(float)

    return df


def get_raw_cytokines(annot_df, covariates, remove_samples_with_missing_covariates=False):
    cyto_df = pd.read_csv(SELECTED_RAW_CYTO_FN)
    cyto_df = cyto_df.drop_duplicates()
    cyto_df['ID_V'] = cyto_df['ID_V'].str.split('_').apply(
        lambda x: '300BCG{:03d}_{}_PBMC'.format(int(x[0]), x[1]))
    cyto_df = cyto_df.set_index('ID_V')
    assert cyto_df.index.is_unique
    cyto_df = cyto_df.loc[~cyto_df.isnull().all(axis=1)]
    cyto_df['Box'] = 'Box_' + cyto_df['Box'].astype(int).astype(str)

    if annot_df is not None:
        cyto_df = cyto_df.reindex(annot_df.index)

    if covariates:
        assert annot_df is not None
        cyto_df.index = pd.MultiIndex.from_arrays(
            [cyto_df.index] + \
            [cyto_df[c] for c in set(covariates).intersection(cyto_df.columns)] + \
            [annot_df[c] for c in set(covariates).intersection(annot_df.columns)]
        )

    cyto_df = cyto_df.drop(['Box'], axis=1)
    cyto_df.columns = ['CYTO:{}'.format(c) for c in cyto_df.columns]
    if remove_samples_with_missing_covariates:
        cyto_df = cyto_df.loc[~cyto_df.index.to_frame().isnull().any(axis=1)]
    return cyto_df


def read_sysmex_data(fn=SYSMEX_DATA_FN, which='PBMC', fractions=True, include_WBC=False):
    assert which in ['PBMC', 'WB']
    _celltypes = {'NEUT': 'NEUTRO', 'LYMPH': 'LYMPHO', 'MONO': 'MONO', 'EO': 'EO', 'BASO': 'BASO', 'IG': 'IG'}
    _suffix = '%(%)' if fractions else '#(10^3/uL)'
    _sysmex_df = pd.read_excel(fn, na_values=['----', '++++', '    '], usecols=['New_name', 'Include']
                + ['{}{}'.format(c, _suffix) for c in _celltypes] + (['WBC(10^3/uL)'] if include_WBC else []))
    _sysmex_df = _sysmex_df.loc[_sysmex_df['Include'] == 1]
    wb_sysmex_df = _sysmex_df.loc[_sysmex_df['New_name'].str.match('WB_.*')]
    pbmc_sysmex_df = _sysmex_df.loc[_sysmex_df['New_name'].str.match('PBMC_.*')]

    _rename_index = lambda x: '300BCG{}_{}_PBMC'.format(x[3].upper(), x[2].upper())
    _rename_columns = {**{'WBC(10^3/uL)': 'SYSMEX:WBC'},
                       **{'{}{}'.format(c, _suffix): 'SYSMEX:{}'.format(_celltypes[c]) for c in _celltypes}}
    for _sysmex_df in [wb_sysmex_df, pbmc_sysmex_df]:
        _sysmex_df.set_index(_sysmex_df['New_name'].str.split('_').map(_rename_index), inplace=True)
        _sysmex_df.rename(_rename_columns, axis=1, inplace=True)
        _sysmex_df.drop(['New_name', 'Include'], axis=1, inplace=True)

    return (pbmc_sysmex_df if which == 'PBMC' else wb_sysmex_df).astype({'SYSMEX:{}'.format(_celltypes[c]): float for c in _celltypes})


def read_facs_data(fn=FACS_DATA_FN):
    facs_df = pd.read_excel(fn)
    facs_df = facs_df.set_index(
        facs_df['Id'].map(lambda x: '300{}_{}_PBMC'.format(x[0:6], x[6:].upper()))
    ).drop('Id', axis=1)
    facs_df.columns = facs_df.columns.map(lambda x: 'FACS:{}'.format(x))
    return facs_df


def read_genetic_outliers_data(fn=GENETIC_OUTLIERS_FN):
    df = pd.read_csv(fn, index_col=0)
    df.index.name = None
    return df['DONOR:SNP_OUTLIER']


def read_visit_dates(fn=CASTOR_METADATA_FN):
    df = pd.read_excel(fn, index_col=0)[['V1_ICDATE', 'V1_DATE', 'V2_DATE', 'V3_DATE']]
    for col in df.columns:
        for donor in ['300BCG009', '300BCG168']:
            df.loc[donor, col] = np.nan
    return df[['V1_ICDATE', 'V1_DATE', 'V2_DATE', 'V3_DATE']]


def read_extra_donor_data(fn=DONOR_EXTRA_DATA_FN, which_columns='all'):
    assert which_columns in ['all', 'donor', 'sample']
    df = pd.read_excel(fn, index_col=0, true_values=['yes'], false_values=['no'])
    df.index = '300' + df.index

    _visit_columns = ['{}_v{}'.format(p, v) for p in DONOR_EXTRA_VISIT_COLUMNS_PREFIXES for v in [1, 2, 3]]

    if which_columns == 'donor':
        df = df.loc[:, ~df.columns.isin(_visit_columns)]
        df.columns = 'DONOR:' + df.columns
    elif which_columns == 'sample':
        df = df.loc[:, _visit_columns]
        df.columns = 'SAMPLE:' + df.columns
    else:
        df.columns = ['SAMPLE:{}'.format(c) if c in _visit_columns else 'DONOR:{}'.format(c) for c in df.columns]

    # mean scar size
    if which_columns in ['all', 'donor']:
        for adjust_col in ['DONOR:sizeOfVaccinationSpot_v2', 'DONOR:sizeOfVaccinationSpot_LR_v2']:
            idx = df.index[df[adjust_col] > 2]
            df.loc[df[adjust_col] > 2, adjust_col] /= 10
        df['DONOR:sizeOfVaccinationSpot_avg_v2'] = df[['DONOR:sizeOfVaccinationSpot_v2', 'DONOR:sizeOfVaccinationSpot_LR_v2']].mean(axis=1)
        df['DONOR:sizeOfVaccinationSpot_v2_binary'] = (df['DONOR:sizeOfVaccinationSpot_v2'] > 0).astype('boolean')
        df.loc[df['DONOR:sizeOfVaccinationSpot_v2'].isnull(), 'DONOR:sizeOfVaccinationSpot_v2_binary'] = np.nan
        assert df['DONOR:sizeOfVaccinationSpot_v2_binary'].isnull().equals(df['DONOR:sizeOfVaccinationSpot_v2'].isnull())

        df['DONOR:scarSize_avg_v3'] = df[['DONOR:scarSize_v3', 'DONOR:scarSize_LR_v3']].mean(axis=1)
        df['DONOR:scarSize_v3_binary'] = (df['DONOR:scarSize_v3'] >= 0.5).astype('boolean')
        df.loc[df['DONOR:scarSize_v3'].isnull(), 'DONOR:scarSize_v3_binary'] = np.nan
        assert df['DONOR:scarSize_v3_binary'].isnull().equals(df['DONOR:scarSize_v3'].isnull())

        df['DONOR:parentBcg'] = df[['DONOR:motherBcg', 'DONOR:fatherBcg']].any(axis=1)
        df.loc[df[['DONOR:motherBcg', 'DONOR:fatherBcg']].isnull().all(axis=1), 'DONOR:parentBcg'] = np.nan
        df['DONOR:parentBcg'] = df['DONOR:parentBcg'].astype('boolean')

        assert df['DONOR:postMenopausal'].isin(['yes', 'yes_current', 'no', np.nan]).all()
        df['DONOR:currentMenopausal'] = df['DONOR:postMenopausal'].apply(
                lambda x: True if x == 'yes_current' else False if x == 'no' else np.nan).astype('boolean')
        df['DONOR:postMenopausal_binary'] = df['DONOR:postMenopausal'].apply(
                lambda x: True if x in ['yes', 'yes_current'] else False if x == 'no' else np.nan).astype('boolean')

        assert df['DONOR:vaginalBirth'].isin(['vaginal', 'cesarean', np.nan]).all()
        df['DONOR:vaginalBirth_binary'] = df['DONOR:vaginalBirth'].apply(
            lambda x: True if x == 'vaginal' else False if x == 'cesarean' else np.nan).astype('boolean')

        assert df['DONOR:happiness'].isin([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, np.nan]).all()
        df['DONOR:happiness_binary'] = df['DONOR:happiness'].apply(
            lambda x: True if x in [9, 10] else False if x in [0, 1, 2, 3, 4, 5, 6] else np.nan).astype('boolean')

        assert df['DONOR:qualityTeeth'].isin([0, 1, 2, np.nan]).all()
        df['DONOR:qualityTeeth_binary'] = df['DONOR:qualityTeeth'].apply(
            lambda x: True if x == 2 else False if x in [0, 1] else np.nan).astype('boolean')

        assert df['DONOR:exerciseDuringDay'].isin([0, 1, 2, np.nan]).all()
        df['DONOR:exerciseDuringDay_binary'] = df['DONOR:exerciseDuringDay'].apply(
            lambda x: True if x == 2 else False if x == 0 else np.nan).astype('boolean')

    # binarized mental health and perceived stress, etc.
    donor_01234_vars = ['DONOR:mentalHealth_3weeksBefore', 'DONOR:sleepQuality_lastTwoWeeks_v1', 'DONOR:stress_2weeksBefore']
    sample_01234_vars = ['SAMPLE:stress_dayBefore_v1', 'SAMPLE:stress_dayBefore_v2', 'SAMPLE:stress_dayBefore_v3']
    for var in (donor_01234_vars if which_columns in ['all', 'donor'] else []) + \
               (sample_01234_vars if which_columns in ['all', 'sample'] else []):
        assert df[var].isin([0, 1, 2, 3, 4, np.nan]).all(), df[var].drop_duplicates()
        if var.startswith('SAMPLE:'):
            assert var[-3:] in ['_v1', '_v2', '_v3']
            binarized_var = '{}_binary_{}'.format(var[:-3], var[-2:])
        else:
            binarized_var = '{}_binary'.format(var)
        df[binarized_var] = df[var].apply(
            lambda x: True if x == 4 else False if x in [0, 1, 2] else np.nan).astype('boolean')

    return df


def read_donor_data(fn=DONOR_DATA_FN, bmi_fn=DONOR_BMI_FN, corona_preriod_fn=CORONA_PERIOD_ANNOT_FN,
                    extra_donor_fn=DONOR_EXTRA_DATA_FN, extra_visit1_fn=DONOR_EXTRA_DATA_FN,
                    genetic_outliers_fn=GENETIC_OUTLIERS_FN):
    donor_df = pd.read_csv(fn, header=0,
                           names=['DONOR', 'DONOR:SEX', 'DONOR:BDATE', 'DONOR:IC_DATE', 'DONOR:IC_TIME'],
                           parse_dates=False)
    donor_df['DONOR:SEX'] = donor_df['DONOR:SEX'].apply(
        lambda code: 'F' if code == 1 else 'M' if code == 0 else np.nan)
    donor_df['DONOR:BDATE'] = pd.to_datetime(donor_df['DONOR:BDATE'].apply(lambda x: correct_date(x, century=1900)),
                                             format='%d.%m.%Y')
    donor_df['DONOR:IC_DATE'] = pd.to_datetime(donor_df['DONOR:IC_DATE'].apply(lambda x: correct_date(x, century=2000)),
                                               format='%d.%m.%Y')
    donor_df['DONOR:IC_TIME'] = pd.to_datetime(donor_df['DONOR:IC_TIME']).dt.time

    donor_df.loc[(donor_df['DONOR:IC_TIME'] <= datetime.time(9, 0)),
                 'DONOR:IC_TIME_CAT'] = 'MOR8_9'
    donor_df.loc[
        (donor_df['DONOR:IC_TIME'] > datetime.time(9, 0)) & (donor_df['DONOR:IC_TIME'] <= datetime.time(10, 0)),
        'DONOR:IC_TIME_CAT'] = 'MOR9_10'
    donor_df.loc[
        (donor_df['DONOR:IC_TIME'] > datetime.time(10, 0)) & (donor_df['DONOR:IC_TIME'] <= datetime.time(11, 0)),
        'DONOR:IC_TIME_CAT'] = 'MOR10_11'
    donor_df.loc[
        (donor_df['DONOR:IC_TIME'] > datetime.time(11, 0)) & (donor_df['DONOR:IC_TIME'] <= datetime.time(13, 0)),
        'DONOR:IC_TIME_CAT'] = 'MOR11_13'
    donor_df.loc[donor_df['DONOR:IC_TIME'] > datetime.time(13, 0),
                 'DONOR:IC_TIME_CAT'] = 'EVE'

    donor_df['DONOR:IC_EVENING'] = donor_df['DONOR:IC_TIME'] > datetime.time(13, 0)

    donor_df['DONOR:AGE'] = (donor_df['DONOR:IC_DATE'] - donor_df['DONOR:BDATE']) / np.timedelta64(1, 'Y')
    assert not (donor_df['DONOR:AGE'] >= 90).any()
    donor_df = donor_df.drop('DONOR:BDATE', axis=1)
    donor_df = donor_df.set_index('DONOR')

    if bmi_fn:
        bmi_df = pd.read_csv(bmi_fn, index_col=0, sep='\t',
                             usecols=['Castor Record ID', 'DEM3', 'DEM3A']).rename(
            {'DEM3': 'height', 'DEM3A': 'weight'}, axis=1)
        bmi_df = bmi_df.loc[~bmi_df['height'].isnull() & ~bmi_df['weight'].isnull()]
        donor_df['DONOR:BMI'] = bmi_df['weight'] / (np.power(bmi_df['height'] / 100, 2))
        donor_df['DONOR:HEIGHT'] = bmi_df['height']
        donor_df['DONOR:WEIGHT'] = bmi_df['weight']

    if corona_preriod_fn:
        corona_preriod_df = read_corona_period_data(corona_preriod_fn)
        donor_df = pd.concat([donor_df, corona_preriod_df], axis=1)

    if extra_donor_fn:
        extra_df = read_extra_donor_data(extra_donor_fn, which_columns='donor')
        donor_df = pd.concat([donor_df, extra_df], axis=1)

    for female_col, inclusive_col, male_value in [
        ('DONOR:postMenopausal', 'DONOR:postMenopausalIncludingMen', 'no'),
        ('DONOR:currentMenopausal', 'DONOR:currentMenopausalIncludingMen', False),
        ('DONOR:postMenopausal_binary', 'DONOR:postMenopausalIncludingMen_binary', False),
        ('DONOR:oralContraceptives', 'DONOR:oralContraceptivesIncludingMen', False)
    ]:
        if female_col in donor_df:
            donor_df[inclusive_col] = donor_df[female_col]
            donor_df.loc[donor_df['DONOR:SEX'] == 'M', inclusive_col] = male_value

    if extra_visit1_fn:
        visit1_df = read_extra_donor_data(extra_visit1_fn, which_columns='sample')
        visit1_df = visit1_df.loc[:, visit1_df.columns.str.upper().str.endswith('_V1')].rename(
            lambda c: 'DONOR:IC_' + '_'.join(':'.join(c.split(':')[1:]).split('_')[:-1]), axis=1)
        donor_df = pd.concat([donor_df, visit1_df], axis=1)

    if genetic_outliers_fn:
        genetic_outliers = read_genetic_outliers_data(genetic_outliers_fn)
        donor_df = pd.concat([donor_df, genetic_outliers], axis=1)

    return donor_df


def read_corona_period_data(corona_preriod_fn=CORONA_PERIOD_ANNOT_FN):
    corona_preriod_df = pd.read_csv(corona_preriod_fn, index_col=0, sep='\t')
    corona_preriod_df.columns = 'CTP:' + corona_preriod_df.columns.str.replace('CTP$', '')
    corona_preriod_df.index.name = 'DONOR'
    return corona_preriod_df


def read_circulating_mediators_data(circ_mediators_fn=CORR_CIRC_MEDIATORS_FN):
    circ_mediators_df = pd.read_excel(circ_mediators_fn, index_col=0)
    circ_mediators_df.index = convert_short_index(circ_mediators_df.index)
    circ_mediators_df.columns = 'CM:' + circ_mediators_df.columns
    return circ_mediators_df


def read_cyto_batches(cyto_fn=RAW_CYTO_FN):
    df = pd.read_excel(cyto_fn, sep=';', usecols=['Patient_ID', 'visit', 'Box'])
    assert (df.groupby(['Patient_ID'])['Box'].nunique() == 1).all()
    df.index = ['300BCG{:03d}_V{}_PBMC'.format(int(idx), str(visit)) for idx, visit in
                zip(df['Patient_ID'], df['visit'])]
    df['DONOR:CYTO_BATCH'] = 'Box' + df['Box'].astype(str)
    return df['DONOR:CYTO_BATCH']


def read_cyto_data(cyto_fn=CORR_CYTO_FN, qual_fn=CYTO_QUAL_FN, log2=False):
    cyto_df = pd.read_excel(cyto_fn, index_col=0)
    _qual_xls_file = pd.ExcelFile(qual_fn)
    qual = pd.DataFrame()
    for sheet_name in _qual_xls_file.sheet_names:
        _n, name = sheet_name.split('.')
        assert int(_n) in range(1, 6)
        _df = _qual_xls_file.parse(sheet_name, header=None)  # read a specific sheet to DataFrame
        _df[1] = name.replace('justForLukas', 'bad.with.visit.effects')
        qual = pd.concat([qual, _df.set_index(0)], axis=0)
    qual = qual[1]
    assert qual.index.is_unique
    assert set(qual.index) == set(cyto_df.columns)

    new_columns = 'CYTO:' + cyto_df.columns + '_' + qual.loc[cyto_df.columns]
    cyto_df.columns = new_columns.loc[cyto_df.columns].values
    cyto_df.index = convert_short_index(cyto_df.index)

    _rename = {'CYTO:LPS.100ng_7d_PBMC_IFNg_good': 'CYTO:MTB_7d_PBMC_IFNg_good',
               'CYTO:LPS.100ng_7d_PBMC_IL.17_good': 'CYTO:MTB_7d_PBMC_IL.17_good'}
    for r in _rename:
        if r in cyto_df.columns:
            print('Renaming', r, '-->', _rename[r])
    cyto_df = cyto_df.rename(_rename, axis=1)

    return np.log2(cyto_df) if log2 else cyto_df


def read_exclusions_data(fn=EXCLUSION_FN):

    CYTO_7D_REMOVE_REASON = 'Incorrect 7d stimulation duration'
    CYTO_POS_NEG_CONTROL_REASON = 'Positive negative control'

    exclusions = pd.Series(name='SAMPLE:EXCLUSION', dtype=str)
    cyto_pos_neg_control = pd.Series(name='CYTO_POS_NEG_CONTROL', dtype='boolean')
    cyto_7days_remove = pd.Series(name='CYTO_7D_REMOVE', dtype='boolean')

    df = pd.read_csv(fn, index_col=0)

    for donor in df.index:
        for visit in df.loc[donor, 'visit'].split(','):
            visit = visit.strip()
            for celltype in ['PBMC', 'cd8t', 'monocyte', 'nkcell']:
                sample_idx = '{}_{}_{}'.format(donor, visit, celltype)
                if celltype == 'PBMC':
                    if df.loc[donor, 'reason'] == CYTO_7D_REMOVE_REASON:
                        cyto_7days_remove.loc[sample_idx] = True
                    elif df.loc[donor, 'reason'] == CYTO_POS_NEG_CONTROL_REASON:
                        cyto_pos_neg_control.loc[sample_idx] = True
                if df.loc[donor, 'reason'] not in [CYTO_7D_REMOVE_REASON, CYTO_POS_NEG_CONTROL_REASON]:
                    exclusions.loc[sample_idx] = df.loc[donor, 'reason']
    return exclusions.str.strip(), cyto_pos_neg_control, cyto_7days_remove


def make_sample_annot(donor_df=None, sample_df=None, lab_df=None, sysmex_df=None, sysmex_counts_df=None,
                      wb_sysmex_df=None, wb_sysmex_counts_df=None, facs_df=None, cyto_df=None, circ_mediators_df=None,
                      season_corrected_donor_scores=False,
                      cyto_formula='SAMPLE_VISIT + SAMPLE_DONOR + SAMPLE_VISIT_TIME_REAL + ' + replace_chars_for_LM(' + '.join(BLOOD)),
                      cyto_formula_cols=CYTO_MODEL_COLS,
                      CM_formula='SAMPLE_VISIT + SAMPLE_DONOR + SAMPLE_VISIT_TIME_REAL + ' + replace_chars_for_LM(' + '.join(WHOLE_BLOOD)),
                      V1_blood_cols = ['DONOR:AGE', 'DONOR:SEX', 'SAMPLE:VISIT_TIME_REAL'],
                      V1_cyto_cols = ['DONOR:AGE', 'DONOR:SEX', 'SAMPLE:VISIT_TIME_REAL'] + BLOOD,
                      V1_CM_cols = ['DONOR:AGE', 'DONOR:SEX', 'SAMPLE:VISIT_TIME_REAL', 'DONOR:BMI', 'DONOR:oralContraceptivesIncludingMen'] + WHOLE_BLOOD,
                      CM_formula_cols=CM_MODEL_COLS, cytokine_V1_adj_scores=False, standard_scaling_scores=False,
                      split_T_cells=True, split_monocytes=True, split_NK_cells=True,
                      fix_V1_visit_time=True, remove_evening=False):

    ###############################
    # DONOR AND SAMPLE ATTRIBUTES #
    ###############################

    if donor_df is None:
        donor_df = read_donor_data()
    assert donor_df.columns.str.contains('^DONOR:|^CTP:').all()

    if sample_df is None:
        sample_df = read_extra_donor_data(which_columns='sample')
    assert sample_df.columns.str.contains('^SAMPLE:.*_v[1-3]$|^SAMPLE:.*_v[1-3]_binary$').all(), sample_df.columns

    annot_df = pd.DataFrame()
    for tissue in ['PBMC', 'nkcell', 'monocyte', 'cd8t', 'neutrophil']:
        for visit in ['V1', 'V2', 'V3']:
            _df = donor_df.copy()
            _df['SAMPLE:DONOR'] = _df.index
            _df['SAMPLE:VISIT'] = visit
            _df['SAMPLE:TISSUE'] = tissue
            _df = pd.concat([_df,
                             sample_df.loc[:, sample_df.columns.str.upper().str.endswith('_{}'.format(visit))].rename(
                                 lambda c: '_'.join(c.split('_')[:-1]), axis=1
                             )], axis=1)
            _df.index = _df['SAMPLE:DONOR'] + '_' + _df['SAMPLE:VISIT'] + '_' + _df['SAMPLE:TISSUE']
            annot_df = pd.concat([annot_df, _df], axis=0)
    assert annot_df.columns.str.contains('^SAMPLE:|^DONOR:|^CTP:').all()

    if lab_df is None:
        lab_df = pd.read_csv(SAMPLE_ANNOT_FN.format(which=''), index_col=0)
    assert lab_df.index.str.endswith('_ATAC_R1').all()
    lab_df.index = lab_df.index.str.replace('_ATAC_R1$', '')
    lab_df = lab_df.loc[:, ~lab_df.columns.str.contains('^SAMPLE:|^DONOR:|^CTP:|SYSMEX|^FACS:|^CYTO:')]
    assert not lab_df['DEMUX:BARCODE'].isnull().any()
    assert lab_df.index.isin(annot_df.index).all()
    annot_df = pd.concat([annot_df, lab_df], axis=1)

    # remove non-PBMC samples without ATAC-seq libraries
    annot_df = annot_df.loc[(annot_df['SAMPLE:TISSUE'] == 'PBMC') | (annot_df['QC:PASS'] == True)]

    # setup a bool variable to denote whether there was a misplaced sample
    _correct_batch = np.zeros((annot_df.shape[0],), dtype=bool)
    for celltype, batch_prefix in [('PBMC', 'BCG_P'), ('nkcell', 'BCG_KILLER'), ('monocyte', 'BCG_MONO'),
                                   ('cd8t', 'BCG_TCELL'), ('neutrophil', 'BCG_NEUTRO')]:
        _correct_batch |= (annot_df['SAMPLE:TISSUE'] == celltype) & annot_df['LAB:BATCH'].str.startswith(batch_prefix)
    annot_df['LAB:WRONG_BATCH'] = ~_correct_batch
    annot_df.loc[annot_df['LAB:BATCH'].isnull(), 'LAB:WRONG_BATCH'] = np.nan
    annot_df['LAB:WRONG_BATCH'] = annot_df['LAB:WRONG_BATCH'].astype('boolean')
    assert (annot_df.loc[annot_df['LAB:WRONG_BATCH'] == False].groupby(['SAMPLE:DONOR', 'SAMPLE:TISSUE'])[
                'LAB:BATCH'].nunique() == 1).sum()

    #####################
    # BLOOD COMPOSITION #
    #####################

    for _df, which, fractions, include_WBC in [
        (sysmex_df, 'PBMC', True, False),
        (sysmex_counts_df, 'PBMC', False, True),
        (wb_sysmex_df, 'WB', True, False),
        (wb_sysmex_counts_df, 'WB', False, True)
    ]:
        if _df is None:
            _df = read_sysmex_data(which=which, fractions=fractions, include_WBC=include_WBC)
        assert not _df.columns.str.contains('^SAMPLE:|^DONOR:|^CTP:').any(), _df.columns.tolist()
        assert _df.index.isin(annot_df.index).all()
        _prefix = '{}_SYSMEX_{}:'.format(which, 'PERC' if fractions else 'PER_ML')
        _df.columns = _df.columns.str.replace('SYSMEX:', _prefix)
        annot_df = pd.concat([annot_df, _df], axis=1)

    if facs_df is None:
        facs_df = read_facs_data()
    assert not facs_df.columns.str.contains('^SAMPLE:|^DONOR:|^CTP:').any()
    assert facs_df.index.isin(annot_df.index).all()
    annot_df = pd.concat([annot_df, facs_df], axis=1)

    for cell_type in ['MONO', 'LYMPHO', 'NEUTRO', 'BASO', 'EO']:
        annot_df['PBMC_PERC:{}'.format(cell_type)] = annot_df['PBMC_SYSMEX_PERC:{}'.format(cell_type)]
    for lympho_type in ['T', 'B', 'NK', 'NKT']:
        annot_df['PBMC_PERC:{}'.format(lympho_type)] = (annot_df['FACS:LYMPHO/{}'.format(lympho_type)] / annot_df['FACS:LYMPHO']) * annot_df['PBMC_SYSMEX_PERC:LYMPHO']
    if split_T_cells:
        for T_type in ['T/CD8', 'T/CD4', 'T/CD4/TREG']:
            annot_df['PBMC_PERC:{}'.format(T_type)] = (annot_df['FACS:LYMPHO/{}'.format(T_type)] / annot_df['FACS:LYMPHO']) * annot_df['PBMC_SYSMEX_PERC:LYMPHO']
    if split_monocytes:
        for mono_type in ['CLASSICAL', 'INTERMEDIATE', 'NON_CLASSICAL']:
            annot_df['PBMC_PERC:MONO/{}'.format(mono_type)] = (annot_df['FACS:MONO/{}'.format(mono_type)] / annot_df['FACS:MONO']) * annot_df['PBMC_SYSMEX_PERC:MONO']
    if split_NK_cells:
        for NK_type in ['NK/BRIGHT', 'NK/DIM', 'NK/OTHER']:
            annot_df['PBMC_PERC:{}'.format(NK_type)] = (annot_df['FACS:LYMPHO/{}'.format(NK_type)] / annot_df['FACS:LYMPHO']) * annot_df['PBMC_SYSMEX_PERC:LYMPHO']

    annot_df['PBMC_PER_ML:WBC'] = annot_df['PBMC_SYSMEX_PER_ML:WBC']
    for cell_type in annot_df.columns[annot_df.columns.str.contains('^PBMC_PERC:')]:
        cell_type = cell_type[len('PBMC_PERC:'):]
        if cell_type in ['MONO', 'LYMPHO', 'NEUTRO', 'BASO', 'EO']:
            annot_df['PBMC_PER_ML:{}'.format(cell_type)] = annot_df['PBMC_SYSMEX_PER_ML:{}'.format(cell_type)]
            x = annot_df['PBMC_PERC:{}'.format(cell_type)] / 100 * annot_df['PBMC_PER_ML:WBC']
            assert (x - annot_df['PBMC_PER_ML:{}'.format(cell_type)]).abs().max() < 0.075
        else:
            annot_df['PBMC_PER_ML:{}'.format(cell_type)] = annot_df['PBMC_PERC:{}'.format(cell_type)] / 100 * annot_df['PBMC_PER_ML:WBC']

    for cell_type in annot_df.columns[annot_df.columns.str.contains('^FACS:')]:
        cell_type = cell_type[len('FACS:'):]
        if cell_type not in ['LEUCO', 'WBC']:
            annot_df['WB_PERC:{}'.format(cell_type.replace('LYMPHO/', ''))] = (annot_df['FACS:{}'.format(cell_type)] / annot_df['FACS:LEUCO']) * 100
            annot_df['WB_PER_ML:{}'.format(cell_type.replace('LYMPHO/', ''))] = (annot_df['FACS:{}'.format(cell_type)] / annot_df['FACS:LEUCO']) * annot_df['FACS:WBC']
    annot_df['WB_PER_ML:WBC'] = annot_df['FACS:WBC']
    annot_df = annot_df.drop(annot_df.columns[annot_df.columns.str.contains('^FACS:')], axis=1)

    #######################################
    # CYTOKINES AND CIRCULATING MEDIATORS #
    #######################################

    if cyto_df is None:
        cyto_df = read_cyto_data(cyto_fn=CORR_CYTO_FN, qual_fn=CYTO_QUAL_FN, log2=True)
    assert not cyto_df.columns.str.contains('^SAMPLE:|^DONOR:|^CTP:').any()
    assert cyto_df.index.isin(annot_df.index).all()
    annot_df = pd.concat([annot_df, cyto_df], axis=1)

    cyto_batches = read_cyto_batches()
    annot_df = pd.concat([annot_df, cyto_batches], axis=1)

    if circ_mediators_df is None:
        circ_mediators_df = read_circulating_mediators_data(circ_mediators_fn=CORR_CIRC_MEDIATORS_FN)
    assert not circ_mediators_df.columns.str.contains('^SAMPLE:|^DONOR:|^CTP:').any()
    assert circ_mediators_df.index.isin(annot_df.index).all()
    annot_df = pd.concat([annot_df, circ_mediators_df], axis=1)

    #####################
    # EXCLUSIONS AND QC #
    #####################

    exclusions, cyto_pos_neg_control, cyto_7days_remove = read_exclusions_data()
    exclusions = exclusions.loc[exclusions.index.isin(annot_df.index)]
    annot_df = pd.concat([annot_df, exclusions], axis=1)
    # check that all cytokine data is ALREADY excluded
    # thus also LFC_CYTO, IC_CYTO, and responder status are excluded
    assert cyto_pos_neg_control.all() and annot_df.loc[
        cyto_pos_neg_control.index, annot_df.columns.str.startswith('CYTO:')].isnull().all().all()
    assert cyto_7days_remove.all() and annot_df.loc[
        cyto_7days_remove.index, annot_df.columns.str.contains('^CYTO:.*_7d_.*')].isnull().all().all()
    assert annot_df.loc[exclusions.index, annot_df.columns.str.startswith('CYTO:')].isnull().all().all()

    # set nan to all data for exclusions and non-PBMC samples
    for prefix in ['CYTO:', 'CM:',
                   'PBMC_PERC:', 'PBMC_PER_ML:', 'WB_PERC:', 'WB_PER_ML:',
                   'PBMC_SYSMEX_PERC', 'PBMC_SYSMEX_PER_ML', 'WB_SYSMEX_PERC', 'WB_SYSMEX_PER_ML']:
        assert annot_df.columns.str.startswith(prefix).sum() != 0
        annot_df.loc[exclusions.index, annot_df.columns.str.startswith(prefix)] = np.nan
        annot_df.loc[annot_df['SAMPLE:TISSUE'] != 'PBMC', annot_df.columns.str.startswith(prefix)] = np.nan

    ####################################################
    # CYTOKINES AND CIRCULATING MEDIATORS FOLD-CHANGES #
    ####################################################

    for df in [cyto_df.copy(), circ_mediators_df.copy()]:
        df.index = df.index.str.split('_', expand=True)
        df.index.names = ['SAMPLE:DONOR', 'SAMPLE:VISIT', 'SAMPLE:TISSUE']
        lfc_df = pd.DataFrame()
        for visit in ['V2', 'V3']:
            _diff_df = visit_diff(df, visit=visit, base_visit='V1')
            annot_df = annot_df.join(_diff_df.rename(lambda x: 'LFC_{}_{}'.format(visit, x), axis=1), on='SAMPLE:DONOR')
            _diff_df = _diff_df.rename(lambda x: 'LFC_{}'.format(x), axis=1).rename(lambda x: '{}_{}_PBMC'.format(x, visit), axis=0)
            lfc_df = pd.concat([lfc_df, _diff_df], axis=0)
        annot_df = pd.concat([annot_df, lfc_df], axis=1)

    # set nan to all data for exclusions and non-PBMC samples
    for prefix in ['LFC_CYTO:', 'LFC_V2_CYTO:', 'LFC_V3_CYTO:',
                   'LFC_CM:', 'LFC_V2_CM:', 'LFC_V3_CM:']:
        assert annot_df.columns.str.startswith(prefix).sum() != 0
        annot_df.loc[exclusions.index, annot_df.columns.str.startswith(prefix)] = np.nan
        annot_df.loc[annot_df['SAMPLE:TISSUE'] != 'PBMC', annot_df.columns.str.startswith(prefix)] = np.nan

    ########################################################
    # V1 levels of cytokines, CMs, whole-blood composition #
    ########################################################

    _v1_df = annot_df.loc[
        annot_df['SAMPLE:VISIT'] == 'V1',
        annot_df.columns.str.contains('^CYTO:|^CM:|^WB_PERC:|^WB_PER_ML:|^WB_SYSMEX_PERC:|^WB_SYSMEX_PER_ML:|^SAMPLE:DONOR$')
    ].set_index('SAMPLE:DONOR')
    assert (~_v1_df.columns.str.contains('DATE|TIME')).all()
    _v1_df.columns = 'IC_' + _v1_df.columns
    annot_df = annot_df.join(_v1_df.loc[~_v1_df.isnull().all(axis=1)], on='SAMPLE:DONOR')

    # set nan to all data for exclusions and non-PBMC samples
    for prefix in ['IC_CYTO:', 'IC_CM:',
                   'IC_WB_PERC:', 'IC_WB_PER_ML:', 'IC_WB_SYSMEX_PERC:', 'IC_WB_SYSMEX_PER_ML:']:
        assert annot_df.columns.str.startswith(prefix).sum() != 0
        annot_df.loc[exclusions.index, annot_df.columns.str.startswith(prefix)] = np.nan
        annot_df.loc[annot_df['SAMPLE:TISSUE'] != 'PBMC', annot_df.columns.str.startswith(prefix)] = np.nan

    #######################
    # VISIT DATE AND TIME #
    #######################

    dates_df = read_visit_dates()
    assert dates_df.index.is_unique
    for d in dates_df.index:
        if annot_df.loc[annot_df['SAMPLE:DONOR'] == d, 'DONOR:IC_DATE'].iloc[0] != dates_df.loc[d, 'V1_ICDATE']:
            if not (pd.isnull(annot_df.loc[annot_df['SAMPLE:DONOR'] == d, 'DONOR:IC_DATE'].iloc[0]) and pd.isnull(dates_df.loc[d, 'V1_ICDATE'])):
                print('Difference in annotations!')
                print(annot_df.loc[annot_df['SAMPLE:DONOR'] == d, 'DONOR:IC_DATE'])
                print('IC', dates_df.loc[d, 'V1_ICDATE'])
                print('V1', dates_df.loc[d, 'V1_DATE'])
                print('V2', dates_df.loc[d, 'V2_DATE'])
                print('V3', dates_df.loc[d, 'V3_DATE'])
                print('')

    for visit in ['V1', 'V2', 'V3']:
        v_times = pd.read_csv(os.path.join(METADATA, '{}_times.tsv'.format(visit)), sep='\t', index_col=2)['Time']
        assert v_times.index.is_unique
        v_times = pd.to_datetime(v_times.str.replace('.', ':')).dt.time
        for donor in v_times.index:
            annot_df.loc[(annot_df['SAMPLE:VISIT'] == visit) & \
                         (annot_df['SAMPLE:DONOR'] == donor), 'SAMPLE:VISIT_TIME'] = v_times.loc[donor]

        for donor in dates_df.index:
            date = dates_df.loc[donor, f'{visit}_DATE']
            if not pd.isnull(date):
                annot_df.loc[(annot_df['SAMPLE:VISIT'] == visit) & \
                             (annot_df['SAMPLE:DONOR'] == donor), 'SAMPLE:VISIT_DATE'] = date

    for _prefix, _type in [('SAMPLE', 'VISIT'), ('DONOR', 'IC')]:
        annot_df['{}:{}_DAYS_FROM_20170101'.format(_prefix, _type)] = \
            annot_df['{}:{}_DATE'.format(_prefix, _type)].copy() - pd.to_datetime('2017-01-01')
        annot_df['{}:{}_DATE_REAL'.format(_prefix, _type)] = \
            annot_df['{}:{}_DAYS_FROM_20170101'.format(_prefix, _type)] / np.timedelta64(1, 'Y')
        annot_df['{}:{}_DAYS_FROM_20170101'.format(_prefix, _type)] = \
            annot_df['{}:{}_DAYS_FROM_20170101'.format(_prefix, _type)] / np.timedelta64(1, 'D')
        annot_df['{}:{}_DAYS_FROM_0101'.format(_prefix, _type)] = \
            annot_df['{}:{}_DAYS_FROM_20170101'.format(_prefix, _type)].copy() % 365
        annot_df['{}:{}_DATE_2PI'.format(_prefix, _type)] = \
            annot_df['{}:{}_DATE_REAL'.format(_prefix, _type)] * 2 * np.pi
        annot_df['{}:{}_DATE_2PI_SIN'.format(_prefix, _type)] = np.sin(
            annot_df['{}:{}_DATE_2PI'.format(_prefix, _type)])
        annot_df['{}:{}_DATE_2PI_COS'.format(_prefix, _type)] = np.cos(
            annot_df['{}:{}_DATE_2PI'.format(_prefix, _type)])
        annot_df['{}:{}_MONTH_REAL'.format(_prefix, _type)] = \
            annot_df['{}:{}_DATE_REAL'.format(_prefix, _type)] * 12 + 1
        annot_df.loc[annot_df['{}:{}_MONTH_REAL'.format(_prefix, _type)] >= 13,
                     '{}:{}_MONTH_REAL'.format(_prefix, _type)] -= 12
        assert not ((annot_df['{}:{}_MONTH_REAL'.format(_prefix, _type)] < 1).any() or (
                annot_df['{}:{}_MONTH_REAL'.format(_prefix, _type)] >= 13).any())

    _time_to_real = lambda x: (pd.to_timedelta(pd.to_datetime(x, format='%H:%M:%S').strftime('%H:%M:%S'),
                                               unit='m').total_seconds() / 3600) if not pd.isnull(x) else np.nan
    annot_df['DONOR:IC_TIME_REAL'] = annot_df['DONOR:IC_TIME'].apply(_time_to_real)
    annot_df['SAMPLE:VISIT_TIME_REAL'] = annot_df['SAMPLE:VISIT_TIME'].apply(_time_to_real)

    for c in ['SAMPLE:VISIT_DATE', 'DONOR:IC_DATE', 'SAMPLE:VISIT_TIME', 'DONOR:IC_TIME']:
        col = annot_df[c].copy()
        col.loc[~col.isnull()] = col.loc[~col.isnull()].astype(str)
        col.loc[col.isnull()] = np.nan
        annot_df[c] = col

    if fix_V1_visit_time:
        annot_df.loc[(annot_df['SAMPLE:VISIT'] == 'V1') & (
                annot_df['DONOR:IC_TIME_REAL'] <= 13), 'SAMPLE:VISIT_TIME_REAL'] = \
            annot_df.loc[
                (annot_df['SAMPLE:VISIT'] == 'V1') & (
                        annot_df['DONOR:IC_TIME_REAL'] <= 13), 'DONOR:IC_TIME_REAL']
        annot_df.loc[
            (annot_df['SAMPLE:VISIT'] == 'V1') & (annot_df['DONOR:IC_TIME_REAL'] <= 13), 'SAMPLE:VISIT_TIME'] = \
            annot_df.loc[
                (annot_df['SAMPLE:VISIT'] == 'V1') & (annot_df['DONOR:IC_TIME_REAL'] <= 13), 'DONOR:IC_TIME']
        for donor in set(
                annot_df.loc[~annot_df['SAMPLE:VISIT_TIME_REAL'].isnull() & annot_df['DONOR:IC_TIME_REAL'].isnull(), 'SAMPLE:DONOR']):
            annot_df.loc[annot_df['SAMPLE:DONOR'] == donor, 'DONOR:IC_TIME_REAL'] = annot_df.loc[
                '{}_V1_PBMC'.format(donor), 'SAMPLE:VISIT_TIME_REAL']
            annot_df.loc[annot_df['SAMPLE:DONOR'] == donor, 'DONOR:IC_TIME'] = annot_df.loc[
                '{}_V1_PBMC'.format(donor), 'SAMPLE:VISIT_TIME']

    if remove_evening:
        print('Remove evening')
        annot_df = annot_df.loc[annot_df['DONOR:IC_TIME_REAL'] <= 13]
        print('annot_df', annot_df.shape)
    else:
        circadian_df = pd.read_csv(CIRCADIAN_DATA_FN, index_col=0)
        for circadian_group in ['EVE', 'MOR']:
            annot_df.loc[(annot_df['SAMPLE:TISSUE'] == 'PBMC') & \
                         annot_df['SAMPLE:DONOR'].isin(
                             circadian_df.loc[circadian_df['GROUP'] == circadian_group].index),
                         'DONOR:CIRCAD_REPLIC'] = circadian_group
        assert (annot_df['DONOR:CIRCAD_REPLIC'] == 'EVE').equals(annot_df['DONOR:IC_EVENING'])

    #################################################
    # CORRECTED CYTOKINES AND CIRCULATING MEDIATORS #
    #################################################

    for prefix, design, cols in [
        (r'^CYTO:', cyto_formula, cyto_formula_cols),
        (r'^CM:', CM_formula, CM_formula_cols)
    ]:
        X_df, Y_df = drop_na_for_LM(annot_df[cols], annot_df.loc[:, annot_df.columns.str.contains(prefix)])
        assert X_df.index.str.contains('_PBMC$').all()
        _, _, corrected_df, _ = fit_linear_model(X_df, Y_df, design=design, do_not_correct=['SAMPLE_VISIT'],
                                                 return_corrected_X=True, random_state=RANDOM_STATE, verbose=True)
        corrected_df = corrected_df.rename(lambda c: 'CORR_{}'.format(c), axis=1)
        annot_df = pd.concat([annot_df, corrected_df], axis=1)

        # Fold changes
        corrected_df.index = corrected_df.index.str.split('_', expand=True)
        corrected_df.index.names = ['SAMPLE:DONOR', 'SAMPLE:VISIT', 'SAMPLE:TISSUE']
        lfc_df = pd.DataFrame()
        for visit in ['V2', 'V3']:
            _diff_df = visit_diff(corrected_df, visit=visit, base_visit='V1')
            annot_df = annot_df.join(_diff_df.rename(lambda x: 'LFC_{}_{}'.format(visit, x), axis=1), on='SAMPLE:DONOR')
            _diff_df = _diff_df.rename(lambda x: 'LFC_{}'.format(x), axis=1).rename(lambda x: '{}_{}_PBMC'.format(x, visit), axis=0)
            lfc_df = pd.concat([lfc_df, _diff_df], axis=0)
        annot_df = pd.concat([annot_df, lfc_df], axis=1)

    # V1 only
    for prefix, cols in [
        (r'^WB_PER_ML:', V1_blood_cols),
        (r'^CYTO:', V1_cyto_cols),
        (r'^CM:', V1_CM_cols),
    ]:
        V1_df = annot_df.loc[(annot_df['SAMPLE:VISIT'] == 'V1') & (annot_df['SAMPLE:TISSUE'] == 'PBMC')].copy()
        V1_df = V1_df.loc[~V1_df[cols].isnull().any(axis=1)]
        _, _, V1_df, _ = fit_linear_model(
            X_df=V1_df.loc[:, cols], Y_df=V1_df.loc[:, V1_df.columns.str.contains(prefix)],
            design=' + '.join([str(1)] + [_encode_coef(c, not is_numeric_dtype(V1_df[c].dtype)) for c in cols]),
            lmm_groups=None, do_not_correct=None, return_corrected_X=True, just_correction=True, random_state=RANDOM_STATE)
        V1_df.columns = 'V1_CORR_' + V1_df.columns
        annot_df = pd.concat([annot_df, V1_df], axis=1)

    # set nan to all data for exclusions and non-PBMC samples
    for prefix in ['V1_CORR_WB_PER_ML:', 'V1_CORR_CYTO:', 'V1_CORR_CM:',
                   'CORR_CYTO:', 'LFC_CORR_CYTO:', 'LFC_V2_CORR_CYTO:', 'LFC_V3_CORR_CYTO:',
                   'CORR_CM:', 'LFC_CORR_CM:', 'LFC_V2_CORR_CM:', 'LFC_V3_CORR_CM:']:
        assert annot_df.columns.str.startswith(prefix).sum() != 0
        annot_df.loc[exclusions.index, annot_df.columns.str.startswith(prefix)] = np.nan
        annot_df.loc[annot_df['SAMPLE:TISSUE'] != 'PBMC', annot_df.columns.str.startswith(prefix)] = np.nan

    _v1_df = annot_df.loc[
        annot_df['SAMPLE:VISIT'] == 'V1',
        annot_df.columns.str.contains('^V1_CORR_WB_PER_ML:|^SAMPLE:DONOR$')].set_index('SAMPLE:DONOR')
    assert (~_v1_df.columns.str.contains('DATE|TIME')).all()
    _v1_df.columns = 'IC_' + _v1_df.columns
    annot_df = annot_df.join(_v1_df.loc[~_v1_df.isnull().all(axis=1)], on='SAMPLE:DONOR')

    ###############################################
    # Rename cytokines excluded from the analysis #
    ###############################################

    annot_df.columns = annot_df.columns.str.replace(':C.albicans.yeast_24h_PBMC_IFNg_good',
                                                    ':C.albicans.yeast_24h_PBMC_IFNg_excluded')

    ################
    # DONOR SCORES #
    ################
    donor_scores_df, _ = make_donor_scores(
        annot_df.loc[:, annot_df.columns.str.contains('^CYTO:.*_good$|^CYTO:C.albicans.yeast_24h_PBMC_IFNg_excluded$')] \
            if not season_corrected_donor_scores else pd.read_csv(SEASON_CORRECTED_CYTO, index_col=0),
        visit='V3',
        X_df=annot_df[cyto_formula_cols] if not season_corrected_donor_scores else None,
        design=cyto_formula if not season_corrected_donor_scores else None,
        do_not_correct=['Intercept', 'SAMPLE_DONOR', 'SAMPLE_VISIT'],
        remove_V1_production=cytokine_V1_adj_scores, standard_scaling=standard_scaling_scores,
        methods=['top_half_mean'], max_impute_iter=10,
        random_state=RANDOM_STATE, save_fig=True, fig_dir=make_dir('results', 'figures'))
    donor_scores_df.columns = donor_scores_df.columns.str.replace(
        '^top_half_mean\.', 'thm.').str.replace('^top_half_floor_mean\.', 'thfm.').str.replace('^trimmed_mean\.', 'tm.').str.replace('^mean_without_worst\.', 'mww.').str.replace('^mean\.', 'm.')
    annot_df = annot_df.join(donor_scores_df, on='SAMPLE:DONOR')

    for fold_change in [1.2]:
        responders_df = binarize_donor_scores(donor_scores_df,
                                              fraction_responders=None,
                                              fold_change=fold_change,
                                              save_fig=True, fig_dir=make_dir('results', 'figures'))
        annot_df = annot_df.join(responders_df, on='SAMPLE:DONOR')

    # set nan to all data for exclusions and non-PBMC samples
    for col in donor_scores_df.columns.tolist() + responders_df.columns.tolist():
        assert col in annot_df.columns
        annot_df.loc[exclusions.index, col] = np.nan

    ###########
    # SUMMARY #
    ###########

    all_null = annot_df.loc[:, ~annot_df.columns.str.contains('^SAMPLE:|^DONOR:|^CTP:|^LAB:WRONG_BATCH$')].isnull().all(
        axis=1)
    print('\nWrong batch in {} samples'.format((annot_df['LAB:WRONG_BATCH'] == True).sum()))
    print('Samples with no annotations: {}'.format(all_null.sum()))
    print('Circadian analysis:',
          'evening cohort', len(set(annot_df.loc[annot_df['DONOR:CIRCAD_REPLIC'] == 'EVE', 'SAMPLE:DONOR'])),
          'morning cohort', len(set(annot_df.loc[annot_df['DONOR:CIRCAD_REPLIC'] == 'MOR', 'SAMPLE:DONOR'])))
    print('annot_df', annot_df.shape)
    annot_df.index.name = 'SAMPLE:ID'
    return annot_df


def impute_cytokines(cyto_df, max_iter=10, random_state=RANDOM_STATE):
    cyto_df = cyto_df.loc[~cyto_df.isnull().all(axis=1)].copy()
    imputed_df = pd.DataFrame(
        data=IterativeImputer(max_iter=max_iter, random_state=random_state).fit_transform(cyto_df),
        index=cyto_df.index, columns=cyto_df.columns
    )
    assert imputed_df.shape == cyto_df.shape
    assert not imputed_df.isnull().any().any()
    return imputed_df


def binarize_donor_scores(donor_scores_df, fraction_responders=None, fold_change=1.2,
                          save_fig=False, show_fig=False, fig_dir=None):
    assert fraction_responders is not None or fold_change is not None
    assert fraction_responders is None or fold_change is None
    assert not save_fig or fig_dir is not None

    if fraction_responders is not None:
        assert fold_change is None
        responders_df = binarize(donor_scores_df,
                                 negative=(0.0, 1 - fraction_responders),
                                 positive=(1 - fraction_responders, 1.0))
    else:
        assert fold_change is not None
        responders_df = donor_scores_df >= np.log2(fold_change)
        responders_df = responders_df.mask(donor_scores_df.isnull())
    assert responders_df.isnull().equals(donor_scores_df.isnull())
    assert responders_df.isin([0, 1, np.nan]).all().all()
    responders_df = responders_df.mask(responders_df == 1, 'R')
    responders_df = responders_df.mask(responders_df == 0, 'N')
    real_fractions = (responders_df == 'R').sum() / responders_df.shape[0]
    print('Responders (fraction {}, fold_change {}):\n{}'.format(
        fraction_responders, fold_change, real_fractions))

    if save_fig or show_fig:
        fig, axs = plt.subplots(1, responders_df.shape[1], figsize=(4 * responders_df.shape[1], 4))
        for ax, col in zip(axs, responders_df.columns):
            ax = sns.distplot(donor_scores_df[col], kde=False, ax=ax)
            min_resp_l2fc = donor_scores_df[col].loc[responders_df[col] == 'R'].min()
            ax.axvline(min_resp_l2fc)
            ax.set_ylabel('Frequency')
            ax.set_xlabel(col)
            title = 'Resp. {} ~ {}'.format(
                'fraction {}'.format(fraction_responders) if fraction_responders else 'FC {}'.format(fold_change),
                'FC {} {:.2f}'.format(GREATER_EQUAL, np.power(2, min_resp_l2fc)) if fraction_responders else 'fraction of {:.2f}'.format(real_fractions.loc[col])
            )
            ax.set_title(title)
            sns.despine()
        if save_fig:
            savefig(os.path.join(fig_dir, 'BCG.responders_{}.{}.pdf'.format('frac{:g}'.format(fraction_responders) if fraction_responders else 'FC{:g}'.format(fold_change), col)))
        if show_fig:
            plt.show()

    responders_df.columns += '_{}_responder'.format('frac{:g}'.format(fraction_responders) if fraction_responders else 'FC{:g}'.format(fold_change))
    return responders_df


def get_fold_change_cytogroups_masks(index, which=r'^CYTO:.*_good$'):

    return {
        'adaptive_MTB_7d':
            index.str.contains(which) &
            index.str.contains(':MTB_') &
            index.str.contains('_7d_'),
        'adaptive_7d':
            index.str.contains(which) &
            index.str.contains('_7d_'),
        'IFNg_MTB_7d':
            index.str.contains(which) &
            index.str.contains(':MTB_') &
            index.str.contains('_7d_') &
            index.str.contains('_IFNg'),
        'heterologous_nonspecific_7d':
            index.str.contains(which) &
            ~index.str.contains(':MTB_') &
            index.str.contains('_7d_'),
        'innate_MTB_24h_wo_LAC':
            index.str.contains(which) &
            index.str.contains(':MTB_') &
            index.str.contains('_24h_') &
            ~index.str.contains('_IFNg|_IL\.17|_lactate'),
        'innate_nonspecific_24h':
            index.str.contains(which) &
            ~index.str.contains(':MTB_') &
            index.str.contains('_24h_') &
            ~index.str.contains('_IFNg|_IL\.17'),
        'innate_nonspecific_24h_wo_LAC':
            index.str.contains(which) &
            ~index.str.contains(':MTB_') &
            index.str.contains('_24h_') &
            ~index.str.contains('_IFNg|_IL\.17|_lactate'),
        'innate_nonspecific_24h_wo_LAC_IL10_IL1ra':
            index.str.contains(which) &
            ~index.str.contains(':MTB_') &
            index.str.contains('_24h_') &
            ~index.str.contains('_IFNg|_IL\.17|_lactate|_IL\.1ra|_IL\.10'),
        'lactate':
            index.str.contains(which) &
            index.str.contains('_lactate'),
        'innate_nonspecific_24h_IL10_IL1ra':
            index.str.contains(which) &
            ~index.str.contains(':MTB_') &
            index.str.contains('_24h_') &
            index.str.contains('_IL\.1ra|_IL\.10'),
        'innate_24h_wo_LAC':
            index.str.contains(which) &
            index.str.contains('_24h_') &
            ~index.str.contains('_IFNg|_IL\.17|_lactate'),
        'innate_24h_wo_LAC_IL10_IL1ra':
            index.str.contains(which) &
            index.str.contains('_24h_') &
            ~index.str.contains('_IFNg|_IL\.17|_lactate|_IL\.1ra|_IL\.10'),
        'innate_24h_wo_LAC_IL12':
            index.str.contains(which) &
            index.str.contains('_24h_') &
            ~index.str.contains('_IFNg|_IL\.17|_IL\.12|_lactate'),
        'innate_incl_LAC':
            index.str.contains(which) &
            index.str.contains('_24h_') &
            ~index.str.contains('_IFNg|_IL\.17'),
    }


def get_V1_cytogroups_masks(index, which=r'^CYTO:.*_good$'):

    return {
        'adaptive_7d':
            index.str.contains(which) &
            index.str.contains('_7d_'),

        'innate_24h_pro_inflammatory':
            index.str.contains(which) &
            index.str.contains('_24h_') &
            ~index.str.contains('_IFNg|_IL\.17|_IL\.1ra|_IL\.10|_lactate'),

        'innate_24h_anti_inflammatory':
            index.str.contains(which) &
            index.str.contains('_24h_') &
            index.str.contains('_IL\.1ra|_IL\.10'),

        'innate_24h_wo_LAC':
            index.str.contains(which) &
            index.str.contains('_24h_') &
            ~index.str.contains('_IFNg|_IL\.17|_lactate'),
    }


def make_donor_scores(cyto_df, visit, X_df, design, do_not_correct=['Intercept', 'SAMPLE_DONOR', 'SAMPLE_VISIT'],
                      cytokine_filter=None,
                      remove_V1_production=False, standard_scaling=False,
                      methods='top_half_mean', cytogroup=None, return_masked_LFCs=False, cytokine_renaming=True,
                      max_impute_iter=10, padj_method='fdr_bh', fdr=0.05, random_state=RANDOM_STATE,
                      save_fig=False, show_fig=False, fig_dir=None, verbose=1):
    if isinstance(methods, str):
        methods = [methods]
    assert 'non_negative_trimmed_mean' not in methods
    if isinstance(cytokine_filter, str):
        cytokine_filter = [cytokine_filter]
    assert not cytokine_filter or all([f in ['positive', 'fdr'] for f in cytokine_filter])
    assert all([m in ['mean', 'mean_without_worst', 'top_half_mean', 'top_half_floor_mean', 'trimmed_mean', 'non_negative_trimmed_mean'] for m in methods])
    assert not save_fig or fig_dir is not None
    assert X_df is None or cyto_df.index.equals(X_df.index)

    if set(do_not_correct) not in [set(['SAMPLE_VISIT']), set(['Intercept', 'SAMPLE_DONOR', 'SAMPLE_VISIT']), set(['Intercept', 'SAMPLE_VISIT']), set(['SAMPLE_DONOR', 'SAMPLE_VISIT'])]:
        print('\n C A R E F U L\n do_not_correct = {}\n'.format(do_not_correct))

    cyto_df = impute_cytokines(cyto_df, max_iter=max_impute_iter, random_state=random_state)
    cyto_df.sort_index(axis=1).to_csv(IMPUTED_CYTO_FN)

    if design is not None:
        X_df, cyto_df = drop_na_for_LM(X_df.loc[cyto_df.index], cyto_df, verbose=verbose)
        _, results_df, cyto_df, _ = fit_linear_model(X_df, cyto_df, design=design, do_not_correct=do_not_correct,
                                                  return_corrected_X=True, random_state=RANDOM_STATE, verbose=verbose)
        results_df = results_df.loc[results_df.index.get_level_values('contrast') == 'SAMPLE_VISIT[T.{}]'.format(visit)].droplevel('contrast')
        assert not cyto_df.isnull().any().any()


        cyto_mask = np.ones((results_df.shape[0],), dtype=bool)
        if cytokine_filter and 'positive' in cytokine_filter:
            cyto_mask &= (results_df['Coef'] > 0)
        if cytokine_filter and 'fdr' in cytokine_filter:
            padj = adjusted_pvals(results_df['p.value'], method=padj_method)
            cyto_mask &= (padj < fdr)
        cyto_df = cyto_df.loc[:, results_df.index[cyto_mask]]

    cyto_df.index = cyto_df.index.str.split('_', expand=True)
    cyto_df.index.names = ['SAMPLE:DONOR', 'SAMPLE:VISIT', 'SAMPLE:TISSUE']
    lfc_df = visit_diff(cyto_df, visit=visit, base_visit='V1')
    lfc_df = lfc_df.loc[~lfc_df.isnull().all(axis=1)]

    if remove_V1_production:
        assert (cyto_df.index.get_level_values('SAMPLE:TISSUE') == 'PBMC').all()
        cyto_df = cyto_df.loc[cyto_df.index.get_level_values('SAMPLE:VISIT') == 'V1']
        cyto_df.index = cyto_df.index.get_level_values('SAMPLE:DONOR')
        lfc_df.columns = 'LFC_' + lfc_df.columns
        corr_lfc_df = []
        for cyto in cyto_df.columns:
            _, _, _corr_lfc_df, _ = fit_linear_model(
                cyto_df.loc[lfc_df.index][[cyto]], lfc_df[['LFC_{}'.format(cyto)]], design='1 + {}'.format(replace_chars_for_LM(cyto)),
                do_not_correct=['Intercept'], return_corrected_X=True, random_state=RANDOM_STATE, verbose=verbose)
            corr_lfc_df.append(_corr_lfc_df)
        lfc_df = pd.concat(corr_lfc_df, axis=1)
        lfc_df.columns = lfc_df.columns.str.replace('LFC_', '')

    if standard_scaling:
        lfc_df = pd.DataFrame(StandardScaler().fit_transform(lfc_df), index=lfc_df.index, columns=lfc_df.columns)

    cytogroups_masks = get_fold_change_cytogroups_masks(lfc_df.columns)
    donor_scores_df = pd.DataFrame()
    for cytogroup in cytogroups_masks.keys() if cytogroup is None else [cytogroup]:
        _lfc_df = lfc_df.loc[:, cytogroups_masks[cytogroup]]
        if cytokine_renaming:
            _lfc_df.columns = rename_cytokines(_lfc_df.columns)
        for method in methods:
            if method == 'non_negative_trimmed_mean':
                assert False
                _df = _lfc_df.where(_lfc_df >= 0, other=0).copy()
            else:
                _df = _lfc_df.copy()
            # if standard_scaling:
            #     _df = pd.DataFrame(StandardScaler().fit_transform(_df), index=_df.index, columns=_df.columns)
            if method == 'mean':
                pass
            elif method == 'mean_without_worst':
                if _df.shape[1] > 1:
                    _ranks = _df.rank(ascending=False, axis=1, method='first')
                    _df = _df.mask(_ranks == _ranks.max())
            elif method == 'top_half_mean':
                _top_half = _df.rank(ascending=False, axis=1, method='first') <= np.ceil(_df.shape[1] / 2)
                assert _top_half.equals(_df.rank(ascending=False, axis=1) <= np.ceil(_df.shape[1] / 2))
                assert _top_half.equals(_df >= _df.median(axis=1).values.reshape((-1, 1)))
                _df = _df.where(_top_half)
            elif method == 'top_half_floor_mean':
                _top_half = _df.rank(ascending=False, axis=1, method='first') <= np.floor(_df.shape[1] / 2)
                _df = _df.where(_top_half)
            elif method in ['trimmed_mean', 'non_negative_trimmed_mean']:
                if _df.shape[1] > 2:
                    _ranks = _df.rank(ascending=False, axis=1, method='first')
                    _df = _df.mask((_ranks == _ranks.min()) | (_ranks == _ranks.max()))
            else:
                raise ValueError

            if return_masked_LFCs:
                return _df

            donor_scores = _df.mean(axis=1).rename('{}.{}_{}'.format(method, cytogroup, visit))
            donor_scores_df = donor_scores_df.append(donor_scores)

            if save_fig or show_fig:
                cg = clustermap(_lfc_df.iloc[np.argsort(donor_scores)], figsize=(4, 3.5),
                                row_cluster=False, col_cluster=_lfc_df.shape[1] > 1, xticklabels=True, yticklabels=False,
                                z_score=None, cmap=PALETTE_CODES['diverging'], center=0, robust=True)
                cg.fig.suptitle(donor_scores.name, y=1.1)
                if save_fig:
                    savefig(os.path.join(fig_dir, 'BCG.donor_scores.{}.pdf'.format(donor_scores.name)))
                if show_fig:
                    plt.show()

    return donor_scores_df.T, lfc_df


def rename_celltype(celltype, plural=True, capitalize=False):
    celltype = 'NK cell' if celltype == 'nkcell' else \
        'monocyte' if celltype == 'monocyte' else \
            'CD8+ T cell' if celltype == 'cd8t' else \
                'PBMC' if celltype == 'PBMC' else \
                    celltype

    if plural:
        celltype = '{}s'.format(celltype)

    if capitalize:
        celltype = '{}{}'.format(celltype[0].upper(), celltype[1:])

    return celltype


def fix_annot_names(annot_df):
    annot_df.columns = annot_df.columns.str.replace(
        '^CIRCAD_REPLIC$', 'DONOR:CIRCAD_REPLIC').str.replace(
        '^SAMPLE:IC_DATE_2PI$', 'DONOR:IC_DATE_2PI').str.replace(
        '^DONOR:VISIT_DATE_REAL$', 'SAMPLE:VISIT_DATE_REAL').str.replace(
        '^DONOR:VISIT_MONTH_REAL$', 'SAMPLE:VISIT_MONTH_REAL').str.replace(
        '^DONOR:VISIT_TIME_REAL$', 'SAMPLE:VISIT_TIME_REAL').str.replace(
        '^DONOR:VISIT_TIME$', 'SAMPLE:VISIT_TIME')
    return annot_df


def get_deseq_norm_counts(celltype, model='all_samples.batch.sex.age{blood}.TSS_enr'):
    if '{blood}' in model:
        model = model.format(blood='.blood' if celltype == 'PBMC' else '')
    fn = DESEQ_DATA_FN.format(celltype=celltype, model=model, data='norm')
    print(fn)
    return pd.read_csv(fn, index_col=0)


def get_vst_norm_counts(celltype, model='all_samples.batch.sex.age{blood}.TSS_enr', blind=False):
    if '{blood}' in model:
        model = model.format(blood='.blood' if celltype == 'PBMC' else '')
    fn = DESEQ_DATA_FN.format(celltype=celltype, model=model, data='{}vst'.format('blind_' if blind else ''))
    print(fn)
    return pd.read_csv(fn, index_col=0)


def get_vst_batch_corrected_counts(celltype, model='all_samples.batch.sex.age{blood}.TSS_enr', blind=False):
    if '{blood}' in model:
        model = model.format(blood='.blood' if celltype == 'PBMC' else '')
    fn = DESEQ_DATA_FN.format(celltype=celltype, model=model,
                              data='{}vst_batch_corrected'.format('blind_' if blind else ''))
    print(fn)
    return pd.read_csv(fn, index_col=0)


def pval_col(coef=None):
    return 'p.value{}'.format('.{}'.format(coef) if coef else '')


def stat_col(coef=None, F_test=False):
    return '{}{}'.format('F' if F_test else 't', '.{}'.format(coef) if coef else '')


def coef_col(coef=None):
    return 'Coef{}'.format('.{}'.format(coef) if coef else '')


def padj_col(coef=None):
    return 'padj{}'.format('.{}'.format(coef) if coef else '')


def lfc_times_pval_col(coef):
    return 'Coef_times_p.value.{coef}'.format(coef=coef)


def lfc_times_neg_log10_pval_col(coef):
    return 'Coef_times_neg_log10_p.value.{coef}'.format(coef=coef)


def is_dream_LMM(model):
    return model.startswith('LMM.') or '.LMM.' in model


def read_dream_results(celltype, model, contrasts, results_dir='results', extras=None):
    if isinstance(contrasts, str):
        contrasts = [contrasts]

    de_df = pd.DataFrame()
    for coef in contrasts:
        df = pd.read_csv(de_fn(celltype, model, data='results_{}'.format(coef), results_dir=results_dir, extras=extras),
                         index_col=0)
        de_df = pd.concat(
            [de_df, df[['logFC', 'P.Value']].rename({'logFC': coef_col(coef), 'P.Value': pval_col(coef)}, axis=1)],
            axis=1
        )
    de_df['A'] = df['AveExpr']

    return de_df


def read_de(celltype, model, contrasts=None, data='results_p5', F_test=False, adjust='fdr_bh', annot_fn=None, peaks_df=None,
            swap_to_non_resp=False, results_dir='results', extras=None, legacy=False,
            add_coef_times_pval=False, effect_size_filter=None):
    if isinstance(contrasts, str):
        contrasts = [contrasts]

    if contrasts:
        df_cols = np.asarray(
            [[coef_col(c) if not F_test else None, stat_col(c, F_test=F_test), pval_col(c)] for c in contrasts]
        ).flatten()
        df_cols = ['A'] + df_cols[~pd.isnull(df_cols)].tolist()

    if is_dream_LMM(model):
        assert contrasts
        df = read_dream_results(celltype, model, contrasts, results_dir=results_dir, extras=extras)
    else:
        fn = de_fn(celltype, model, data=data, results_dir=results_dir, extras=extras, legacy=legacy)
        df = pd.read_csv(fn, usecols=(['genes'] + df_cols) if contrasts else None).set_index('genes')
    if contrasts:
        df = df[df_cols]

    if adjust:
        assert contrasts
        df = pd.concat([df] + [adjusted_pvals(df, c, method=adjust) for c in contrasts], axis=1)

    if effect_size_filter:
        assert contrasts
        for c in contrasts:
            mask = df[coef_col(c)].abs() >= effect_size_filter
            coef_cols = sorted(set(df.columns).intersection(
                [coef_col(c), pval_col(c), padj_col(c), 't.{}'.format(c), 'Res.{}'.format(c)]))
            df.loc[mask, coef_cols] = np.nan

    if swap_to_non_resp:
        assert contrasts
        for c in contrasts:
            if c.endswith('_V3_FC1.2_R'):
                for col in [coef_col(c) if not F_test else None, stat_col(c, F_test=F_test)]:
                    if col is not None:
                        df[col] *= -1
            elif c.endswith('_R'):
                raise ValueError('Unexpected: cannot determine if to switch responder phenotypeor not')

    if add_coef_times_pval:
        for c in contrasts:
            df[lfc_times_pval_col(c)] = df[coef_col(c)] * (1 - df[pval_col(c)])
            df[lfc_times_neg_log10_pval_col(c)] = df[coef_col(c)] * (-np.log10(df[pval_col(c)]))

    if annot_fn:
        assert peaks_df is None
        peaks_df = get_peak_annot(annot_fn)

    if peaks_df is not None:
        df = pd.concat([df, peaks_df.loc[df.index]], axis=1)

    return df


def read_de_samples(celltype, model, legacy=False):
    return pd.read_csv(de_fn(celltype=celltype, model=model, data='counts', legacy=legacy), index_col=0,
                       nrows=1).columns


def get_de_effective_N(celltype, model, visit1, visit2):
    samples = read_de_samples(celltype, model)
    samples = samples.str.split('_', expand=True).to_frame()
    effective_n = (samples.loc[samples[1].isin([visit1, visit2])].groupby([0]).count()[1] == 2).sum()
    return effective_n


def volcano(df, contrasts=None, pval_thr=None, fdr=None, fdr_with_hue=False, fdr_prefix='', rasterized=True, dpi=300, figsize=(5, 4.5),
            hue='A', hue_order=None, style=None, style_order=None,
            palette=BLUE_GREEN_CMAP, fdr_pass_color=BLUE, fdr_fail_color=LIGHT_GRAY,
            sharex=False, sharey=False, centre=True, alpha=0.5, size=60, xlabel='Log2 fold change', ylabel='{}log10{}P-value'.format(MINUS, NBSPC),
            legend=True, auto_sort=False, on_top=None, on_top_shuffle=False, on_top_annot=None, random_state=RANDOM_STATE, xticks=None,
            legend_kwargs=dict(bbox_to_anchor=(1, 1), title='Average\nlog$_2$CPM'),
            scatter_kwargs={}, ylim=None, xlim_scale=None, title=None,
            fig_fn=None):
    if fdr and not is_iterable(fdr):
        fdr = [fdr]
    if isinstance(contrasts, str):
        contrasts = [contrasts]
    elif contrasts is None:
        contrasts = df.columns[df.columns.str.startswith('p.value')].str.replace('^p\.value\.', '').tolist()
        print('Found the following contrasts:', contrasts)

    fig, axs = plt.subplots(1, len(contrasts), figsize=(len(contrasts) * figsize[0], figsize[1]),
                            sharex=sharex, sharey=sharey, squeeze=False)
    fig.subplots_adjust(wspace=0.4)

    for ax, coef in zip(axs[0], contrasts):
        if hue or fdr_with_hue:
            if fdr and fdr_with_hue:
                assert len(fdr) == 1
                fdr_thr = fdr[0]
                fdr_hue = df[padj_col(coef)] <= fdr_thr
                fdr_hue_order = np.unique(fdr_hue)[::-1]
                if len(fdr_hue_order) > 1:
                    _palette = [fdr_pass_color[coef] if isinstance(fdr_pass_color, dict) else fdr_pass_color, fdr_fail_color]
                else:
                    _palette = [(fdr_pass_color[coef] if isinstance(fdr_pass_color, dict) else fdr_pass_color) if fdr_hue_order[0] else fdr_fail_color]

                _hue = nans_like(fdr_hue).astype(str)
                _hue[fdr_hue] = '{}{} {}'.format(fdr_prefix, LESS_EQUAL, fdr_thr)
                _hue[~fdr_hue] = '{}> {}'.format(fdr_prefix, fdr_thr)
                _hue_order = nans_like(fdr_hue_order).astype(str)
                _hue_order[fdr_hue_order] = '{}{} {}'.format(fdr_prefix, LESS_EQUAL, fdr_thr)
                _hue_order[~fdr_hue_order] = '{}> {}'.format(fdr_prefix, fdr_thr)

                if hue:
                    assert hue in df.columns
                    assert palette is not None
                    _hue = np.copy(df[hue].values)
                    _hue[~fdr_hue] = '{}> {}'.format(fdr_prefix, fdr_thr)
                    if hue_order is not None:
                        _hue_order = list(hue_order) + ['{}> {}'.format(fdr_prefix, fdr_thr)]
                    else:
                        _hue_order = list(np.unique(_hue))
                    _palette = sns.color_palette(palette, n_colors=len(_hue_order))
                    if fdr_fail_color not in _palette:
                        _palette[-1] = fdr_fail_color

            elif hue in df.columns:
                _hue = df[hue].values
                _hue_order = hue_order
                _palette = palette

            else:
                _hue = df['{}.{}'.format(hue, coef)].abs().values
                _hue_order = hue_order
                _palette = palette

        if auto_sort:
            sort = np.argsort(-np.log10(df[pval_col(coef)]))
            not_pass_fdr = df[padj_col(coef)].values[sort] > fdr_thr
            sort[not_pass_fdr] = np.random.RandomState(random_state).choice(sort[not_pass_fdr], size=np.sum(not_pass_fdr), replace=False)
            if auto_sort == 'two_stage_random':
                sort[~not_pass_fdr] = np.random.RandomState(random_state).choice(sort[~not_pass_fdr], size=np.sum(~not_pass_fdr), replace=False)
        else:
            sort = np.arange(df.shape[0])

        if on_top is not None:
            if on_top_annot is not None:
                assert on_top_annot.isin(on_top).all(), (on_top_annot, _on_top)
                on_top = on_top[~on_top.isin(on_top_annot)]
            _on_top = df.reset_index().loc[df.index.isin(on_top)].index.values
            _on_top = _on_top[np.argsort(df.iloc[_on_top][pval_col(coef)])[::-1]]
            if on_top_shuffle:
                np.random.RandomState(0).shuffle(_on_top)
            if on_top_annot is not None:
                _on_top_annot = df.reset_index().loc[df.index.isin(on_top_annot)].index.values
                _on_top_annot = _on_top_annot[np.argsort(df.iloc[_on_top_annot][pval_col(coef)])[::-1]]
            else:
                _on_top_annot = np.asarray([], dtype=int)
            sort = np.concatenate([sort[~np.isin(sort, _on_top) & ~np.isin(sort, _on_top_annot)], _on_top, _on_top_annot])

        assert len(sort) == len(np.unique(sort))
        ax = sns.scatterplot(x=df[coef_col(coef)].values[sort], y=-np.log10(df[pval_col(coef)])[sort],
                             hue=_hue[sort] if hue or fdr_with_hue else None, hue_order=_hue_order,
                             style=df[style] if isinstance(style, str) else style, style_order=style_order,
                             palette=_palette, s=size, alpha=alpha, ax=ax, rasterized=rasterized,
                             **(scatter_kwargs if coef not in scatter_kwargs else scatter_kwargs[coef]))
        if fdr and not fdr_with_hue:
            _deep_pal = sns.color_palette('deep')
            _red = _deep_pal[3]
            del _deep_pal[3]
            _deep_pal = [_red] + _deep_pal
            fdr_colors = _deep_pal[:len(fdr)] if fdr[0] < 0.05 else _deep_pal[1:len(fdr) + 1]
            if len(fdr) == 1:
                fdr_colors = ['black']
            for fdr_thr, fdr_color in zip(fdr, fdr_colors):
                _fdr_to_pval = fdr_to_pval(df, fdr=fdr_thr, fdr_col=padj_col(coef), pval_col=pval_col(coef), ascending=True)
                ax.axhline(-np.log10(_fdr_to_pval), c=fdr_color, ls='--', label='FDR {} {}'.format(LESS_EQUAL, fdr_thr))
        if pval_thr is not None:
            ax.axhline(-np.log10(pval_thr), c=GRAY, ls='--')
            ax.annotate('P {} {}'.format(LESS_EQUAL, pval_thr), (ax.get_xlim()[1], -np.log10(pval_thr)), fontsize=SMALL_FONT)

        ax.tick_params(reset=True, top=False, right=False)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel).set_visible(True)
        ax.set_title(coef if title is None else title)
        if ylim:
            if is_iterable(ylim):
                ax.set_ylim(ylim)
            else:
                _ylim, _ = ax.get_ylim()
                ax.set_ylim((_ylim, ylim))
        if xlim_scale:
            xlim = ax.get_xlim()
            ax.set_xlim((xlim[0] * xlim_scale, xlim[1] * xlim_scale))
        if centre:
            _xlim = np.max(np.abs(ax.get_xlim()))
            ax.set_xlim((-_xlim, _xlim))
        if xticks:
            ax.set_xticks(xticks)
        ax.legend().set_visible(False)
        sns.despine(ax=ax)

    if legend:
        # if hue is None:
        #     if 'title' in legend_kwargs:
        #         del legend_kwargs['title']
        if fdr and not fdr_with_hue:
            handles, labels = ax.get_legend_handles_labels()
            ax.legend(handles[len(fdr):] + handles[:len(fdr)], labels[len(fdr):] + labels[:len(fdr)], **legend_kwargs)
        else:
            ax.legend(**legend_kwargs)

    if fig_fn:
        savefig(fig_fn, dpi=dpi)
    return axs[0]


def find_top_hits(de_df, coef, fdr, show_n, unique_genes=False, highlights=None, max_highlights=None, region_filter=None,
                  model=None, celltype=None, top_n=None, library=None, enr_fdr=None, enr_region_filter=None,
                  results_dir='results'):
    if enr_region_filter is None:
        enr_region_filter = region_filter

    if region_filter:
        de_df = de_df.loc[peaks_to_genes(de_df, **PEAKS_TO_GENES[region_filter])]

    _highlights = []
    if highlights:
        for gene in highlights.get(coef, []):
            _h = de_df.loc[(de_df[padj_col(coef)] <= fdr) & (de_df['gene_name'] == gene)]
            if unique_genes:
                _h = _h.loc[_h.groupby('gene_name')[pval_col(coef)].idxmin()]
            _h = _h.index.tolist()
            if len(_h) == 0:
                _h = de_df.loc[(de_df['gene_name'] == gene) & (de_df[coef_col(coef)] > 0)].sort_values(pval_col(coef)).head(1).index.tolist() + \
                     de_df.loc[(de_df['gene_name'] == gene) & (de_df[coef_col(coef)] < 0)].sort_values(pval_col(coef)).head(1).index.tolist()
            _highlights.extend(_h)

    if max_highlights:
        _h_df = de_df.loc[_highlights]
        _h_df = pd.concat(
            [_h_df.loc[de_df[coef_col(coef)] * d > 0].sort_values(pval_col(coef)).head(max_highlights) for d in [1, -1]]
        )
        _highlights = _h_df.index.tolist()


    top_hits = {}
    pathways = {}

    def _filter_hits(hits):
        if unique_genes:
            hits = hits.loc[hits.groupby('gene_name')[pval_col(coef)].idxmin()]
        hits = hits.sort_values(pval_col(coef)).head(show_n).index.tolist()
        return hits

    if not model:
        for direction in [-1, 1]:
            top_hits[direction] = _filter_hits(de_df.loc[(de_df[padj_col(coef)] <= fdr) &
                                                         (de_df[coef_col(coef)] * direction > 0)])
    else:
        if isinstance(library, str):
            library = [library]
        for direction in [-1, 1]:
            enr_df = read_enrichr_or_lola(celltype, model, coef=coef, library=library, effect_size_filter=None,
                                      rank_metric='p.value', top_n=top_n, region_filter=region_filter,
                                      method='enrichr', direction=direction, results_dir=results_dir)

            if region_filter == enr_region_filter:
                enr_df = enr_df.loc[enr_df['Adjusted P-value'] <= enr_fdr]
            else:
                _df = read_enrichr_or_lola(celltype, model, coef=coef, library=library, effect_size_filter=None,
                                          rank_metric='p.value', top_n=top_n, region_filter=enr_region_filter,
                                          method='enrichr', direction=direction, results_dir=results_dir)
                _df = _df.loc[_df['Adjusted P-value'] <= enr_fdr]
                enr_df = enr_df.loc[enr_df.index.intersection(_df.index)]

            enr_genes = np.unique([g for genes in enr_df['Genes'].values for g in genes.split(';')])
            top_hits[direction] = _filter_hits(de_df.loc[(de_df[padj_col(coef)] <= fdr) &
                                                        de_df.index.isin(enr_genes)])
            for region in top_hits[direction]:
                if enr_df['Genes'].str.contains(region).any():
                    pathways[region] = ' | '.join((enr_df.loc[enr_df['Genes'].str.contains(region)].index.values))

    annot_df = de_df.loc[set(top_hits[1] + top_hits[-1] + _highlights)].sort_values(pval_col(coef))
    annot_df['pathways'] = None
    for region in pathways:
        assert region in annot_df.index
        annot_df.loc[region, 'pathways'] = pathways[region]

    return annot_df


def annotate_volcano(ax, annot_df, coef):
    for idx in annot_df.index:
        ax.annotate(annot_df.loc[idx, 'gene_name'],
                    (annot_df.loc[idx, coef_col(coef)], -np.log10(annot_df.loc[idx, pval_col(coef)])),
                    fontsize=SMALL_FONT, style='italic')


def plot_de_agreement(df, baseline_df, contrasts, baseline_contrasts,
                      plot_lfc=True, plot_pval=True, plot_rank_lfc=True, plot_rank_pval=True,
                      baseline_name=None, figsize=(5, 4.5), rasterized=True, scatter_kwargs={}):
    if 's' not in scatter_kwargs:
        scatter_kwargs['s'] = 15
    if 'alpha' not in scatter_kwargs:
        scatter_kwargs['alpha'] = 0.7
    lfc_getter = lambda _df, _coef: _df[coef_col(_coef)]
    lfc_label = 'log$_2$ fold change'
    pval_label = '$-$log$_{10}$ pval'
    plot_cols = np.asarray([(lfc_label, lfc_getter, False), (pval_label, signed_log10_pvals, False),
                            (lfc_label, lfc_getter, True), (pval_label, signed_log10_pvals, True)])
    _plot_mask = np.asarray([plot_lfc, plot_pval, plot_rank_lfc, plot_rank_pval])
    plot_cols = plot_cols[_plot_mask]
    fig, axs = plt.subplots(len(plot_cols), len(contrasts),
                            figsize=(figsize[0] * len(contrasts), figsize[1] * len(plot_cols)),
                            squeeze=False)
    plt.subplots_adjust(wspace=0.6, hspace=0.8 if baseline_name else 0.5)
    for row, (label, getter, rank) in enumerate(plot_cols):
        for col, (coef, baseline_coef) in enumerate(zip(contrasts, baseline_contrasts)):
            x, y = getter(baseline_df, baseline_coef), getter(df, coef)
            if rank:
                x, y = x.rank(ascending=False), y.rank(ascending=False)
            ax = sns.scatterplot(x=x, y=y, ax=axs[row, col], rasterized=rasterized, **scatter_kwargs)
            ax.set_title('{}\n(Spearman R {:.3f})'.format(coef, spearmanr(x, y)[0]))
            ax.set_xlabel('{}{}{}'.format('Rank of ' if rank else '', label,
                                          '\n({})'.format(baseline_name) if baseline_name else ''))
            ax.set_ylabel('{}{}'.format('Rank of ' if rank else '', label))
            sns.despine()
    return axs


def de_overlap(df, baseline_df, coef, baseline_coef):
    overlap = {}
    for direction in [1, -1]:
        _df1 = baseline_df.loc[baseline_df[coef_col(baseline_coef)] * direction > 0].sort_values(pval_col(baseline_coef))
        _df2 = df.loc[df[coef_col(coef)] * direction > 0].sort_values(pval_col(coef))
        _overlaps, regions1, regions2 = [], set([]), set([])
        for i in range(max(len(_df1), len(_df2))):
            if i < len(_df1):
                regions1.add(_df1.index[i])
            if i < len(_df2):
                regions2.add(_df2.index[i])
            _overlaps.append(len(regions1.intersection(regions2)))
        overlap[direction] = np.asarray(_overlaps)
    return overlap


def plot_de_overlap(overlap, figsize=(5, 4.5), fig_fn=None):
    fig, axs = plt.subplots(1, len(overlap), figsize=(len(overlap) * figsize[0], figsize[1]),
                            sharex=False, sharey='row', squeeze=False)
    fig.subplots_adjust(wspace=0.4)
    for ax, coef in zip(axs[0], overlap):
        for direction in overlap[coef]:
            _overlap = overlap[coef][direction]
            _size = list(range(1, len(_overlap) + 1))
            ax.plot(_size, _overlap / _size,
                    label='Up' if direction == 1 else 'Down' if direction == -1 else direction)
        ax.set_xscale('log')
        ax.tick_params(reset=True, top=False, right=False)
        ax.set_xlabel('Top $n$ hits')
        ax.set_ylabel('Fraction overlapping')
        ax.set_title(coef)
        legend = ax.legend(bbox_to_anchor=(1, 1), title='Fold change')
        legend.set_visible(False)
        sns.despine()
    legend.set_visible(True)

    if fig_fn:
        savefig(fig_fn)
    return axs[0]


def _get_rank_order_args(rank_metric, coef, direction=None):
    assert direction in [None, -1, 1]
    if rank_metric == 'p.value':
        _sort_col, _ascending = pval_col(coef), True
    elif rank_metric == 'Coef':
        _sort_col, _ascending = coef_col(coef), (direction == -1)
    elif rank_metric == 'Coef_times_p.value':
        _sort_col, _ascending = lfc_times_pval_col(coef), (direction == -1)
    elif rank_metric == 'Coef_times_neg_log10_p.value':
        _sort_col, _ascending = lfc_times_neg_log10_pval_col(coef), (direction == -1)
    elif rank_metric == 'FVE':
        _sort_col, _ascending = 'FVE', False
    else:
        raise ValueError
    return _sort_col, _ascending


def get_direction_mask(df, coef, direction):
    assert direction in [-1, 1]
    return df[coef_col(coef)] * direction > 0


def get_top_n_regions(df, coef, direction, rank_metric, top_n):
    assert direction in [None, -1, 1]
    _lfc_mask = get_direction_mask(df, coef, direction) if direction else np.ones((df.shape[0],), dtype=bool)
    # checking that selecting for direction gets rid of NaNs which are there because of effect_size_filter
    if pval_col(coef) in df:
        assert not df.loc[_lfc_mask, pval_col(coef)].isnull().any()
    if coef_col(coef) in df:
        assert not df.loc[_lfc_mask, coef_col(coef)].isnull().any()
    if top_n >= 1:
        assert rank_metric in ['p.value', 'Coef', 'Coef_times_p.value', 'Coef_times_neg_log10_p.value', 'FVE']
        assert direction or rank_metric == 'p.value' or rank_metric == 'FVE'
        _sort_col, _ascending = _get_rank_order_args(rank_metric, coef, direction)
        regions = df.loc[_lfc_mask].sort_values(_sort_col, ascending=_ascending).head(top_n).index
    else:
        assert rank_metric == 'p.value'
        regions = df.loc[_lfc_mask & (df[padj_col(coef)] < top_n)].sort_values(pval_col(coef)).index
    return regions


def get_top_n_regions_both_directions(de_df, coef, top_n, rank_metric, region_filter, peaks_df=None):
    if region_filter is not None:
        de_df = de_df.loc[peaks_to_genes(de_df if peaks_df is None else peaks_df.loc[de_df.index],
                                         **PEAKS_TO_GENES[region_filter])]
    if top_n is not None:
        regions = pd.Index(np.concatenate(
            [get_top_n_regions(de_df, coef=coef, direction=d, rank_metric=rank_metric, top_n=top_n) for d in [-1, 1]]
        ))
    else:
        regions = de_df.index
    return regions


def de_heatmaps(atac_df, annot_df, de_df, coef, rank_metric, top_n, annotations=None, palettes=None,
                pseudo_sort=False, pseudosort_col=None, pseudosort_base_value=None,
                pseudosort_hist_height=0.75, pseudosort_hist_fn=None,
                rasterized=True, legend_bbox='auto', figsize=(8, 8), fig_fn=None, dpi=300, verbose=True):
    if isinstance(annotations, str):
        annotations = [annotations]
    if isinstance(palettes, str):
        palettes = [palettes]

    _directions = {-1: 'Down', 1: 'Up'}
    if pseudo_sort:
        sort, regions = {}, {}
        for direction in _directions.keys():
            _regions = get_top_n_regions(de_df, coef=coef, direction=direction, rank_metric=rank_metric, top_n=top_n)
            if len(_regions) > 0:
                assert coef in ['V3', 'V2']
                mean1 = atac_df.loc[annot_df[pseudosort_col] == pseudosort_base_value, _regions].mean()
                mean2 = atac_df.loc[annot_df[pseudosort_col] == coef, _regions].mean()
                mean_diff = mean2 - mean1
                distances = (atac_df[_regions] - mean_diff)
                _sort = np.argsort(distances.mean(axis=1))
                if direction < 0:
                    _sort = _sort[::-1]
                sort[_directions[direction]] = _sort
                regions[_directions[direction]] = _regions

                if pseudosort_hist_fn:
                    _ranked_pseudosort_col = annot_df[pseudosort_col].iloc[_sort].reset_index(drop=True)
                    fig, ax = plt.subplots(1, 1, figsize=(figsize[0], pseudosort_hist_height))
                    for value in annot_df[pseudosort_col].drop_duplicates().sort_values():
                        sns.distplot(_ranked_pseudosort_col.index.values[_ranked_pseudosort_col.values == value],
                                     bins=10, label=value, ax=ax)
                    ax.set_title(coef)
                    ax.tick_params(axis='both', which='both',
                                   bottom=False, left=False, labelbottom=False, labelleft=False)
                    ax.set_xlim((0, _ranked_pseudosort_col.shape[0]))
                    ax.legend(bbox_to_anchor=(1, 1))
                    sns.despine(ax=ax)
                    savefig(pseudosort_hist_fn.format(direction=_directions[direction].lower()))
                    plt.close(fig)
            elif verbose:
                print('No regions for {} {} {} {}'.format(coef, top_n, direction, 'pseudosort'))
    else:
        sort = {'Up/Down': np.arange(atac_df.shape[0])}
        regions = {'Up/Down': set([])}
        for direction in _directions.keys():
            regions['Up/Down'] = regions['Up/Down'].union(
                get_top_n_regions(de_df, coef=coef, direction=direction, rank_metric=rank_metric, top_n=top_n))

    cg = {}
    for direction in sort:
        if len(regions[direction]) > 0:
            if annotations and legend_bbox == 'auto':
                _legend_bbox = (1.3 + (0.1 * len(annotations)), 1) if not pseudo_sort else (1.05, 1)
            else:
                _legend_bbox = legend_bbox
            cg[direction] = clustermap(
                (atac_df[regions[direction]].iloc[sort[direction]].T, None, annot_df.iloc[sort[direction]]),
                col_anns=annotations, col_anns_palettes=palettes,
                col_anns_legend=annotations, col_legend_bbox=_legend_bbox,
                method='average', metric='euclidean', z_score=0, standard_scale=None, figsize=figsize,
                row_cluster=True, col_cluster=not pseudo_sort,
                vmin=None, vmax=None, robust=True, cmap=sns.cm.rocket_r,
                xticklabels=False, yticklabels=False,
                show_row_dendrogram=True, show_col_dendrogram=True,
                ylabel='{} {} differential regions'.format(coef, direction.lower()),
                xlabel='Samples',
                raster_heatmap=rasterized,
                cbar_kws=dict(label='Z-score'))
            if fig_fn:
                savefig(fig_fn.format(direction=direction.lower().replace('/', '_')),
                        fig=cg[direction].fig, dpi=dpi)
        elif verbose:
            print('No regions for {} {} {} {}'.format(
                coef, top_n, direction, 'pseudosort' if pseudo_sort else 'cluster'))

    return cg if pseudo_sort else cg.get('Up/Down')


def homer_analysis(celltype, model, contrasts, directions, effect_size_filter, rank_metrics, top_ns,
                   size, peak_annot_fn, universe_fn=None, random=False, seed=None, results_dir='results', n_jobs=8):
    assert not random or seed is not None
    _BED_COLS = ['chr', 'start', 'end', 'gene_name', 'characterization']
    if isinstance(rank_metrics, str):
        rank_metrics = [rank_metrics]
    if isinstance(top_ns, int) or isinstance(top_ns, float):
        top_ns = [top_ns]

    if size is None:
        size = 'given'

    make_dir(os.path.split(homer_fn(
        celltype, model, data='{data}', coef=None, effect_size_filter=None, rank_metric=None,
        top_n=None, direction=None, results_dir=results_dir))[0])

    to_run = _save_and_plot_regions_for_enr(
        celltype=celltype, model=model, contrasts=contrasts, directions=directions, effect_size_filter=effect_size_filter,
        rank_metrics=rank_metrics, top_ns=top_ns, peak_annot_fn=peak_annot_fn, results_dir=results_dir,
        fn_func=homer_fn, universe_fn=universe_fn, output_as_dirs=True, random=random, seed=seed, verbose=False
    )

    for bed_fn, universe_fn, output_dir in to_run:
        homer(bed_fn=bed_fn, universe_fn=universe_fn, output_dir=f'{output_dir}.size_{size}', size=size, n_jobs=n_jobs)


def homer(bed_fn, universe_fn, output_dir, size, genome='hg38', htest=True, n_jobs=8):
    assert size == 'given' or isinstance(size, int)
    cmd = f'findMotifsGenome.pl {bed_fn} {genome} {output_dir}{" -h" if htest else ""} -size {size} -bg {universe_fn} -p {n_jobs}'
    print(cmd)
    subprocess.run(cmd.split())


def _lola_dbs_to_str(databases):
    if databases:
        return '_'.join([db[:-len(os.path.sep) - 4] if db.endswith('{}hg19'.format(os.path.sep)) or db.endswith(
            '{}hg38'.format(os.path.sep)) else db.replace(os.path.sep, '_') for db in databases])
    else:
        return None


def _save_and_plot_regions_for_enr(celltype, model, contrasts, region_filters, effect_size_filter, rank_metrics, top_ns,
                                   peak_annot_fn, results_dir, fn_func, universe_fn=None, output_as_dirs=False, save_bed=True,
                                   plots=[], figsize=(5, 5), palette='deep', BED_cols=None, directions=[-1, 1],
                                   random=False, seed=None, verbose=True):
    assert not random or seed is not None
    if region_filters is None or isinstance(region_filters, str):
        region_filters = [region_filters]
    assert all([f in [None, TSS_PROXIMAL, GENE_AND_DISTAL_10kb, DISTAL_1Mb, PROMOTERS, ALL_GENES] for f in region_filters])

    if BED_cols is None:
        BED_cols = ['chr', 'start', 'end', 'gene_name', 'characterization']

    de_df = read_de(celltype=celltype, model=model, contrasts=contrasts, annot_fn=peak_annot_fn,
                    add_coef_times_pval=True, effect_size_filter=effect_size_filter, results_dir=results_dir)
    de_df = simplify_peak_annot(de_df)

    to_run = []
    for rank_metric in rank_metrics:
        assert rank_metric in ['p.value', 'Coef', 'Coef_times_p.value', 'Coef_times_neg_log10_p.value']
        for region_filter in region_filters:
            region_filter_mask = peaks_to_genes(de_df, **PEAKS_TO_GENES[region_filter]) if region_filter else np.ones(
                (de_df.shape[0],), dtype=bool)
            if universe_fn is None and save_bed:
                universe_fn = fn_func(celltype, model, data='universe', regions=region_filter,
                                      coef=None, effect_size_filter=None,
                                      rank_metric=None, top_n=None, direction=None, results_dir=results_dir)
                make_bed(regions=de_df.loc[region_filter_mask].index, fn=universe_fn, peak_annot=de_df[BED_cols], extras=[])

            for top_n in top_ns:
                _no_regions = True
                fig, axs, order = {}, {}, {}
                for p in plots:
                    fig[p], axs[p] = plt.subplots(
                        2, len(contrasts), figsize=(len(contrasts) * figsize[0], 2 * figsize[1]), sharey='row',
                        squeeze=False)
                    fig[p].subplots_adjust(wspace=0.4, hspace=0.9 if p == 'characterization' else 1.1)
                for col, coef in enumerate(contrasts):
                    for row, direction in enumerate(directions):
                        top_regions = get_top_n_regions(
                            de_df.loc[region_filter_mask], coef=coef, direction=direction, rank_metric=rank_metric, top_n=top_n)

                        if random:
                            if verbose:
                                print('RANDOM')
                            n_regions_per_location = de_df.loc[top_regions].groupby(GENOMIC_LOCATION_COL)[coef_col(coef)].count()
                            print(n_regions_per_location)
                            assert len(top_regions) == n_regions_per_location.sum(), (len(top_regions), n_regions_per_location.sum())
                            top_regions = pd.Index([])
                            for location in n_regions_per_location.index:
                                top_regions = top_regions.union(de_df.loc[de_df[GENOMIC_LOCATION_COL] == location].sample(
                                    n=n_regions_per_location.loc[location], random_state=seed).index)
                            assert len(top_regions) == n_regions_per_location.sum(), (len(top_regions), n_regions_per_location.sum())

                        if verbose:
                            print('{} {} {}: {} regions'.format(
                                coef, '{}{}'.format('top ' if top_n >= 1 else 'FDR < ', top_n),
                                'DOWN' if direction == -1 else 'UP', len(top_regions)))
                        if len(top_regions) > 0:
                            _no_regions = False
                            for p in plots:
                                ax = sns.countplot(de_df.loc[top_regions, p],
                                                   order=PEAK_FEATURE_ORDER[p], palette=palette, ax=axs[p][row, col])
                                _xticklabels = ax.get_xticklabels()
                                ax.tick_params(reset=True, top=False, right=False)
                                ax.set_xticklabels(_xticklabels, rotation=45, ha='right')
                                ax.set_xlabel(None)
                                ax.set_ylabel('Number of regions')
                                ax.set_title('{}: {} regions\n({})'.format(
                                    coef, 'DOWN' if direction == -1 else 'UP',
                                    'top {}{}'.format(len(top_regions), ', FDR$<${}'.format(top_n) if top_n < 1 else '')))
                                sns.despine(ax=axs[p][row, col])

                            if save_bed:
                                bed_fn = fn_func(celltype, model, data='top_regions', coef=coef, regions=region_filter,
                                                 effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                                 top_n=top_n, direction=direction, random_seed=seed if random else None,
                                                 results_dir=results_dir)
                                make_bed(regions=top_regions, fn=bed_fn, peak_annot=de_df[BED_cols], extras=[])
                                res_fn = fn_func(celltype, model, data='results', coef=coef, regions=region_filter,
                                                 effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                                 top_n=top_n, direction=direction, ext=None if output_as_dirs else 'tsv',
                                                 random_seed=seed if random else None, results_dir=results_dir)
                                to_run.append((bed_fn, universe_fn, res_fn))
                        else:
                            for p in plots:
                                axs[p][row, col].get_xaxis().set_visible(False)
                                axs[p][row, col].get_yaxis().set_visible(False)
                                sns.despine(ax=axs[p][row, col], bottom=True, left=True)

                if not _no_regions:
                    for p in plots:
                        fig_fn = fn_func(celltype, model, data='barplot_{}'.format(p), regions=region_filter,
                                         effect_size_filter=effect_size_filter, rank_metric=None, top_n=top_n, coef=None,
                                         direction=None, db=None, random_seed=seed if random else None, results_dir=results_dir)
                        savefig(fig_fn, fig=fig[p])

                for p in fig:
                    plt.close(fig[p])

    return to_run


def lola_analysis(celltype, model, contrasts, region_filters, effect_size_filter, rank_metrics, top_ns, peak_annot_fn, databases,
                  db_dir=LOLA_DB_DIR, universe_fn=None, run_lola=True, save_bed=True, plot=True, results_dir='results',
                  plots=None, palette='deep', figsize=(5, 5), random=False, seed=None, n_jobs=8, verbose=True):

    assert not random or seed is not None
    if isinstance(databases, str):
        databases = [databases]
    if isinstance(rank_metrics, str):
        rank_metrics = [rank_metrics]
    if isinstance(databases, str):
        databases = [databases]
    if isinstance(top_ns, int) or isinstance(top_ns, float):
        top_ns = [top_ns]
    if not plot:
        plots = []
    elif plots is None:
        plots = [GENOMIC_LOCATION_COL, REGULATORY_ANNOT_COL, FEATURE_TYPE_COL]
    elif isinstance(plots, str):
        plots = [plots]

    assert all([p in [GENOMIC_LOCATION_COL, REGULATORY_ANNOT_COL, FEATURE_TYPE_COL] for p in plots])
    assert save_bed or not run_lola

    to_run = _save_and_plot_regions_for_enr(
        celltype=celltype, model=model, contrasts=contrasts, region_filters=region_filters, effect_size_filter=effect_size_filter,
        rank_metrics=rank_metrics, top_ns=top_ns, peak_annot_fn=peak_annot_fn, universe_fn=universe_fn,
        results_dir=results_dir, fn_func=lola_fn, save_bed=save_bed, plots=plots, figsize=figsize, palette=palette,
        random=random, seed=seed, verbose=verbose)

    if run_lola:
        bed_files, universe_files, output_files = zip(*to_run)
        assert len(set(universe_files)) == 1
        lola(bed_files=bed_files, universe_file=universe_files[0], output_files=output_files,
             databases=[os.path.join(db_dir, db) for db in databases], cpus=n_jobs)


def quick_lola(top_regions, fn_template, peak_annot, databases, db_dir=LOLA_DB_DIR, n_jobs=1):
    universe_fn = fn_template.format(data='universe')
    make_bed(regions=peak_annot.index, fn=universe_fn, peak_annot=peak_annot, extras=[])
    bed_fn = fn_template.format(data='top_regions')
    make_bed(regions=top_regions, fn=bed_fn, peak_annot=peak_annot, extras=[])
    res_fn = fn_template.format(data='results')
    lola(bed_files=bed_fn, universe_file=universe_fn, output_files=res_fn,
         databases=[os.path.join(db_dir, db) for db in databases], cpus=n_jobs)
    return res_fn


def map_gene_sets_to_regions(library, peaks_and_genes, make_upper=False):
    region_set_library = {}
    for term, genes in read_gene_sets(library, description_NA=None, assert_upper=False, make_upper=make_upper):
        region_set_library[term] = np.unique(peaks_and_genes.loc[peaks_and_genes.isin(genes)].index).tolist()
    return region_set_library


def run_gr_enrichr(celltype, model, contrasts, effect_size_filter, rank_metrics, top_ns, region_filters, peak_annot_fn,
                   gene_sets, db_dir, F_test=False,
                   results_dir='results', strict_background=True, padj_union=False):
    if isinstance(gene_sets, str):
        gene_sets = [gene_sets]
    if isinstance(contrasts, str):
        contrasts = [contrasts]
    if isinstance(rank_metrics, str):
        rank_metrics = [rank_metrics]
    if isinstance(region_filters, str):
        region_filters = [region_filters]
    if isinstance(top_ns, int) or isinstance(top_ns, float):
        top_ns = [top_ns]

    region_filters = [GENE_AND_DISTAL_10kb if f == 'DISTAL_10000' else f for f in region_filters]
    assert all([f in [None, TSS_PROXIMAL, GENE_AND_DISTAL_10kb, DISTAL_1Mb, PROMOTERS, ALL_GENES] for f in region_filters])

    de_df = read_de(celltype=celltype, model=model, contrasts=contrasts, F_test=F_test, annot_fn=peak_annot_fn,
                    add_coef_times_pval=True, effect_size_filter=effect_size_filter, results_dir=results_dir)

    for rank_metric in rank_metrics:
        assert rank_metric in ['p.value', 'Coef', 'Coef_times_p.value', 'Coef_times_neg_log10_p.value']
        for region_filter in region_filters:
            for top_n in top_ns:
                for coef in contrasts:
                    for direction in [-1, 1]:
                        top_regions, enr_df = gr_enrichr(de_df=de_df, region_filter=region_filter, top_n=top_n,
                                                         coef=coef, direction=direction, rank_metric=rank_metric,
                                                         db_dir=db_dir, gene_sets=gene_sets,
                                                         strict_background=strict_background,
                                                         padj_union=padj_union)
                        if len(top_regions) > 0:
                            _fn = enrichr_fn(celltype=celltype, model=model, data='{data}', coef=coef,
                                             effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                             top_n=top_n, direction=direction, db=None,
                                             regions=region_filter, results_dir=results_dir)
                            with open(_fn.format(data='top_regions'), 'w') as f:
                                f.writelines('{}\n'.format('\n'.join(top_regions)))
                            enr_df.to_csv(_fn.format(data='results'), index=False, sep='\t')


def get_gene_sets_for_region_filter(peaks_df, region_filter, db_dir, gene_sets, strict_background=True):
    peaks_mask = peaks_to_genes(peaks_df, **PEAKS_TO_GENES[region_filter]) if region_filter else np.ones(
        (peaks_df.shape[0],), dtype=bool)
    background = (peaks_df.loc[peaks_mask] if strict_background else peaks_df).index.values
    peaks_and_genes = (peaks_df.loc[peaks_mask] if strict_background else peaks_df)['gene_name'].str.upper()
    peaks_and_genes = peaks_and_genes.str.split(', ', expand=True).stack()
    peaks_and_genes.index = peaks_and_genes.index.get_level_values(0)
    gene_set_libraries = [(
        gs,
        map_gene_sets_to_regions(os.path.join(db_dir, '{}.gmt'.format(gs)), peaks_and_genes, make_upper=True)
        if region_filter else
        gene_set_library(os.path.join(db_dir, '{}.gmt'.format(gs)),
                         description_NA=None, as_arrays=True, assert_upper=False, make_upper=True)
    ) for gs in gene_sets]
    return peaks_mask, background, gene_set_libraries


def gr_enrichr(de_df, region_filter, top_n, coef, direction, rank_metric, db_dir, gene_sets,
               strict_background=True, padj_union=False):
    assert direction in [None, -1, 1]
    assert region_filter in [None, TSS_PROXIMAL, GENE_AND_DISTAL_10kb, DISTAL_1Mb, PROMOTERS, ALL_GENES]
    assert rank_metric in ['p.value', 'Coef', 'Coef_times_p.value', 'Coef_times_neg_log10_p.value', 'FVE']

    if isinstance(gene_sets, str):
        gene_sets = [gene_sets]

    peaks_mask, background, gene_set_libraries = get_gene_sets_for_region_filter(
        de_df, region_filter, db_dir, gene_sets, strict_background=strict_background)
    top_regions = get_top_n_regions(de_df.loc[peaks_mask], coef=coef, direction=direction,
                                    rank_metric=rank_metric, top_n=top_n)
    if len(top_regions) > 0:
        enr_df = gene_set_enrichment_test(top_regions, gene_set_libraries=gene_set_libraries,
                                          background=background, padj_method='fdr_bh', padj_union=padj_union)
    else:
        enr_df = None

    return top_regions, enr_df


def enrichr_online(celltype, model, contrasts, effect_size_filter, rank_metrics, top_ns, region_filters, peak_annot_fn,
                   gene_sets,
                   results_dir='results', verbose=True):
    import gseapy
    if isinstance(gene_sets, str):
        gene_sets = [gene_sets]
    if isinstance(contrasts, str):
        contrasts = [contrasts]
    if isinstance(rank_metrics, str):
        rank_metrics = [rank_metrics]
    if isinstance(region_filters, str):
        region_filters = [region_filters]
    if isinstance(top_ns, int) or isinstance(top_ns, float):
        top_ns = [top_ns]

    region_filters = [GENE_AND_DISTAL_10kb if f == 'DISTAL_10000' else f for f in region_filters]
    assert all([f in [TSS_PROXIMAL, GENE_AND_DISTAL_10kb, DISTAL_1Mb, PROMOTERS, ALL_GENES] for f in region_filters])

    de_df = read_de(celltype=celltype, model=model, contrasts=contrasts, annot_fn=peak_annot_fn,
                    add_coef_times_pval=True, effect_size_filter=effect_size_filter)

    for rank_metric in rank_metrics:
        assert rank_metric in ['p.value', 'Coef', 'Coef_times_p.value', 'Coef_times_neg_log10_p.value']
        for region_filter in region_filters:
            for top_n in top_ns:
                for coef in contrasts:
                    for direction in [-1, 1]:
                        _direction_mask = get_direction_mask(de_df, coef, direction)
                        _peaks_mask = peaks_to_genes(de_df, **PEAKS_TO_GENES[region_filter])
                        _df = de_df.loc[_direction_mask & _peaks_mask].copy()
                        _sort_col, _ascending = _get_rank_order_args(rank_metric, coef, direction)
                        _df['_rank'] = _df[_sort_col].rank(ascending=_ascending)
                        top_regions = get_top_n_regions(_df.loc[_df.groupby('gene_name')['_rank'].idxmin().values],
                                                        coef=coef, direction=direction, rank_metric=rank_metric,
                                                        top_n=top_n)
                        if len(top_regions) > 0:
                            top_genes = _df.loc[top_regions].sort_values('_rank')['gene_name'].str.upper()
                            assert len(top_genes) == len(set(top_genes))
                            _fn = enrichr_fn(celltype=celltype, model=model, data='{data}', coef=coef,
                                             effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                             top_n=top_n, direction=direction, db=None,
                                             regions=region_filter, results_dir=results_dir)
                            top_genes.to_csv(_fn.format(data='top_genes'), header=False, index=False)
                            _enr = gseapy.enrichr(gene_list=top_genes.tolist(), gene_sets=gene_sets,
                                                  organism='human', background='hsapiens_gene_ensembl',
                                                  description=os.path.split(_fn.format(data='top_genes'))[1],
                                                  outdir='tmp', no_plot=True, verbose=verbose)

                            _enr.results[ENRICHR_COLS].to_csv(_fn.format(data='results'), index=False, sep='\t')


def plot_lola(celltype, model, contrasts, effect_size_filter, rank_metrics, top_ns, databases, collections, show_n=20,
              metric='pValueLog',
              remove_replicates=False, results_dir='results', palette=None, barwidth=0.75, figsize=(5, 5),
              celltype_fixes=LOLA_CELLTYPES_FIXES, fdr=None, verbose=True):
    if isinstance(databases, str):
        databases = [databases]
    if isinstance(contrasts, str):
        contrasts = [contrasts]
    if isinstance(rank_metrics, str):
        rank_metrics = [rank_metrics]
    if isinstance(top_ns, int) or isinstance(top_ns, float):
        top_ns = [top_ns]

    palette = dict(Down=BLUE, Up=RED) if palette is None else sns.color_palette(palette) if isinstance(palette,
                                                                                                       str) else palette
    assert metric in ['pValueLog', 'oddsRatio', 'support', 'qValue']
    _ascending = metric == 'qValue'

    for rank_metric in rank_metrics:
        for top_n in top_ns:
            results = defaultdict(lambda: defaultdict(lambda: defaultdict(
                lambda: pd.DataFrame.from_dict({'Term': [''] * show_n, metric: [0] * show_n}))))
            fdr_to_pValueLog = defaultdict(lambda: dict(Up=0, Down=0))
            lola_df = None
            for coef in contrasts:
                for direction in [-1, 1]:
                    res_fn = lola_fn(celltype, model, data='results', coef=coef,
                                     effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                     top_n=top_n, direction=direction,
                                     db=_lola_dbs_to_str(databases), results_dir=results_dir)
                    if os.path.exists(res_fn):
                        lola_df = pd.read_csv(res_fn, index_col=1, sep='\t')
                        if collections:
                            lola_df = lola_df.loc[lola_df['collection'].isin(collections)]
                        assert not lola_df['pValueLog'].isnull().any()
                        lola_df.loc[~np.isfinite(lola_df['pValueLog']) & (lola_df['pValueLog'] > 0), 'pValueLog'] = -np.log(sys.float_info.min)
                        # lola_df.loc[lola_df['pValueLog'].isnull(), 'pValueLog'] = lola_df['pValueLog'].max()
                        assert len(set(lola_df['userSet'])) == 1 and lola_df.index.is_unique

                        if fdr and metric == 'pValueLog':
                            lola_df['padj'] = multipletests(np.power(10, -lola_df['pValueLog']), method='fdr_bh')[1]
                            fdr_to_pValueLog[coef]['Up' if direction == 1 else 'Down'] = \
                                fdr_to_pval(lola_df, fdr=fdr, fdr_col='padj', pval_col=metric, ascending=_ascending, is_neg_log10=True)

                        if celltype_fixes:
                            lola_df['cellType'] = lola_df['cellType'].replace(to_replace=celltype_fixes, regex=True)

                        for collection in lola_df['collection'].drop_duplicates():
                            collection_df = lola_df.loc[lola_df['collection'] == collection].copy()
                            if collection in ['jaspar_motifs', 'roadmap_epigenomics']:
                                collection_df['Term'] = collection_df['filename'].str.upper().str.replace(
                                    '__', '/').str.replace('.BED$', '')
                            else:
                                collection_df['Term'] = collection_df['antibody'] + ' (' + collection_df[
                                    'cellType'] + ')'
                            if remove_replicates:
                                grouped = collection_df.groupby('Term')[metric]
                                idx = grouped.idxmax() if metric != 'qValue' else grouped.idxmin()
                                collection_df = collection_df.loc[idx]
                            collection_df = collection_df.sort_values([metric, 'oddsRatio'], ascending=[_ascending, False])
                            results[collection][coef]['Up' if direction == 1 else 'Down'] = collection_df.head(show_n)
                    elif verbose:
                        print('Missing', res_fn)

            # lola_df being None means no regions for this top_n at all
            if lola_df is not None:
                fig_fn_template = lola_fn(celltype, model, data='barplot_{library}', coef=None,
                                          effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                          top_n=top_n, direction=None, db=_lola_dbs_to_str(databases),
                                          results_dir=results_dir)
                _collections = lola_df['collection'].drop_duplicates()
                fig, axs = make_enrichment_plots(
                    results, libraries=_collections, contrasts=contrasts, fig_fn_template=fig_fn_template,
                    metric=metric,
                    show_n=show_n, term_col='Term', convert_to_neg_log10=(metric == 'qValue'), palette=palette,
                    barwidth=barwidth, figsize=figsize, fdr=fdr,
                    fdr_maps={c: fdr_to_pValueLog for c in _collections} if fdr and metric == 'pValueLog' else None)
                for p in fig:
                    plt.close(fig[p])


def plot_enrichr(celltype, model, contrasts, effect_size_filter, rank_metrics, top_ns, region_filters, gene_sets,
                 show_n=20, metric='P-value',
                 results_dir='results', palette=None, barwidth=0.75, figsize=(5, 5),
                 gene_sets_fixes=GENE_SETS_FIXES, fdr=None, hide_legend=False, verbose=True):
    if isinstance(gene_sets, str):
        gene_sets = [gene_sets]
    if isinstance(contrasts, str):
        contrasts = [contrasts]
    if isinstance(rank_metrics, str):
        rank_metrics = [rank_metrics]
    if isinstance(region_filters, str):
        region_filters = [region_filters]
    if isinstance(top_ns, int) or isinstance(top_ns, float):
        top_ns = [top_ns]

    region_filters = [GENE_AND_DISTAL_10kb if f == 'DISTAL_10000' else f for f in region_filters]
    palette = dict(Down=BLUE, Up=RED) if palette is None else sns.color_palette(palette) if isinstance(palette,
                                                                                                       str) else palette
    assert metric in ['P-value', 'Adjusted P-value', 'Combined Score', 'Odds Ratio', 'Overlap']
    _ascending = metric in ['P-value', 'Adjusted P-value']

    for rank_metric in rank_metrics:
        for region_filter in region_filters:
            for top_n in top_ns:
                results = defaultdict(lambda: defaultdict(lambda: defaultdict(
                    lambda: pd.DataFrame.from_dict({'Term': [''] * show_n, metric: [0] * show_n}))))
                fdr_to_P_value = defaultdict(lambda: defaultdict(lambda: dict(Up=0, Down=0)))
                enr_df = None
                for coef in contrasts:
                    for direction in [-1, 1]:
                        res_fn = enrichr_fn(celltype, model, data='results', coef=coef,
                                            effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                            top_n=top_n, direction=direction, db=_lola_dbs_to_str(gene_sets),
                                            regions=region_filter, results_dir=results_dir)
                        if os.path.exists(res_fn):
                            enr_df = pd.read_csv(res_fn, index_col=None, sep='\t').sort_values(metric,
                                                                                               ascending=_ascending)
                            assert not enr_df['P-value'].isnull().any()
                            if gene_sets_fixes:
                                enr_df['Term'] = enr_df['Term'].replace(to_replace=gene_sets_fixes, regex=True)
                            for library in enr_df['Gene_set'].drop_duplicates():
                                library_df = enr_df.loc[enr_df['Gene_set'] == library]
                                if fdr and metric == 'P-value':
                                    fdr_to_P_value[library][coef]['Up' if direction == 1 else 'Down'] = \
                                        fdr_to_pval(library_df, fdr=fdr, fdr_col='Adjusted P-value', pval_col=metric, ascending=_ascending)
                                results[library][coef]['Up' if direction == 1 else 'Down'] = library_df.head(show_n)
                        elif verbose:
                            print('Missing', res_fn)

                # enr_df being None means no genes for this top_n/region_filter at all
                if enr_df is not None:
                    fig_fn_template = enrichr_fn(celltype, model, data='barplot_{library}', coef=None,
                                                 effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                                 top_n=top_n, direction=None, db=_lola_dbs_to_str(gene_sets),
                                                 regions=region_filter, results_dir=results_dir)

                    fig, axs = make_enrichment_plots(
                        results, libraries=enr_df['Gene_set'].drop_duplicates().values, contrasts=contrasts,
                        fig_fn_template=fig_fn_template, metric=metric,
                        show_n=show_n, term_col='Term',
                        convert_to_neg_log10=(metric in ['P-value', 'Adjusted P-value']),
                        palette=palette, barwidth=barwidth, figsize=figsize, fdr=fdr,
                        fdr_maps=fdr_to_P_value if fdr and metric == 'P-value' else None, legend_bbox_max=5,
                        hide_legend=hide_legend)

    return (fig, axs) if enr_df is not None and len(top_ns) == 1 and len(region_filters) == 1 and len(
        rank_metrics) == 1 else None


def make_enrichment_plots(results, libraries, contrasts, fig_fn_template, metric, show_n=20, term_col='Term',
                          convert_to_neg_log10=False, palette=None, barwidth=0.75, figsize=(5, 5), fdr=None,
                          fdr_maps=None, legend_bbox_min=1.75, legend_bbox_max=2.5, hide_legend=False,
                          strip_term_ids=True):
    if isinstance(libraries, str):
        libraries = [libraries]
    if isinstance(contrasts, str):
        contrasts = [contrasts]

    fig, axs = {}, {}
    for lib in libraries:
        _longest_term = np.max([results[lib][coef][d][term_col].str.len().max() for coef in results[lib] for d in
                                results[lib][coef]]) - (6 if WIKI_PATHWAYS in lib else 13 if GO_BIO_PROCESS in lib else 0)
        fig[lib], axs[lib] = plt.subplots(1, len(contrasts), squeeze=False,
                                          figsize=(len(contrasts) * figsize[0], figsize[1]))
        fig[lib].subplots_adjust(wspace=_longest_term / 10)
        for i, (ax, coef) in enumerate(zip(axs[lib][0], contrasts)):
            enr_data = {d: results[lib][coef][d].reset_index().copy() for d in ['Down', 'Up']}
            for d in enr_data:
                if WIKI_PATHWAYS in lib:
                    enr_data[d][term_col] = enr_data[d][term_col].str.replace(' WP[0-9]+$', '')
                elif GO_BIO_PROCESS in lib:
                    enr_data[d][term_col] = enr_data[d][term_col].str.replace(' \(GO:[0-9]+\)$', '')
            lax, rax = enrichment_plot(
                enr_data, top_n=show_n, score=metric, term_col=term_col, convert_to_neg_log10=convert_to_neg_log10,
                thresholds=fdr_maps[lib][coef] if fdr_maps else None,
                thr_kwargs={'Down': dict(c='k', ls='--'),
                            'Up': dict(c='k', ls='--', label='FDR {}'.format(fdr))},
                color=palette, legend=not hide_legend and (i + 1 == len(contrasts)),
                legend_kwargs=dict(bbox_to_anchor=(min(max(legend_bbox_min, _longest_term / 10), legend_bbox_max), 1),
                                   title='Fold change'),
                width=barwidth, edgecolor='0.15', lw=mpl.rcParams['lines.linewidth'] / 2, ax=ax
            )
            lax.set_title('{}\n{}'.format(lib, coef), y=1.02)
            if 'enrichment' not in lax.get_xlabel().lower():
                lax.set_xlabel('Enrichment {}'.format(lax.get_xlabel()))
            lax.set_ylabel(None)
            rax.set_ylabel(None)
        if fig_fn_template:
            savefig(fig_fn_template.format(library=lib), fig=fig[lib])
    return fig, axs


def fdr_to_pval(df, fdr, fdr_col, pval_col, ascending, is_neg_log10=False):
    if (df[fdr_col] <= fdr).sum() > 1:
        # value = df.loc[(df[fdr_col] - fdr).abs().idxmin(), pval_col]
        value = df.loc[df[fdr_col] <= fdr, pval_col].sort_values(ascending=ascending).iloc[-1]
    else:
        # when there is no test passing FDR, it is equal to FWER
        value = fdr / df.shape[0]
        if is_neg_log10:
            value = -np.log10(value)
    return value


def reduce_regions_to_genes_with_a_rule_of_thumb(group):
    if (group > 0).sum() / len(group) >= 2 / 3:
        return group.max()
    elif (group < 0).sum() / len(group) >= 2 / 3:
        return group.min()
    else:
        return group.mean()


def gsea_collect(celltype, model, coef, effect_size_filter, rank_metric, region_filter, gene_sets, db_dir, results_dir):
    outfile_template = gsea_fn(celltype=celltype, model=model, data='{data}', coef=coef, db=None,
                               effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                               regions=region_filter, results_dir=results_dir, ext=None)
    _GSEA_COLS = ['NAME', 'SIZE', 'ES', 'NES', 'NOM p-val', 'FDR q-val', 'FWER p-val', 'RANK AT MAX', 'LEADING EDGE']
    _directions = {-1: 'neg', 1: 'pos'}
    results = defaultdict(lambda: {})
    for library in gene_sets:
        outdir = outfile_template.format(data=library)
        _all_terms = []
        for direction in _directions.keys():
            results[library][direction] = pd.read_csv(
                os.path.join(outdir, 'gsea_report_for_na_{}.tsv'.format(_directions[direction])),
                sep='\t', index_col=0, usecols=_GSEA_COLS, na_values='---')
            assert (results[library][direction]['ES'] * direction >= 0).all()
            assert ((results[library][direction]['NES'] * direction >= 0) | results[library][direction]['NES'].isnull()).all()

            for term in results[library][direction].index:
                # sets of up and down terms must be disjoint
                assert term not in _all_terms
                _all_terms.append(term)
                _fn = os.path.join(outdir, '{}.xls'.format(term))
                if os.path.exists(_fn):
                    term_df = pd.read_csv(_fn, sep='\t', index_col=0,
                                          usecols=['PROBE', 'RANK METRIC SCORE', 'CORE ENRICHMENT'])
                    assert term_df['CORE ENRICHMENT'].isin(['Yes', 'No']).all()
                    ledge_genes = term_df.loc[term_df['CORE ENRICHMENT'] == 'Yes', 'RANK METRIC SCORE']
                    results[library][direction].loc[term, 'LEADING EDGE'] = ';'.join(
                        ledge_genes.sort_values(ascending=(direction == -1)).index.values)
        assert sum([results[library][direction].shape[0] for direction in results[library]]) == len(_all_terms)

    for library in gene_sets:
        _df = pd.concat([results[library][direction] for direction in results[library]], axis=0).reset_index()
        _df['Gene_set'] = library
        _gene_set_terms = gene_set_library(os.path.join(db_dir, '{}.gmt'.format(library)), assert_upper=False).keys()
        _df['NAME'] = _df['NAME'].replace({t.upper(): t for t in _gene_set_terms})
        _df['_abs_NES'] = _df['NES'].abs()
        _df = _df.sort_values(['FDR q-val', '_abs_NES'], ascending=[True, False])
        results[library] = _df[['Gene_set'] + _GSEA_COLS].rename({'NAME': 'Term'}, axis=1)

    df = pd.concat([results[library] for library in results], axis=0)
    results_fn = gsea_fn(celltype=celltype, model=model, data='results', coef=coef,
                         db=None, effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                         regions=region_filter, results_dir=results_dir)
    df.to_csv(results_fn, sep='\t', index=False)
    return df


def gsea_jobs(celltype, model, contrasts, effect_size_filter, rank_metrics, region_filters, peak_annot_fn, gene_sets,
              db_dir,
              nperm=1000, weight=1, min_size=15, max_size=500, nplots=None, create_svgs=False,
              random_seed=RANDOM_STATE, results_dir='results', reduce_to_genes=False, use_gseapy=False,
              collect_and_plot_cmd=None,
              n_jobs=8, memory='16G', collect_jobs=4, collect_memory='4G', partition='shortq', walltime='12:00:00',
              exclude_inodes=False, run_locally=False, show_n=20, metric='FDR q-value', palette=None, barwidth=0.75, figsize=(5, 5),
              gene_sets_fixes=GENE_SETS_FIXES, fdr=None, verbose=True):
    assert memory[-1].upper() == 'G'
    assert collect_memory[-1].upper() == 'G'
    if isinstance(gene_sets, str):
        gene_sets = [gene_sets]
    if isinstance(contrasts, str):
        contrasts = [contrasts]
    if isinstance(rank_metrics, str):
        rank_metrics = [rank_metrics]
    if isinstance(region_filters, str):
        region_filters = [region_filters]
    if isinstance(gene_sets, str):
        gene_sets = [gene_sets]

    region_filters = [GENE_AND_DISTAL_10kb if f == 'DISTAL_10000' else f for f in region_filters]
    assert all([f in [TSS_PROXIMAL, GENE_AND_DISTAL_10kb, DISTAL_1Mb, PROMOTERS, ALL_GENES] for f in region_filters])

    de_df = read_de(celltype=celltype, model=model, contrasts=contrasts, annot_fn=peak_annot_fn,
                    add_coef_times_pval=True, effect_size_filter=effect_size_filter)

    for rank_metric in rank_metrics:
        assert rank_metric in ['p.value', 'Coef', 'Coef_times_p.value', 'Coef_times_neg_log10_p.value']
        for region_filter in region_filters:
            _df = de_df.loc[peaks_to_genes(de_df, **PEAKS_TO_GENES[region_filter])].copy()
            assert not _df['gene_name'].isnull().any()
            _df = _df.join(_df.groupby('gene_name')['A'].count().rename('MT_correction'), on='gene_name')
            job_ids = []
            for coef in contrasts:
                _sort_col, _ = _get_rank_order_args(rank_metric, coef)
                if rank_metric == 'p.value':
                    _df[_sort_col] = _df[_sort_col] * _df['MT_correction']
                    _df[_sort_col] = signed_log10_pvals(_df, coef=coef)

                if reduce_to_genes:
                    _idx = _df[_sort_col].abs().groupby(_df['gene_name']).idxmax()
                    ranked_features = _df.loc[_idx].set_index('gene_name')[_sort_col].sort_values(ascending=False)
                    ranked_features.index = ranked_features.index.str.upper()
                else:
                    ranked_features = _df[_sort_col].sort_values(ascending=False)

                # there can be nulls because of effect_size_filter
                ranked_features = ranked_features.loc[~ranked_features.isnull()]
                ranked_features_fn = gsea_fn(celltype=celltype, model=model,
                                             data='ranked_genes' if reduce_to_genes else 'ranked_regions',
                                             coef=coef, db=None, effect_size_filter=effect_size_filter,
                                             rank_metric=rank_metric, regions=region_filter, results_dir=results_dir)
                ranked_features.to_csv(ranked_features_fn, header=False, sep='\t')

                if use_gseapy:
                    results_fn = gsea_fn(method='gseapy', celltype=celltype, model=model, data='results', coef=coef,
                                         db=None, effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                         regions=region_filter, results_dir=results_dir)
                    outdir_template = gsea_fn(method='gseapy', celltype=celltype, model=model, data='{data}', coef=coef,
                                              db=None, effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                              regions=region_filter, results_dir=results_dir, ext=None)
                    _res = prerank_gseapy(ranked_features, libraries=gene_sets, db_dir=db_dir,
                                          outfile=results_fn, outdir_template=outdir_template, nperm=nperm,
                                          weight=weight,
                                          min_size=min_size, max_size=max_size, nplots=nplots, random_seed=random_seed,
                                          n_jobs=n_jobs)
                else:
                    outfile_template = gsea_fn(celltype=celltype, model=model, data='{data}', coef=coef, db=None,
                                               effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                               regions=region_filter, results_dir=results_dir, ext=None)
                    for library in gene_sets:
                        gene_set_fn = os.path.join(db_dir, '{}.gmt'.format(library))
                        outdir, job_name = os.path.split(outfile_template.format(data=library))
                        job_id = submit_gsea_job(ranked_features_fn=ranked_features_fn, gene_set_fn=gene_set_fn,
                                                 outdir=outdir,
                                                 job_name=job_name, nperm=nperm, weight=weight, min_size=min_size,
                                                 max_size=max_size, plot_top_x=nplots, create_svgs=create_svgs,
                                                 random_seed=random_seed, n_jobs=n_jobs, memory=memory,
                                                 partition=partition, walltime=walltime,
                                                 exclude_inodes=exclude_inodes, run_locally=run_locally, verbose=verbose)
                        job_ids.append(job_id)
            if collect_and_plot_cmd is not None:
                if run_locally:
                    for rank_metric in rank_metrics:
                        for region_filter in region_filters:
                            for coef in contrasts:
                                print('Collect GSEA {} {}'.format(region_filter, coef))
                                gsea_collect(celltype=celltype, model=model, coef=coef,
                                             effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                             region_filter=region_filter, gene_sets=gene_sets, db_dir=db_dir,
                                             results_dir=results_dir)
                    plot_gsea(celltype=celltype, model=model, contrasts=contrasts,
                              effect_size_filter=effect_size_filter, rank_metrics=rank_metrics,
                              region_filters=region_filters, gene_sets=None, show_n=show_n, metric=metric,
                              results_dir=results_dir, palette=palette, barwidth=barwidth,
                              figsize=figsize, gene_sets_fixes=gene_sets_fixes, fdr=fdr, min_pval=(1 / nperm),
                              use_gseapy=use_gseapy)
                else:
                    _, job_name = os.path.split(
                        gsea_fn(celltype=celltype, model=model, data='{data}', coef=None, db=None,
                                effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                regions=region_filter, results_dir=results_dir,
                                ext=None).format(data='collect'))
                    submit_gsea_collect_and_plot(
                        collect_and_plot_cmd=collect_and_plot_cmd.format(regions=region_filter),
                        after_job_ids=job_ids, outdir=outdir, job_name=job_name, n_cpus=collect_jobs,
                        memory=collect_memory, partition=partition, walltime=walltime,
                        exclude_inodes=exclude_inodes, verbose=verbose)


def submit_gsea_collect_and_plot(collect_and_plot_cmd, after_job_ids, outdir, job_name, n_cpus=4, memory='4G',
                                 partition='shortq', walltime='12:00:00', exclude_inodes=False, verbose=True):
    assert memory[-1].upper() == 'G'
    if verbose:
        print('************* submit_gsea_collect_and_plot')
        print(n_cpus, memory, partition, walltime)
        print(after_job_ids)
        print(collect_and_plot_cmd)
        print(outdir)
        print(job_name)

    with open_file(os.path.join(outdir, '{}.sh'.format(job_name)), 'w') as f:
        print('#!/bin/bash\n', file=f)
        print('python {}'.format(collect_and_plot_cmd), file=f)

    _SBATCH = (
        'sbatch '
        '--parsable --dependency=afterok:{dependency} --kill-on-invalid-dep=yes --job-name={job_name} --output={outdir}/{job_name}.log '
        '--mem={memory} --cpus-per-task={ncpus} --time={walltime} --partition={partition}{exclude_inodes} '
        '{script}'
    ).format(
        dependency=':'.join(after_job_ids),
        script=os.path.join(outdir, '{}.sh'.format(job_name)),
        job_name=job_name,
        outdir=outdir,
        walltime=walltime,
        memory=memory,
        partition=partition,
        ncpus=n_cpus,
        exclude_inodes=' --exclude=i[001-022]' if exclude_inodes else ''
    )
    if verbose:
        print(_SBATCH)
    job_id = subprocess.check_output(_SBATCH.split(' '), universal_newlines=True).strip()
    if verbose:
        print('(submitted, job_id: {})\n'.format(job_id))
    return job_id


def prerank_gseapy(ranked_features, libraries, db_dir, outfile=None, outdir_template=None, nperm=1000, weight=1,
                   min_size=15,
                   max_size=500, nplots=None, random_seed=RANDOM_STATE, n_jobs=8):
    import gseapy
    if nplots is None or nplots > 0:
        assert outdir_template

    if isinstance(libraries, str):
        libraries = [libraries]

    results = {}
    for library in libraries:
        gene_sets = gene_set_library(os.path.join(db_dir, '{}.gmt'.format(library)), as_arrays=False)
        _res = gseapy.prerank(rnk=ranked_features, gene_sets=gene_sets, permutation_num=nperm,
                              weighted_score_type=weight, outdir=None,
                              min_size=min_size, max_size=max_size, ascending=False, processes=n_jobs,
                              seed=random_seed, no_plot=True)
        _res.res2d['Gene_set'] = library
        results[library] = _res

        if nplots is None or nplots > 0:
            _outdir = outdir_template.format(data=library)
            os.makedirs(_outdir, exist_ok=True)
            if nplots is None:
                _to_plot = _res.res2d.index
            else:
                _to_plot = (_res.res2d.sort_values('fdr').head(nplots) if nplots >= 1 \
                                else _res.res2d.loc[_res.res2d['fdr'] < nplots].sort_values('fdr')).index
            with sns.plotting_context('notebook'):
                for term in _to_plot:
                    try:
                        gseapy.plot.gseaplot(
                            ofname=os.path.join(_outdir, 'gseapy_enplot_{}.pdf'.format(
                                term.replace('\\', '_').replace('/', '_').replace(': ', '_').replace(
                                    ':', '_').replace(', ', '_').replace(',', '_').replace(' ', '_'))),
                            rank_metric=_res.ranking, term=term, **_res.results[term])
                    except:
                        pass
    if outfile:
        results_df = pd.concat([results[library].res2d for library in results], axis=0)
        results_df = results_df[['Gene_set'] + results_df.columns[results_df.columns != 'Gene_set'].tolist()]
        results_df.to_csv(outfile, sep='\t')
    return results


def submit_gsea_job(ranked_features_fn, gene_set_fn, outdir, job_name, nperm=1000, weight=1, min_size=15, max_size=500,
                    plot_top_x=None, create_svgs=False, random_seed=RANDOM_STATE, n_jobs=8, memory='16G',
                    partition='shortq', walltime='12:00:00', exclude_inodes=False, run_locally=False, verbose=True):
    assert memory[-1].upper() == 'G'
    if verbose:
        print('submit_gsea_job')
        print(n_jobs, memory, partition, walltime)
        print('outdir: {}\njob_name: {}'.format(outdir, job_name))

    _gene_set_library_size = len(gene_set_library(gene_set_fn, assert_upper=False))
    if verbose:
        print('{}: {} terms'.format(gene_set_fn, _gene_set_library_size))

    _GSEA_SCRIPT = (
        '{home}/tools/GSEA_4.1.0/{script} {memory} GSEAPreranked -gmx {gmx} -collapse No_Collapse '
        '-mode Max_probe -norm meandiv -nperm {nperm} -rnk {rnk} -scoring_scheme {scoring_scheme} '
        '-rpt_label {rpt_label} -create_svgs {create_svgs} -include_only_symbols true -make_sets true '
        '-plot_top_x {plot_top_x} -rnd_seed {rnd_seed} -set_max {set_max} -set_min {set_min} -zip_report false '
        '-out {out}'
    ).format(
        home='$HOME' if not run_locally else os.environ['HOME'],
        script='gsea-cli2.sh' if not run_locally else 'gsea-cli.sh',
        memory=memory if not run_locally else '',
        gmx=gene_set_fn,
        nperm=nperm,
        rnk=ranked_features_fn,
        scoring_scheme=GSEA_WEIGHT_SCHEMES[weight],
        rpt_label=job_name,
        create_svgs=str(create_svgs).lower(),
        plot_top_x=_gene_set_library_size if plot_top_x is None else plot_top_x,
        rnd_seed=random_seed,
        set_max=max_size if max_size else int(1e6),
        set_min=min_size if min_size else 2,
        out=outdir
    )

    with open_file(os.path.join(outdir, '{}.sh'.format(job_name)), 'w') as f:
        print('#!/bin/bash\n', file=f)
        print(_GSEA_SCRIPT, file=f)
        _path = os.path.join(outdir, job_name)
        print('\nls -d {path}*'.format(path=_path), file=f)
        print('\nif [[ -d {path} ]]; then echo "Removing old run in {path}"; rm -r {path}; fi'.format(path=_path),
              file=f)
        print('\nmv {path}.GseaPreranked* {path}'.format(path=_path), file=f)
        print('\ncd {outdir}'.format(outdir=outdir), file=f)
        print('\nln -s {job_name}/index.html {job_name}.html'.format(job_name=job_name), file=f)
        print('\ncd {job_name}'.format(job_name=job_name), file=f)
        for _direction in ['pos', 'neg']:
            print('\nls gsea_report_for_na_{direction}_*.tsv'.format(direction=_direction), file=f)
            print('ln -s gsea_report_for_na_{direction}_*.tsv gsea_report_for_na_{direction}.tsv'.format(
                direction=_direction), file=f)

    if run_locally:
        subprocess.run(['bash', os.path.join(outdir, '{}.sh'.format(job_name))])
    else:
        _SBATCH = (
            'sbatch '
            '--parsable --job-name={job_name} --output={outdir}/{job_name}.log '
            '--mem={memory} --cpus-per-task={ncpus} --time={walltime} --partition={partition}{exclude_inodes} '
            '{script}'
        ).format(
            script=os.path.join(outdir, '{}.sh'.format(job_name)),
            job_name=job_name,
            outdir=outdir,
            walltime=walltime,
            memory=memory,
            partition=partition,
            ncpus=n_jobs,
            exclude_inodes=' --exclude=i[001-022]' if exclude_inodes else ''
        )
        if verbose:
            print(_SBATCH)
        job_id = subprocess.check_output(_SBATCH.split(' '), universal_newlines=True).strip()
        if verbose:
            print('(submitted, job_id: {})\n'.format(job_id))
        return job_id


def plot_gsea(celltype, model, contrasts, effect_size_filter, rank_metrics, region_filters, gene_sets, show_n=20,
              metric='FDR q-value',
              results_dir='results', palette=None, barwidth=0.75, figsize=(5, 5),
              gene_sets_fixes=GENE_SETS_FIXES, fdr=None, min_pval=None, use_gseapy=False, verbose=True):
    if isinstance(gene_sets, str):
        gene_sets = [gene_sets]
    if isinstance(contrasts, str):
        contrasts = [contrasts]
    if isinstance(rank_metrics, str):
        rank_metrics = [rank_metrics]
    if isinstance(region_filters, str):
        region_filters = [region_filters]

    region_filters = [GENE_AND_DISTAL_10kb if f == 'DISTAL_10000' else f for f in region_filters]
    palette = dict(Down=BLUE, Up=RED) if palette is None else sns.color_palette(palette) if isinstance(palette,
                                                                                                       str) else palette
    assert metric in ['NES', 'NOM p-val', 'FDR q-val', 'nes', 'pval', 'fdr']
    if metric.upper() == 'NES':
        metric = metric.upper()
        _sort_col = '_abs_NES'
        _ascending = False
    else:
        _sort_col = [metric, '_abs_NES']
        _ascending = [True, False]

    for rank_metric in rank_metrics:
        for region_filter in region_filters:
            results = defaultdict(lambda: defaultdict(lambda: defaultdict(
                lambda: pd.DataFrame.from_dict({'Term': [''] * show_n, metric: [0] * show_n}))))
            fdr_maps = defaultdict(lambda: defaultdict(lambda: dict(Up=0, Down=0)))
            enr_df = None
            for coef in contrasts:
                res_fn = gsea_fn(method='gseapy' if use_gseapy else 'gsea', celltype=celltype, model=model,
                                 data='results',
                                 coef=coef, db=None, effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                 regions=region_filter, results_dir=results_dir)
                if os.path.exists(res_fn):
                    enr_df = pd.read_csv(res_fn, index_col=None, sep='\t').rename({'nes': 'NES', 'es': 'ES'}, axis=1)
                    enr_df['_abs_NES'] = enr_df['NES'].abs()

                    # sort before replacing 0 P-vals and FDRs on purpose
                    # in case warning is issued, this will be visible in the figure as well
                    # because top hits will not have top P-vals/FDRs
                    enr_df = enr_df.sort_values(_sort_col, ascending=_ascending)
                    # now fix 0
                    if min_pval and metric in ['NOM p-val', 'FDR q-val', 'pval', 'fdr']:
                        _non_zero_min = enr_df.loc[enr_df[metric] != 0, metric].min()
                        if _non_zero_min < min_pval:
                            print('WARNING: Minimum non-zero "{}" is {} but you are using "min_pval" of {}, '
                                  'which is larger'.format(metric, _non_zero_min, min_pval))
                        enr_df.loc[enr_df[metric] == 0, metric] = min_pval

                    if gene_sets_fixes:
                        enr_df['Term'] = enr_df['Term'].replace(to_replace=gene_sets_fixes, regex=True)
                    for direction in [-1, 1]:
                        for library in enr_df['Gene_set'].drop_duplicates():
                            _df = enr_df.loc[(enr_df['ES'] * direction >= 0) & (enr_df['Gene_set'] == library)].head(
                                show_n).copy()
                            _df[metric] = _df[metric].abs()
                            results[library][coef]['Up' if direction == 1 else 'Down'] = _df
                            fdr_maps[library][coef]['Up' if direction == 1 else 'Down'] = fdr
                elif verbose:
                    print('Missing', res_fn)

            # enr_df being None means no genes for this region_filter at all
            if enr_df is not None:
                fig_fn_template = gsea_fn(method='gseapy' if use_gseapy else 'gsea', celltype=celltype, model=model,
                                          data='barplot_{library}', coef=None,
                                          effect_size_filter=effect_size_filter, rank_metric=rank_metric,
                                          db=_lola_dbs_to_str(gene_sets), regions=region_filter,
                                          results_dir=results_dir)
                fig, axs = make_enrichment_plots(
                    results, libraries=enr_df['Gene_set'].drop_duplicates().values, contrasts=contrasts,
                    fig_fn_template=fig_fn_template, metric=metric,
                    show_n=show_n, term_col='Term',
                    convert_to_neg_log10=(metric in ['NOM p-val', 'FDR q-val', 'pval', 'fdr']),
                    palette=palette, barwidth=barwidth, figsize=figsize, fdr=fdr,
                    fdr_maps=fdr_maps if fdr and metric in ['FDR q-val', 'fdr'] else None, legend_bbox_max=5)

    return (fig, axs) if enr_df is not None and len(region_filters) == 1 and len(rank_metrics) == 1 else None


def convert_short_index(short_index, celltype='PBMC'):
    assert short_index.str.contains('^BCG').all()
    assert short_index.str.contains('_V1$|_V2$|_V3$').all()
    return '300' + short_index + '_' + celltype


def compare_limma_models(celltype, model_restr, model_full, weighted=True, multiple_test='fdr_bh', strict=True):
    ssr, df_resid, k, n = {}, {}, {}, None
    for model in [model_restr, model_full]:
        _df_res_df = pd.read_csv(de_fn(celltype=celltype, model=model, data='df_residual'), index_col=0)
        assert (_df_res_df == _df_res_df.iloc[0, 0]).all().all()
        df_resid[model] = _df_res_df.iloc[0, 0]
        _res_df = pd.read_csv(de_fn(celltype=celltype, model=model, data='residuals'), index_col=0)
        if weighted:
            _w_df = pd.read_csv(de_fn(celltype=celltype, model=model, data='weights'), index_col=0)
        else:
            _w_df = 1
        ssr[model] = (_w_df * np.square(_res_df)).sum(axis=1)
        if n:
            assert n == _res_df.shape[1]
            assert not strict or samples == _res_df.columns.tolist()
        else:
            n = _res_df.shape[1]
            samples = _res_df.columns.tolist()
        _coef_df = pd.read_csv(de_fn(celltype=celltype, model=model, data='coefficients'), index_col=0)
        k[model] = _coef_df.shape[1]
        assert df_resid[model] == n - k[model]

    return compare_F_test_with_multiple_tests(ssr_restr=ssr[model_restr], ssr_full=ssr[model_full],
                                              df_restr=df_resid[model_restr], df_full=df_resid[model_full],
                                              multiple_test=multiple_test)


def ml_stats_tests(df, results_df, full_model, restr_models, pval_thr, X_visits, y_visits, estimator, score_name,
                   scoring, use_delong, n_delong_may_not_reject, model_to_str, average_roc=True, verbose=True):
    stars = defaultdict(lambda: {})
    if restr_models:
        for y in set(df['y']):
            full_model_score = df.loc[
                (df['y'] == y) & (df['X_name'] == full_model[0]) & (df['data'] == full_model[1]), score_name].values
            if len(full_model_score) > 0:  # maybe this y was not predicted by the full model
                assert len(full_model_score) == 1
                full_model_score = full_model_score[0]
                for restr_model in restr_models:
                    restr_model_score = df.loc[(df['y'] == y) & (df['X_name'] == restr_model[0]) & (
                            df['data'] == restr_model[1]), score_name].values
                    if len(restr_model_score) > 0:  # maybe this y was not predicted by the this model
                        assert len(restr_model_score) == 1
                        restr_model_score = restr_model_score[0]

                        _mask = (results_df['y'] == y) & (results_df['estimator'] == estimator) & (
                                results_df['X_visits'] == X_visits) & (results_df['y_visits'] == y_visits)

                        _restr_model = results_df.loc[
                            (results_df['X_name'] == restr_model[0]) & (results_df['data'] == restr_model[1]) & _mask]
                        _full_model = results_df.loc[
                            (results_df['X_name'] == full_model[0]) & (results_df['data'] == full_model[1]) & _mask]
                        _seeds = set(_restr_model['seed']).intersection(_full_model['seed'])
                        # have to sort by seed here
                        _restr_model = _restr_model.loc[_restr_model['seed'].isin(_seeds)].sort_values('seed')
                        _full_model = _full_model.loc[_full_model['seed'].isin(_seeds)].sort_values('seed')

                        # paired T-test
                        if len(_seeds) > 1:
                            t_pval = ttest_rel(_restr_model[score_name], _full_model[score_name])[1]
                        else:
                            t_pval = 1

                        # DeLong test
                        if use_delong and scoring == 'roc_auc':
                            if average_roc:
                                full_true, restr_true, full_samples, restr_samples = [], [], [], []
                                full_pred, restr_pred = pd.DataFrame(), pd.DataFrame()
                                for _restr_fn, _full_fn in zip(_restr_model['preds_fn'], _full_model['preds_fn']):
                                    _restr_data = np.load(_restr_fn, allow_pickle=True)
                                    _restr_true, _restr_pred = np.concatenate(_restr_data['y_true']), np.concatenate(
                                        _restr_data['y_pred'])
                                    _restr_samples = np.concatenate(_restr_data['samples'])
                                    _full_data = np.load(_full_fn, allow_pickle=True)
                                    _full_true, _full_pred = np.concatenate(_full_data['y_true']), np.concatenate(
                                        _full_data['y_pred'])
                                    _full_samples = np.concatenate(_full_data['samples'])
                                    assert np.array_equal(_restr_true, _full_true)
                                    assert np.array_equal(_restr_samples, _full_samples)

                                    sort = np.argsort(_full_samples)
                                    restr_samples.append(_restr_samples[sort])
                                    full_samples.append(_full_samples[sort])
                                    restr_true.append(_restr_true[sort])
                                    full_true.append(_full_true[sort])
                                    restr_pred = restr_pred.append(
                                        pd.Series(_restr_pred[sort], index=restr_samples[-1]), ignore_index=True)
                                    full_pred = full_pred.append(pd.Series(_full_pred[sort], index=full_samples[-1]),
                                                                 ignore_index=True)

                                for _r, _f in zip(restr_true, full_true):
                                    assert np.array_equal(_r, restr_true[0])
                                    assert np.array_equal(_f, full_true[0])

                                for _r, _f in zip(restr_samples, full_samples):
                                    assert np.array_equal(_r, restr_samples[0])
                                    assert np.array_equal(_f, full_samples[0])

                                restr_pred = restr_pred.T.rank(axis=0, method='average').mean(axis=1).values
                                full_pred = full_pred.T.rank(axis=0, method='average').mean(axis=1).values

                                delong_pval = np.power(10, delong_roc_test(restr_true[0], restr_pred, full_pred)[0, 0])

                                if verbose:
                                    print(y, full_model_score, roc_auc_score(restr_true[0], full_pred),
                                          restr_model_score, roc_auc_score(restr_true[0], restr_pred), delong_pval)

                                stars[y][model_to_str(restr_model)] = delong_pval < pval_thr

                            else:
                                delong_pvals = []
                                for _restr_fn, _full_fn in zip(_restr_model['preds_fn'], _full_model['preds_fn']):
                                    _restr_data = np.load(_restr_fn, allow_pickle=True)
                                    _restr_true, _restr_pred = np.concatenate(_restr_data['y_true']), np.concatenate(
                                        _restr_data['y_pred'])
                                    _full_data = np.load(_full_fn, allow_pickle=True)
                                    _full_true, _full_pred = np.concatenate(_full_data['y_true']), np.concatenate(
                                        _full_data['y_pred'])
                                    assert np.array_equal(_restr_true, _full_true)
                                    delong_pvals.append(
                                        np.power(10, delong_roc_test(_restr_true, _restr_pred, _full_pred)[0, 0]))
                                stars[y][model_to_str(restr_model)] = \
                                    sorted(delong_pvals)[
                                        -min(len(delong_pvals), 1 + n_delong_may_not_reject)] < pval_thr
                                if verbose:
                                    print(y, full_model_score, model_to_str(restr_model), restr_model_score,
                                          sorted(delong_pvals))
                        else:
                            # TODO some other test for regression? Bootstrapping? Not needed if differences are obvious...
                            stars[y][model_to_str(restr_model)] = t_pval < pval_thr
    return stars


def rename_old_cytokine(old_name, quality='good'):
    cyto, stimulus, _ = old_name.split('_')

    if cyto == 'TNF':
        cyto = 'TNF.a'

    if cyto.startswith('IL'):
        cyto = 'IL.' + cyto[2:]

    assert len(stimulus) == 2

    if stimulus[0] == 'T':
        time = '24h'
    elif stimulus[0] == 'W':
        time = '7d'
    else:
        raise ValueError('Unknown time "{}"'.format(stimulus))

    if stimulus[1] == '3':
        stimulus = 'S.aureus'
    elif stimulus[1] == '4':
        stimulus = 'MTB'
    else:
        raise ValueError('Unknown stimulus "{}"'.format(stimulus))

    new_name = 'CYTO:{stimulus}_{time}_PBMC_{cyto}_{quality}'.format(
        stimulus=stimulus, time=time, cyto=cyto, quality=quality)
    return new_name


def F_analysis(full, restricted, name, padj_method='fdr_bh', threshold=0.05, save_to_fn=None, verbose=True):
    df = pd.DataFrame(columns=['F.', 'p.value.', 'df.diff.'])
    df.columns = df.columns + name
    for c in full:
        if False and full[c].df_model == restricted[c].df_model:
            f_value, p_value, diff_df = None, None, 0
        else:
            df_restr = getattr(restricted[c], 'df_resid_inference', restricted[c].df_resid)
            df_full = getattr(full[c], 'df_resid_inference', full[c].df_resid)
            diff_df = df_restr - df_full
            f_value, p_value = compare_F_test(
                ssr_restr=np.sum(np.square(restricted[c].resid)), ssr_full=np.sum(np.square(full[c].resid)),
                df_restr=df_restr, df_full=df_full
            )
            try:
                _f_value2, _p_value2, _diff_df2 = full[c].compare_f_test(restricted[c])
                assert abs(f_value - _f_value2) < 1e-6
                assert abs(p_value - _p_value2) < 1e-6
                assert abs(diff_df - _diff_df2) < 1e-6
            except AttributeError:
                pass
        df = df.append(pd.Series([f_value, p_value, diff_df], index=df.columns, name=c))
    if padj_method:
        df = pd.concat([df, adjusted_pvals(df, coef=name, method=padj_method)], axis=1)
    if save_to_fn:
        df.to_csv(save_to_fn)

    p_col = (padj_col if padj_method else pval_col)(name)
    if verbose:
        print('F-test found {} hits for {} with {} {} {}'.format(
            (df[p_col] <= threshold).sum(), name, padj_method if padj_method else 'p-value', LESS_EQUAL, threshold))
    return df


def plot_var_expl(df, fn):
    fig, ax = plt.subplots(1, 1)
    df = df.reset_index('contrast')
    order = df.groupby('contrast')['FVE'].median().sort_values(ascending=False).index.tolist()
    if 'Residual' in order:
        order.remove('Residual')
        order.append('Residual')
    ax = sns.violinplot(data=df, x='contrast', y='FVE', order=order, cut=0, scale='width')
    ax = sns.stripplot(data=df, x='contrast', y='FVE', order=order, color='k')
    ax.set_ylim([-0.05, 1.05])
    savefig(fn)


def LR_analysis(model, celltype, annot_fn, Y_name, Y_regex, columns, contrasts=None,
                Y_transform=None, visits=None, design=None, lmm_groups=None, scale=False,
                fold_changes=None, remove_samples=None, permutation=None, do_not_correct=None,
                padj_method='fdr_bh', results_dir='results', save_outputs=False, save_models=False, fn_suffix=None,
                show_raw_distribution=False,
                reml=True, variance_partition=False,
                show_volcano=False,
                show_corr_distribution=False,
                show_fold_changes=False,
                cyto_short_stimuli_name=True, cyto_with_time=False,
                fdr=0.05, circle_size=40, color=None, alpha=1, fig_size=(3, 3),
                do_not_report_intercept=True, return_corrected_X=False, var_groups=None, use_imputed_cytokines=False,
                verbose=1):

    if verbose > 1:
        print('* * * * * * * * * * * * * * * * * * * * * * * * * * *\n'
              '* Running a linear regression analysis for {}\n'
              '* * * * * * * * * * * * * * * * * * * * * * * * * * *\n'.format(Y_name))

    if not visits:
        visits = ['V1', 'V2', 'V3']
    if isinstance(visits, str):
        visits = [visits]
    if isinstance(columns, str):
        columns = [columns]
    if save_outputs:
        fn_template = lr_fn(celltype, model, Y_name, data='{data}', extras=fn_suffix, ext='{ext}', results_dir=results_dir)
        make_dir(results_dir, 'LR', '{}.{}'.format(celltype, model), 'models')
    annot_df = get_sample_annot(fn=annot_fn)

    if use_imputed_cytokines:
        _imputed_df = pd.read_csv(IMPUTED_CYTO_FN, index_col=0)
        annot_df = annot_df.loc[_imputed_df.index]
        annot_df.loc[:, _imputed_df.columns] = _imputed_df

    if remove_samples:
        if verbose:
            print('Removing up to {} samples'.format(len(remove_samples)))
        annot_df = annot_df.loc[~annot_df.index.isin(remove_samples)]

    Y_df = annot_df.loc[(annot_df['SAMPLE:TISSUE'] == celltype) & annot_df['SAMPLE:VISIT'].isin(visits),
                        annot_df.columns.str.contains(Y_regex)].copy()
    if Y_transform:
        if verbose:
            print('Transforming Y with', Y_transform)
        Y_df = Y_transform(Y_df)

    if show_raw_distribution or save_outputs:
        for y in Y_df.columns:
            values = Y_df.loc[~Y_df[y].isnull(), y]
            if len(values) > 0:
                ax = sns.distplot(values, label=y)
        ax.set_xlabel(Y_name)
        ax.set_ylabel('Density')
        ax.legend().set_visible(Y_df.shape[1] < 11)
        sns.despine()
        if save_outputs:
            savefig(fn_template.format(data='histogram', ext='pdf'))
        if show_raw_distribution:
            plt.show()

    _cast_booleans = {c: 'Int64' for c in columns if is_bool_dtype(annot_df[c].dtype)}
    X_df = annot_df.loc[Y_df.index, columns].copy().astype(_cast_booleans)

    if fold_changes:
        assert X_df.index.equals(Y_df.index)
        X_df = X_df.set_index(['SAMPLE:DONOR', fold_changes[0]])
        Y_df.index = X_df.index
        diff_args = dict(diff_col=fold_changes[0], new=fold_changes[2], base=fold_changes[1])
        # X_df = df_diff(X_df, **diff_args)
        _diffs = []
        _blood_correction = X_df.columns.str.contains('^PBMC_PERC:|^WB_PER_ML:')
        _diffs.append(df_diff(X_df.loc[:, ~_blood_correction], diff_op='-', **diff_args))
        _diffs.append(np.log2(df_diff(X_df.loc[:, _blood_correction], diff_op='/', **diff_args)))
        X_df = pd.concat(_diffs, axis=1)
        Y_df = df_diff(Y_df, diff_op='-', **diff_args)

    X_df, Y_df = drop_na_for_LM(X_df, Y_df, verbose=verbose)
    if verbose:
        print('Targets: {} samples, {} variables'.format(Y_df.shape[0], Y_df.shape[1]))

    _LM_results = fit_linear_model(
        X_df, Y_df, design=design if design else '1 + {}'.format(' + '.join(replace_chars_for_LM(X_df.columns))),
        lmm_groups=lmm_groups, scale=scale, reml=reml, variance_partition=variance_partition, contrasts=contrasts,
        permutation=permutation, do_not_correct=do_not_correct, return_corrected_X=return_corrected_X,
        var_groups=var_groups, random_state=RANDOM_STATE, verbose=verbose
    )
    models, results_df = _LM_results[0], _LM_results[1]
    if contrasts:
        contrasts_df = _LM_results[2]
    if return_corrected_X:
        corrected_df = _LM_results[3]
        if save_outputs:
            corrected_df.to_csv(fn_template.format(data='corrected', ext='csv'))
    var_df = _LM_results[-1]

    if do_not_report_intercept:
        results_df = results_df.loc[results_df.index.get_level_values('contrast') != 'Intercept']

    if padj_method and len(results_df) > 0:
        def _adjust_pvalues(df):
            return pd.concat(
                [adjusted_pvals(df.loc[df.index.get_level_values('contrast') == c, 'p.value'],
                                method=padj_method) for c in set(df.index.get_level_values('contrast'))]) \
                if df.shape[1] != 0 else pd.DataFrame()

        results_df = pd.concat([results_df, _adjust_pvalues(results_df)], axis=1)
        if contrasts:
            contrasts_df = pd.concat([contrasts_df, _adjust_pvalues(contrasts_df)], axis=1)

    if save_outputs:
        results_df.to_csv(fn_template.format(data='coefs', ext='csv'))
        if contrasts:
            contrasts_df.to_csv(fn_template.format(data='contrasts', ext='csv'))
        var_df.to_csv(fn_template.format(data='variance', ext='csv'))
        if 'FVE' in var_df:
            plot_var_expl(var_df, fn_template.format(data='variance', ext='pdf'))

    if save_models:
        if lmm_groups is None:
            dump_pickle(fn_template.format(data='models', ext='pickle'), models)
        else:
            import pymer4
            fn = lr_fn(celltype, model, Y_name, data='{data}', extras=fn_suffix, ext='{ext}', subdir='models', results_dir=results_dir)
            for m in models:
                pymer4.io.save_model(models[m], fn.format(data='model_{}'.format(m), ext='hdf5'))

    _return = [models, results_df.sort_index(axis=1)]
    if contrasts:
        _return.append(contrasts_df.sort_index(axis=1))
    if return_corrected_X:
        _return.append(corrected_df)
    _return.append(var_df)
    return _return


def convert_partial_year(number):
    year = int(number)
    d = datetime.timedelta(days=(number - year) * 365)
    day_one = datetime.datetime(2017 + year, 1, 1)
    date = d + day_one
    return date


def make_annot(df, target, contrast, thr=0.05, fdr_col='global_fdr'):
    if (target, contrast) in df.index:
        padj = df.loc[(target, contrast), fdr_col]
    else:
        padj = None
    if padj and padj <= thr:
        return '{:.0e}'.format(padj)
    else:
        return ''


RENAME_CYTO = pd.read_excel(STIMULI_NAMES_FN, sheet_name='cytokine', header=None)[[0, 1]].set_index(1)[0].to_dict()

RENAME_CM = pd.read_excel(STIMULI_NAMES_FN, sheet_name='circMed', header=None)[[0, 1]].set_index(1)[0].to_dict()

RENAME_BLOOD = {
    'WBC': 'white-blood count',
    'LYMPHO': 'lymphocyte',
    'MONO': 'monocyte (CD14+)',
    'MONO_CLASSICAL': 'monocyte (CD14++CD16{})'.format(MINUS),
    'MONO/CLASSICAL': 'monocyte (CD14++CD16{})'.format(MINUS),
    'MONO_INTERMEDIATE': 'monocyte (CD14+CD16+)',
    'MONO/INTERMEDIATE': 'monocyte (CD14+CD16+)',
    'MONO_NON_CLASSICAL': 'monocyte (CD14++CD16+)',
    'MONO/NON_CLASSICAL': 'monocyte (CD14++CD16+)',
    'T': 'T cell (CD3+CD56{})'.format(MINUS),
    'T_CD8': 'T cell (CD8+)',
    'T/CD8': 'T cell (CD8+)',
    'T_CD4': 'T cell (CD4+)',
    'T/CD4': 'T cell (CD4+)',
    'T_CD4_TREG': 'T cell (CD4+CD25++)',
    'T/CD4/TREG': 'T cell (CD4+CD25++)',
    'T_DN': 'T cell (CD4{m}CD8{m})'.format(m=MINUS),
    'T/DN': 'T cell (CD4{m}CD8{m})'.format(m=MINUS),
    'T_DP': 'T cell (CD4+CD8+)',
    'T/DP': 'T cell (CD4+CD8+)',
    'B': 'B cell (CD19+)',
    'NK': 'NK cell (CD3{}CD56+)'.format(MINUS),
    'NK/BRIGHT': 'NK cell (CD3{m}CD56++CD16{m})'.format(m=MINUS),
    'NK_BRIGHT': 'NK cell (CD3{m}CD56++CD16{m})'.format(m=MINUS),
    'NK/DIM': 'NK cell (CD3{}CD56+CD16+)'.format(MINUS),
    'NK_DIM': 'NK cell (CD3{}CD56+CD16+)'.format(MINUS),
    'NK/OTHER': 'NK cell (CD3{m}CD56+CD16{m})'.format(m=MINUS),
    'NK_OTHER': 'NK cell (CD3{m}CD56+CD16{m})'.format(m=MINUS),
    'NKT': 'NKT cell (CD3+CD56+)',
    'NEUTRO': 'neutrophil',
    'BASO': 'basophil',
    'IG': 'immature granulocyte',
    'EO': 'eosinophil',
}

RENAME_FACTOR = {
    'VISIT_DATE_REAL': 'visit date',
    'IC_DATE_REAL': 'vaccination date',
    'VISIT_TIME_REAL': 'circadian rhythm',
    'IC_TIME_REAL': 'vaccination time',
    'AGE': 'age',
    'BMI': 'BMI',
    'SEX': 'sex',
    'oralContraceptivesIncludingMen': 'oral contraceptives', # eptives',
    'currentSmoker': 'smoker',
    'currentMenopausalIncludingMen': 'in menopause',
    'alcoholInLast24h': 'alcohol',
    'IC_alcoholInLast24h': 'alcohol before vaccination',
    'sleepHoursLastNight': 'sleep before visit',
    'IC_sleepHoursLastNight': 'sleep before vaccination',
    'motherBcg': 'mother with BCG',
    'sport_yesNo24hBeforeAppointment': 'sport before visit',
    'IC_sport_yesNo24hBeforeAppointment': 'sport before vaccination',
    'stress_dayBefore': 'stress before visit',
    'IC_stress_dayBefore': 'stress before vaccination',
    'sizeOfVaccinationSpot_v2': 'scar size (d14)',
    'scarSize_v3': 'scar size (d90)',
    'SEASON': 'seasonal effects'
}

RENAME_FACTOR_EXPLAIN = {
    'AGE': f'age\n(18 {EM_DASH} 71 years)',
    'SEX': f'sex\n(female {EM_DASH} male)',
    'BMI': f'BMI\n(17.8 {EM_DASH} 34.2 kg/m2)',
    'oralContraceptivesIncludingMen': f'oral contraceptive use\n(no {EM_DASH} yes)',
    'alcoholInLast24h': f'alcohol within last 24h\n(no {EM_DASH} yes)',
    'VISIT_TIME_REAL': f'circadian rhythm\n(8am {EM_DASH} 12pm)',
}

PEAK_ANNOT_COLS = {
    'region_id': 'Region ID',
    'chr': 'Chromosome',
    'start': 'Start',
    'end': 'End',
    'length': 'Length',
    'gene_id': 'Ensembl gene ID',
    'gene_name': 'Gene symbol',
    'genomic_feature': FEATURE_TYPE_COL,
    'feat_type': f'{FEATURE_TYPE_COL} (fine-grain)',
    'feat_anchor': 'Feature anchor',
    'distance': 'Distance to anchor',
    'location': 'Location relative to feature',
    'genomic_location': GENOMIC_LOCATION_COL,
    'characterization': f'{GENOMIC_LOCATION_COL} (fine-grain)',
    'reg_feature_id': f'{REGULATORY_ANNOT_COL} feature ID',
    'reg_feature': f'{REGULATORY_ANNOT_COL} feature',
}

ATAC_QC_COLS = {
    'SAMPLE:DONOR': 'Donor',
    'SAMPLE:VISIT': 'Visit',
    'SAMPLE:TISSUE': 'Cell type',

    # DEMUX and LAB
    'RUN:PROTOCOL': 'Protocol',
    'DEMUX:FLOWCELL': 'Flowcell',
    'LAB:BATCH': 'Batch ID',
    'QC:PASS': 'QC pass',
    'LAB:CQ': 'PCR Cq',

    # QC
    'RUN:FILTERED_PEAKS': 'Filtered peaks',
    'RUN:TSS_ENRICHMENT': 'TSS enrichment',
    'RUN:FRIP': 'FRiP',
    'RUN:ORACLE_FRIP': 'Regulatory Build FRiP',
    'QC:PROMOTER_FRIP': 'Promoter FRiP',

    # RUN
    'RUN:FASTQC_GC_PERC': 'FastQC GC percentage',
    'RUN:FASTQC_READ_LENGTH': 'FastQC read length',
    'RUN:FASTQC_TOTAL_PASS_FILTER_READS': 'FastQC total pass filter reads',
    'RUN:TRIM_TRIM_LOSS_PERC': 'Trim-loss percentage',
    'RUN:TRIM_TRIMMED_PERC': 'Trimmed percentage',
    'RUN:TRIM_UNTRIMMED_PERC': 'Not-trimmed percentage',
    'RUN:UNIQUE_ALIGNED_PERC': 'Uniquely-aligned percentage',
    'RUN:MULTIPLE_ALIGNED_PERC': 'Not-uniquely-aligned percentage',
    'RUN:NOT_ALIGNED_PERC': 'Not-aligned percentage',
    'RUN:DUPLICATE_PERCENTAGE': 'Duplicate percentage',
    'RUN:MT_MAPPED_READS': 'MT-mapped reads',
    'RUN:MT_DUPLICATE_PERCENTAGE': 'MT-duplicate percentage',
    'RUN:PEAKS': 'Peaks',
}

ATAC_QC_INTEGER_COLS = [
    'RUN:PEAKS', 'RUN:FILTERED_PEAKS', 'RUN:MT_MAPPED_READS', 'RUN:FASTQC_TOTAL_PASS_FILTER_READS',
    'RUN:FASTQC_READ_LENGTH', 'RUN:FASTQC_GC_PERC'
]

DONOR_COLS = {
    # NECESSARY
    'SAMPLE:DONOR': 'Donor',
    'SAMPLE:VISIT': 'Time point',

    'DONOR:AGE': 'Age',
    'DONOR:SEX': 'Sex',
    'DONOR:BMI': 'BMI',
    'DONOR:oralContraceptivesIncludingMen': 'Oral contraceptives',
    'SAMPLE:alcoholInLast24h': 'Alcohol within last 24h',

    'DONOR:scarSize_v3': 'Scar size (d90)',
    'DONOR:sizeOfVaccinationSpot_v2': 'Scar size (d14)',

    'DONOR:IC_DATE': 'Vaccination date',
    'DONOR:IC_TIME': 'Vaccination time',
    'DONOR:IC_EVENING': 'Evening vaccination',

    'SAMPLE:VISIT_DATE': 'Sample collection date',
    'SAMPLE:VISIT_TIME': 'Sample collection time',

    # TECHNICAL AND USEFUL
    'SAMPLE:EXCLUSION': 'Exclusion',
    'QC:PASS': 'ATAC-seq QC pass',
    'DONOR:SNP_OUTLIER': 'SNP outlier',
    'DONOR:CYTO_BATCH': 'Cytokine batch ID',
    'LAB:BATCH': 'ATAC-seq batch ID',

    # NOT USED IN A PAPER -- SHARE AS A RESOURCE?
    'DONOR:HEIGHT': 'Height',
    'DONOR:WEIGHT': 'Weight',

    'DONOR:alcohol_unitsWeekAverage': 'Alcohol-units per week',
    'SAMPLE:alcohol_unitsLast24h': 'Alcohol-units within last 24h',

    'DONOR:postMenopausal': 'Post menopause',

    'DONOR:qualityTeeth': 'Teeth quality',

    'DONOR:currentSmoker': 'Current smoker',
    'DONOR:packYears': 'Smoking pack-years',

    'DONOR:exerciseDuringDay': 'Exercise level per day',
    'DONOR:sport_hoursPerWeek': 'Sport-hours per week',
    'DONOR:sport_timesPerMonth': 'Sport-units per month',
    'SAMPLE:sport_hoursIn24hBeforeAppointment': 'Sport within last 24h',

    'DONOR:vaginalBirth': 'Vaginal birth',
    'DONOR:motherBcg': 'Mother BCG vaccine',
    'DONOR:motherPox': 'Mother pox vaccine',
    'DONOR:fatherBcg': 'Father BCG vaccine',
    'DONOR:fatherPox': 'Father pox vaccine',

    'DONOR:happiness': 'General happiness',
    'DONOR:mentalHealth_3weeksBefore': 'Mental health 3 weeks before vaccination',
    'DONOR:stress_2weeksBefore': 'Stress 2 weeks before vaccination',
    'SAMPLE:stress_dayBefore': 'Stress within last 24h',

    'DONOR:sleepPerNightHours': 'Sleep-hours per night',
    'DONOR:sleepQuality_lastTwoWeeks_v1': 'Sleep quality 2 weeks before vaccination',
    'SAMPLE:sleepHoursLastNight': 'Sleep-hours within last 24h'
}

DONOR_INTEGER_COLS = [
    'DONOR:HEIGHT', 'DONOR:exerciseDuringDay', 'DONOR:sport_timesPerMonth', 'DONOR:happiness',
    'DONOR:mentalHealth_3weeksBefore', 'DONOR:stress_2weeksBefore', 'SAMPLE:stress_dayBefore', 'DONOR:qualityTeeth',
    'DONOR:sleepQuality_lastTwoWeeks_v1'
]

DONOR_BOOLEAN_COLS = [
    'DONOR:oralContraceptivesIncludingMen', 'SAMPLE:alcoholInLast24h', 'DONOR:IC_EVENING', 'QC:PASS',
    'DONOR:currentSmoker', 'DONOR:motherBcg', 'DONOR:motherPox', 'DONOR:fatherBcg', 'DONOR:fatherPox'
]


def camel_case(s):
    if len(s) < 2:
        return s.lower()
    else:
        return s[0] + ''.join([w[0].upper() + w[1:] for w in s.replace('-', ' ').replace('.', ' ').split()])[1:]


def rename_factor(factor, caps=True, multiline=False, short_rhythm=False, short_contraceptives=False,
                  short_alcohol=False, explain_factors=False, include_V1_prefix=True,
                  use_sample_collect_time=False, camel_case=False):
    if factor in RENAME_FACTOR:
        if explain_factors:
            factor = RENAME_FACTOR_EXPLAIN[factor]
        else:
            factor = RENAME_FACTOR[factor]
            if use_sample_collect_time:
                factor = factor.replace('circadian rhythm', 'sample-collection time')
            if short_rhythm:
                factor = factor.replace('circadian rhythm', 'circadian rh.').replace('vaccination time', 'vacc. time')
            if short_contraceptives:
                factor = factor.replace('oral contraceptives', 'oral contrac.')
            if short_alcohol:
                factor = factor.replace('alcohol before vaccination', 'alcohol before vacc.')
            if multiline:
                factor = factor.replace(' ', '\n')
    else:
        vaccination = factor.startswith('IC_')
        if vaccination:
            factor = factor.replace('IC_', '')
        if factor.startswith('PBMC_PERC_'):
            factor = rename_blood(factor.replace('PBMC_PERC_', ''))
        elif factor.startswith('PBMC_PER_ML_'):
            factor = rename_blood(factor.replace('PBMC_PER_ML_', '')) + ' (PBMC per ml)'
        elif factor.startswith('WB_PERC_'):
            factor = rename_blood(factor.replace('WB_PERC_', '')) + ' (whole-blood %)'
        elif factor.startswith('WB_PER_ML_'):
            factor = rename_blood(factor)
        elif factor.startswith('CM_'):
            factor = rename_CM(factor)
        elif factor.startswith('CYTO_'):
            factor = rename_cytokine(factor)
        else:
            pass
        if include_V1_prefix and vaccination:
            factor = 'V1 ' + factor

    if caps:
        factor = factor[0].upper() + factor[1:]

    if camel_case:
        factor = camel_case(factor)

    return factor


def rename_factors(factors, **kwargs):
    if isinstance(factors, str):
        return rename_factor(factors, **kwargs)
    else:
        return np.asarray([rename_factor(factor, **kwargs) for factor in factors])


RENAME_Y = {
    'CYTO': 'cytokine',
    'CORR_CYTO': 'corrected cytokine',
    'CM': 'inflammatory marker',
    'CORR_CM': 'corrected inflammatory marker',
    'WB_PERC': 'blood percentage',
    'WB_PER_ML': 'blood count',
    'PBMC_PERC': 'PBMC percentage',
    'PBMC_PER_ML': 'PBMC count',
    'SEASON': 'seasonality',
    'ATAC': 'chromatin accessibility',
    'SCORE': 'vaccination score'
}


def rename_regions(index, peaks_df=None):
    assert isinstance(index, pd.Index)
    if peaks_df is None:
        peaks_df = get_peak_annot()
    return pd.Index(
        peaks_df.loc[index, 'chr'].astype(str)
        + '_' + peaks_df.loc[index, 'start'].astype(str)
        + '_' + peaks_df.loc[index, 'end'].astype(str), name=REGION_COL
    )


def renamed_regions_index_with_annotation(index, peaks_df=None, extra_levels=None):
    assert isinstance(index, pd.Index)
    if peaks_df is None:
        peaks_df = get_peak_annot()

    multi_index = [rename_regions(index, peaks_df=peaks_df.loc[index])]

    for region_filter in [TSS_PROXIMAL, GENE_AND_DISTAL_10kb]:
        mapping = get_region_to_gene_mapping(peaks_df.loc[index], region_filter)
        multi_index.append(pd.Index(mapping.values, name=mapping.name))

    if extra_levels is not None:
        multi_index.extend(extra_levels)

    return pd.MultiIndex.from_arrays(multi_index)


def rename_Y(Y_name, caps=True, plural=True):
    if Y_name in RENAME_Y:
        Y_name = RENAME_Y[Y_name]
        if caps:
            Y_name = Y_name[0].upper() + Y_name[1:]
        if plural and not Y_name.endswith('lity') and not Y_name.endswith('date'):
            Y_name += 's'
    return Y_name


def rename_CM(protein):
    assert protein.startswith('CM:') or protein.startswith('CM_')
    return RENAME_CM[protein.split('_')[-1]]


def rename_CMs(protein):
    if isinstance(protein, str):
        return rename_CM(protein)
    else:
        return np.asarray([rename_CM(p) for p in protein])


def rename_blood(celltype, plural=True, drop_cell_suffix=False, short_mono=False, capitalize=False):
    assert celltype.startswith('WB_PER_ML:') or celltype.startswith('WB_PER_ML_') or celltype in RENAME_BLOOD, celltype
    if celltype.startswith('WB_PER_ML'):
        celltype = celltype[len('WB_PER_ML') + 1:]
    blood = RENAME_BLOOD[celltype]
    if plural:
        blood = blood.replace('cyte', 'cytes').replace('cell', 'cells').replace('phil', 'phils')
    if drop_cell_suffix:
        blood = blood.replace('cells', '').replace('cell', '')
    if short_mono:
        blood = blood.replace('monocytes', 'mono.').replace('monocyte', 'mono.')
    if capitalize:
        blood = '{}{}'.format(blood[0].upper(), blood[1:])
    return blood


def rename_bloods(celltype, **kwargs):
    if isinstance(celltype, str):
        return rename_blood(celltype, **kwargs)
    else:
        return np.asarray([rename_blood(c, **kwargs) for c in celltype])


def rename_cytokine(cyto, short_stimuli_name=True, with_time=False, only_cyto=False, only_stimulus=False,
                    only_time=False, remove_prefix=True, capitalize=True, time_stim_prot=False, stimuli_italics=True):
    assert cyto.startswith('CYTO:') or cyto.startswith('CYTO_')
    if only_time:
        with_time = True
    _MT = MT if short_stimuli_name else 'M. tuberculosis'
    _SA = SA if short_stimuli_name else 'S. aureus'
    _CA = CA if short_stimuli_name else 'C. albicans'
    if stimuli_italics and not short_stimuli_name:
        _MT = f'${_MT}$'
        _SA = f'${_SA}$'
        _CA = f'${_CA}$'
    _RENAME_STIMULUS = {'MTB': _MT, 'C.albicans.yeast': _CA, 'S.aureus': _SA, 'LPS.100ng': LPS}

    stimulus, duration, _, protein, _ = cyto[5:].split('_')
    stimulus = _RENAME_STIMULUS[stimulus]
    protein = RENAME_CYTO[protein]

    if time_stim_prot:
        cyto = '{}_{}_{}'.format(duration, stimulus, protein)
    else:
        if capitalize:
            protein = protein[0].upper() + protein[1:]

        if only_cyto:
            cyto = protein
        elif only_stimulus:
            cyto = '{}{}'.format(stimulus, ', {}'.format(duration) if with_time else '')
        elif only_time:
            cyto = duration
        else:
            cyto = '{} ({}{})'.format(protein, stimulus, ', {}'.format(duration) if with_time else '')

        if not remove_prefix:
            cyto = 'CYTO:' + cyto

    return cyto


def rename_cytokines(cyto, **kwargs):
    if isinstance(cyto, str):
        return rename_cytokine(cyto, **kwargs)
    else:
        return np.asarray([rename_cytokine(c, **kwargs) for c in cyto])


def rename_Y_vars(Y_name, var, caps_for_blood=True, **kwargs):
    if Y_name == 'CYTO':
        rename_func = rename_cytokines
        allowed_kwargs = list(inspect.signature(rename_cytokine).parameters)
        kwargs = {arg: kwargs[arg] for arg in kwargs if arg in allowed_kwargs}
    elif Y_name == 'CM':
        rename_func = rename_CMs
        allowed_kwargs = list(inspect.signature(rename_CM).parameters)
        kwargs = {arg: kwargs[arg] for arg in kwargs if arg in allowed_kwargs}
    elif Y_name == 'WB_PER_ML':
        rename_func = rename_bloods
        allowed_kwargs = list(inspect.signature(rename_blood).parameters)
        kwargs = {arg: kwargs[arg] for arg in kwargs if arg in allowed_kwargs}
        if caps_for_blood:
            assert 'capitalize' not in kwargs or kwargs['capitalize']
            kwargs['capitalize'] = True
    else:
        raise ValueError
    return rename_func(var, **kwargs)


def rename_targets(targets):
    if all([x.startswith('CYTO:') for x in targets]):
        return rename_cytokines(targets)
    elif all([x.startswith('CM:') for x in targets]):
        return rename_CMs(targets)
    elif all([x.startswith('WB_PER_ML:') for x in targets]):
        return rename_bloods(targets)
    else:
        raise ValueError


def argsort_cytokines(index, short_form=False):
    if not short_form:
        index = pd.Index(rename_cytokines(index))
    return index.str.replace('IFN-γ \(', 'IFN-γ Zzz(').str.replace('IL-17 \(', 'A-IL-17 Zzz(').str.replace(MT, 'Zzz').str.split(' ', expand=True).sortlevel([1, 0])[1]


def rename_pathway(pthw, capitalize=True):
    pthw = pthw.replace(
         'interferon', 'IFN').replace(
         'gamma', GAMMA).replace(
         'Fc gamma R', 'FcγR').replace(
         'tumor necrosis factor', 'TNF').replace(
         'epidermal growth factor', 'EGF').replace(
         # 'Staphylococcus aureus', 'S. aureus').replace(
         'Herpes simplex virus 1', 'HSV-1').replace(
         'Kaposi sarcoma-associated herpesvirus', 'KSHV').replace(
         'polymerase II', 'pol2').replace(
         'interleukin-10', 'IL-10').replace(
         'vascular endothelial growth factor', 'VEGF')
         # .replace(
         # 'negative regulation of ', 'neg. ').replace(
         # 'positive regulation of ', 'pos. ').replace(
         # 'regulation of ', 'reg. ').replace(
         # 'cellular response to ', 'cell resp. ').replace(
         # 'activation of ', '').replace(
         # 'response to ', 'resp. ')

    if pthw.endswith(' pathway'):
        pthw = pthw[:-len(' pathway')]
    # if pthw.endswith(' process'):
    #     pthw = pthw[:-len(' process')]

    if capitalize and not (pthw.startswith('mRNA') or pthw.startswith('rRNA') or pthw.startswith('lncRNA')):
        pthw = pthw[0].upper() + pthw[1:]

    return pthw


def rename_pathways(pthw, **kwargs):
    if isinstance(pthw, str):
        return rename_pathway(pthw, **kwargs)
    else:
        return np.asarray([rename_pathway(p, **kwargs) for p in pthw])
    
    
def _rename_targets_and_factors(multi_index, Y_name, only_cyto, **kwargs):
    # Rename everything
    if Y_name == 'CYTO':
        renamed_targets = rename_cytokines(multi_index.get_level_values('target'),
                                           only_cyto=only_cyto)
    elif Y_name == 'CM':
        renamed_targets = rename_CMs(multi_index.get_level_values('target'))
    elif Y_name == 'WB_PER_ML':
        renamed_targets = rename_bloods(multi_index.get_level_values('target'))
    else:
        renamed_targets = multi_index.get_level_values('target')

    multi_index = pd.MultiIndex.from_arrays([
        pd.Index(renamed_targets, name='target'),
        pd.Index(rename_factors(multi_index.get_level_values('contrast').str.replace(
            'DONOR_', '').str.replace('SAMPLE_', '').str.replace('\[T\.M\]$', ''), **kwargs), name='contrast')
    ])
    return multi_index


def factor_analysis(
        celltype, model, Y_name, factors_kind, annot_col, res_fn=None,
        visits=False, visit_interaction=False, fold_changes=False, LMM=False, scale=False, remove_evening=False, subset_factors=False, subset_targets=False, exclude_targets=False,
        factor_volcanos=False, joint_volcano=False, heatmap=True, fdr=0.05, age_effect_per_N_years=10, color_by_stimulus=False, marker_by_duration=False,
        scatter_kws={}, legend_kws={}, volcano_kws={}, annot_size=10, max_annot=10,
        cell_width=0.5, cell_height=0.5, names_without_time=False, names_without_stimulus=False, blank_no_significance=False,
        pval_asterisks=None, pval_fmt='.0e', factor_rename_kws={}, force_order=None,
        cbar_label='Signed {}log10{}P-value'.format(MINUS, NBSPC), annot_kws={}, cbar_kws={}, heatmap_kws={},
        heatmap_title=None, heatmap_xlabel=None, heatmap_ylabel=None, explain_heatmap_annot=True,
        padj_method='fdr_bh', transpose=False, fig_dir=None, save_fig=False, show_fig=True, fig_format='svg', dpi=DPI,
):
    _return = []
    if isinstance(fdr, str):
        fdr = [fdr]

    if Y_name == 'SCORE' and visits != ['V1']:
        return

    if visits == ['V1']:
        remove_evening = Y_name == 'SCORE'

    if res_fn is None:
        res_fn = summary_lr_fn(celltype=celltype, model=model, Y_name=Y_name, visits=visits,
                           fold_changes=fold_changes, visit_interaction=visit_interaction, LMM=LMM, scale=scale, remove_evening=remove_evening,
                           **({factors_kind: True} if factors_kind and factors_kind.startswith('include_') else {'extras': factors_kind}  if factors_kind else {}))
        assert os.path.basename(res_fn).startswith('LR_results_')
        assert res_fn.endswith('.csv.gz') or res_fn.endswith('.csv')
    fig_fn = os.path.join(fig_dir if fig_dir else os.path.dirname(res_fn),
                          os.path.basename(res_fn).replace(
                              'LR_results_', 'LR_{data}_').replace('.csv.gz', '.{ext}').replace('.csv', '.{ext}'))

    if visits:
        print('\n----------------\nV I S I T --- {}\n----------------'.format('_'.join(visits)))
    elif visit_interaction:
        print('\n----------------------------\nI N T E R A C T I O N --- {}\n----------------------------'.format(visit_interaction))
    elif fold_changes:
        print('\n------------------------------\nF O L D - C H A N G E S --- {}\n------------------------------'.format(fold_changes))
    elif LMM:
        print('\n----------------------------\nL M M\n----------------------------')
    else:
        print('\n----------------------------\n F I X E D    E F F E C T S\n----------------------------')

    if not os.path.exists(res_fn):
        print('MISSING:', res_fn)
    else:
        print(res_fn)

    res_df = pd.read_csv(res_fn, index_col=['target', 'contrast'])

    # rename interaction contrasts so that they are simple, short factor names
    if visit_interaction:
        res_df = res_df.loc[res_df.index.get_level_values('contrast').str.contains('SAMPLE_VISIT\[T\.{}\]'.format(visit_interaction))]
        res_df.index = pd.MultiIndex.from_arrays(
            [res_df.index.get_level_values('target'),
             res_df.index.get_level_values('contrast').str.replace('SAMPLE_VISIT\[T\.{}\]:'.format(visit_interaction), '')])

    # rename season factor to "SEASON" and fill in "Coef" with random positive effect size
    if factors_kind == 'include_season':
        res_df.index = pd.MultiIndex.from_arrays(
            [res_df.index.get_level_values('target'),
             res_df.index.get_level_values('contrast').str.replace(
                 'DONOR_IC_DATE_2PI_SIN__DONOR_IC_DATE_2PI_COS', 'SEASON').str.replace(
                 'SAMPLE_VISIT_DATE_2PI_SIN__SAMPLE_VISIT_DATE_2PI_COS', 'SEASON').str.replace(
                 'DONOR_', '').str.replace('SAMPLE_', '').str.replace('\[T\.M\]$', '')])
        res_df.loc[res_df.index.get_level_values('contrast') == 'SEASON', 'Coef'] = np.random.rand((res_df.index.get_level_values('contrast') == 'SEASON').sum())
        res_df = res_df.loc[res_df.index.get_level_values('contrast').str.contains('SEASON')]

    # select a subset of contrasts to show
    _contrasts = res_df.index.get_level_values('contrast')
    if subset_factors == 'auto':
        mask = np.ones((res_df.shape[0],), dtype=bool)
        if factors_kind == 'include_scores':
            if Y_name == 'CYTO':
                mask &= ~_contrasts.str.contains('^thm')
        elif factors_kind != 'include_season':
            mask &= _contrasts != 'SEASON'
        else:
            if Y_name != 'SCORE':
                mask &= ~_contrasts.str.contains('^scar|^sizeOfVac')
        mask &= ~_contrasts.str.contains('[sS]leep|[sS]port|[sS]tress|[sS]moker|[mM]enopausal|[fF]ather|[pP]arent')
        mask &= ~_contrasts.str.contains('binary$|_FC1.2_responder$')
        if not (visit_interaction or fold_changes):
            mask &= ~_contrasts.str.contains('mother')
        res_df = res_df.loc[mask]
    elif subset_factors:
        if isinstance(subset_factors, str):
            subset_factors = [subset_factors]
        res_df = res_df.loc[_contrasts.isin(
            ['{}[T.M]'.format(f) if f == 'DONOR_SEX' else f.replace('[T_', '[T.') for f in replace_chars_for_LM(subset_factors)])]

    if subset_targets:
        if isinstance(subset_targets, str):
            res_df = res_df.loc[res_df.index.get_level_values('target').str.contains(subset_targets)]
        else:
            res_df = res_df.loc[res_df.index.get_level_values('target').isin(subset_targets)]

    if exclude_targets:
        if isinstance(exclude_targets, str):
            res_df = res_df.loc[~res_df.index.get_level_values('target').str.contains(exclude_targets)]
        else:
            res_df = res_df.loc[~res_df.index.get_level_values('target').isin(exclude_targets)]

    # drop the obvious
    if Y_name == 'SCORE' and factors_kind is None:
        res_df = res_df.drop([
            ('DONOR:scarSize_v3', 'sizeOfVaccinationSpot_v2'),
            ('DONOR:sizeOfVaccinationSpot_v2', 'scarSize_v3')
        ])

    n_contrasts = len(set(res_df.index.get_level_values('contrast')))
    n_targets = len(set(res_df.index.get_level_values('target')))

    # adjusted p-values
    res_df['contrast_fdr'] = np.nan
    for trgt in np.unique(res_df.index.get_level_values('target')):
        _this_target = (res_df.index.get_level_values('target') == trgt)
        res_df.loc[_this_target, 'contrast_fdr'] = adjusted_pvals(res_df.loc[_this_target, 'p.value'], method=padj_method)
    res_df['target_fdr'] = np.nan
    for cntrst in np.unique(res_df.index.get_level_values('contrast')):
        _this_contrast = (res_df.index.get_level_values('contrast') == cntrst)
        res_df.loc[_this_contrast, 'target_fdr'] = adjusted_pvals(res_df.loc[_this_contrast, 'p.value'], method=padj_method)
    res_df['global_fdr'] = adjusted_pvals(res_df['p.value'], method=padj_method)
    res_df['padj'] = res_df[annot_col]

    if factor_volcanos:
        de_df = res_df.unstack('contrast')
        de_df.columns = ['{}.{}'.format(x, y) for x, y in
                         zip(de_df.columns.get_level_values(0), de_df.columns.get_level_values(1))]
        if Y_name == 'CYTO':
            if color_by_stimulus:
                de_df['Stimulus'] = rename_cytokines(de_df.index, only_stimulus=True, with_time=False)
                de_df['Duration'] = rename_cytokines(de_df.index, only_time=True)
            de_df['y'] = rename_cytokines(de_df.index, only_cyto=True)

            style = 'Duration' if marker_by_duration else None
            style_order = list(DURATION_COLORS.keys()) if marker_by_duration else None
            hue = 'Stimulus' if color_by_stimulus else None
            hue_order = list(STIMULUS_COLORS.keys()) if color_by_stimulus else None
            if 'title' not in legend_kws:
                legend_kws['title'] = 'Stimulus' if color_by_stimulus else Y_name
        else:
            if Y_name == 'CM':
                de_df['y'] = rename_CMs(de_df.index)
            elif Y_name == 'WB_PER_ML':
                de_df['y'] = rename_bloods(de_df.index)
            else:
                de_df['y'] = de_df.index.str.replace('^{}:'.format(Y_name), '')

            style = None
            style_order = None
            hue = None
            hue_order = None
            if 'title' not in legend_kws:
                legend_kws['title'] = rename_Y(Y_name)

        _contrasts = res_df.index.get_level_values('contrast').drop_duplicates().values
        axs = volcano(de_df, contrasts=_contrasts[np.argsort(rename_factors(_contrasts, **factor_rename_kws))],
                      hue=hue, hue_order=hue_order,
                      style=style, style_order=style_order,
                      fdr=fdr, scatter_kwargs=scatter_kws, legend_kwargs=legend_kws, **volcano_kws)

        for ax in axs:
            contrast = ax.get_title()
            ax.set_title('{}: {}'.format(rename_Y(Y_name), rename_factors(contrast.replace('DONOR_', '').replace('SAMPLE_', '').replace('[T.M]', '').replace('VISIT[T.', '').replace(']', ''), **factor_rename_kws)))
            _de_df = de_df.sort_values(pval_col(contrast))
            if max_annot == 'all':
                targets = _de_df.index
            else:
                targets = _de_df.index[_de_df[padj_col(contrast)] <= np.min(fdr)][:max_annot]
            for target in targets:
                ax.annotate(de_df.loc[target, 'y'],
                            (de_df.loc[target, 'Coef.{}'.format(contrast)], -np.log10(de_df.loc[target, 'p.value.{}'.format(contrast)])),
                            fontsize=annot_size)
        if save_fig:
            savefig(fig_fn.format(data='volcanos', ext=fig_format if isinstance(fig_format, str) else fig_format['volcanos']), dpi=dpi)
        if show_fig:
            plt.show()
        plt.close()
    else:
        de_df = None

    _original_index = res_df.index
    _renamed_index = _rename_targets_and_factors(res_df.index, Y_name=Y_name, only_cyto=names_without_stimulus, **factor_rename_kws)
    res_df.index = _renamed_index

    if joint_volcano:
        joint_df = res_df.copy()
        joint_df.columns = joint_df.columns + '.joint'
        joint_df['contrast'] = joint_df.index.get_level_values('contrast')
        joint_df['target'] = joint_df.index.get_level_values('target')
        if age_effect_per_N_years:
            joint_df.loc[joint_df.index.get_level_values('contrast').str.lower().str.startswith('age'), 'Coef.joint'] *= age_effect_per_N_years

        if 'hue_order' not in volcano_kws:
            volcano_kws['hue_order'] = np.unique(joint_df['contrast'])
        axs = volcano(joint_df, contrasts='joint', hue='contrast', fdr=fdr,
                      scatter_kwargs=scatter_kws, legend_kwargs=legend_kws, **volcano_kws)
        for ax in axs:
            ax.set_title(rename_Y(Y_name))
            _joint_df = joint_df.sort_values(pval_col('joint'))
            if max_annot == 'all':
                targets = _joint_df.index
            else:
                targets = _joint_df.index[_joint_df[padj_col('joint')] <= np.min(fdr)][:max_annot]
            for trgt in targets:
                ax.annotate(joint_df.loc[trgt, 'target'],
                            (joint_df.loc[trgt, coef_col('joint')], -np.log10(joint_df.loc[trgt, pval_col('joint')])),
                            fontsize=annot_size)
        if save_fig:
            savefig(fig_fn.format(data='joint_volcano', ext=fig_format if isinstance(fig_format, str) else fig_format['joint_volcano']), dpi=dpi)
        if show_fig:
            plt.show()
        plt.close()
    else:
        joint_df = None

    if heatmap:
        if 'fmt' not in heatmap_kws:
            heatmap_kws['fmt'] = ''
        if 'lw' not in heatmap_kws:
            heatmap_kws['lw'] = 1
        if 'center' not in heatmap_kws:
            heatmap_kws['center'] = 0
        if 'cmap' not in heatmap_kws:
            heatmap_kws['cmap'] = 'RdBu_r'
        if 'xticklabels' not in heatmap_kws:
            heatmap_kws['xticklabels'] = 1
        if 'yticklabels' not in heatmap_kws:
            heatmap_kws['yticklabels'] = 1
        if 'square' not in heatmap_kws:
            heatmap_kws['square'] = True
        if cbar_label is not None:
            cbar_kws['label'] = cbar_label
        if 'size' not in annot_kws:
            annot_kws['size'] = annot_size

        pvals_df = signed_log10_pvals(res_df).unstack().T

        if force_order:
            try:
                pvals_df = pvals_df.loc[force_order]
            except KeyError as e:
                print('This is the index:', pvals_df.index.values)
                print('This is force_order:', force_order)
                print('This is the diff:', set(force_order).symmetric_difference(pvals_df.index))
                raise e

        if Y_name == 'CYTO':
            pvals_df = pvals_df.iloc[:, argsort_cytokines(pvals_df.columns, short_form=True)]

        fig, ax = plt.subplots(1, 1, figsize=((n_targets if not transpose else n_contrasts) * cell_width, (n_contrasts if not transpose else n_targets) * cell_height), squeeze=True, sharey=True)
        annot = heatmap_pval_annot(res_df[padj_col()], rows=pvals_df.index, cols=pvals_df.columns, thr=np.min(fdr), use_asterisk=pval_asterisks, fmt=pval_fmt, asterisks=pval_asterisks)
        if names_without_time:
            pvals_df.columns = pvals_df.columns.str.replace(', 24h|, 7d', '')
        if blank_no_significance:
            pvals_df = pvals_df.mask(pvals_df.abs() < -np.log10(np.min(fdr)), other=0)

        if transpose:
            pvals_df = pvals_df.T
            annot = annot.T
        ax = sns.heatmap(pvals_df, annot=annot, ax=ax,
                         cbar_kws=cbar_kws, annot_kws=annot_kws, **heatmap_kws)
        ax.set_xlabel(heatmap_xlabel)
        ax.set_ylabel(heatmap_ylabel)
        ax.set_title('{}{}{}'.format(
            heatmap_title if heatmap_title else '',
            '\n' if heatmap_title is not None and explain_heatmap_annot else '',
            'Comparisons with "{}" {} {} are annotated{}'.format(
                annot_col, LESS_EQUAL, np.min(fdr), ', nominal P-values > {} are in light gray'.format(np.min(fdr)) if blank_no_significance else '') if explain_heatmap_annot else ''
        ))
        if save_fig or show_fig:
            if save_fig:
                savefig(fig_fn.format(data='pvals_heatmap', ext=fig_format if isinstance(fig_format, str) else fig_format['pvals_heatmap']), dpi=dpi)
            if show_fig:
                plt.show()
            plt.close()
        else:
            _return.append(ax)
    else:
        pvals_df = None

    assert res_df.index.equals(_renamed_index)
    res_df.index = _original_index
    _return.extend([res_df, pvals_df, joint_df, de_df])
    return _return


def read_enrichr_or_lola(celltype, model, coef, effect_size_filter,
                         rank_metric, top_n, region_filter, direction, method, library=None,
                         include_fn='auto', leukotypes_regex='_treat_FC2_FDR1.bed$', rename_leukotypes=True,
                         swap_to_non_resp=False, results_dir='results', filename=None):
    if method == 'lola':
        if library is not None and not isinstance(library, str) and len(library) == 1:
            library = library[0]
        if filename is None:
            # because JASPAR does not fit into one LOLA run
            fn_type, ext = ('top_regions', None) if library == JASPAR else ('results', 'auto')
            fn = lola_fn(celltype, model, data=fn_type, coef=coef, effect_size_filter=effect_size_filter,
                         rank_metric=rank_metric, top_n=top_n, direction=direction, regions=region_filter,
                         ext=ext, swap_to_non_resp=swap_to_non_resp, results_dir=results_dir)
        else:
            fn = filename

        if os.path.exists(fn):
            if library == JASPAR:
                df = []
                for i in range(1, 21):
                    df.append(pd.read_csv(os.path.join(fn, f'JASPAR_2022_part{i}', 'allEnrichments.tsv'), index_col='dbSet', sep='\t'))
                    assert len(set(df[-1]['userSet'])) == 1 and df[-1].index.is_unique
                df = pd.concat(df).reset_index(drop=True)
                df.index.name = 'dbSet'
            else:
                df = pd.read_csv(fn, index_col='dbSet', sep='\t')
                df = df.loc[~df.isnull().all(axis=1)]
                assert len(set(df['userSet'])) == 1, set(df['userSet'])
                assert df.index.is_unique, df.index[df.index.duplicated()]
            assert not df['pValueLog'].isnull().any()
            if leukotypes_regex and 'leukotypes' in df['collection'].values:
                df = df.loc[(df['collection'] != 'leukotypes') | df['filename'].str.contains(leukotypes_regex)]
            df['pValueLog'] = np.power(10, -1 * df['pValueLog'])
            df['Overlap'] = df['support'].astype(int).astype(str) + '/' + df[['support', 'b']].sum(axis=1).astype(int).astype(str)
            df['Term'] = df['description']
            if include_fn == True or (include_fn == 'auto' and library == JASPAR):
                df['Term'] = df['Term'] + ' (' + df['filename'].str.replace('.bed', '').str.replace('.narrowPeak', '') + ')'
            if 'leukotypes' in df['collection'].values:
                df.loc[df['collection'] == 'leukotypes', 'Term'] = df.loc[df['collection'] == 'leukotypes', 'cellType'].str.replace(leukotypes_regex.replace('.bed', '', 1) if rename_leukotypes else '\.bed$', '').str.replace('_FDR5$', '*').str.replace('_FDR1$', '**')
            df = df.rename({'collection': 'Gene_set', 'pValueLog': 'P-value', 'oddsRatio': 'Odds Ratio'}, axis=1)
            df = df[['Term', 'Gene_set', 'P-value', 'Odds Ratio', 'Overlap']].set_index('Term')

        else:
            print('Missing:', fn)
            df = None

    elif method == 'enrichr':
        if filename is None:
            fn = enrichr_fn(celltype, model, data='results', coef=coef, effect_size_filter=effect_size_filter,
                            rank_metric=rank_metric, top_n=top_n, direction=direction, db=None,
                            regions=region_filter, swap_to_non_resp=swap_to_non_resp, results_dir=results_dir)
        else:
            fn = filename
        if os.path.exists(fn):
            df = pd.read_csv(fn, index_col=1, sep='\t')
        else:
            print('Missing:', fn)
            df = None
    else:
        raise ValueError

    if library is not None:
        sns.pointplot()
        if isinstance(library, str):
            library = [library]
        df = df.loc[df['Gene_set'].isin(library)]

    return df
        
        
def enr_heatmap(celltype, model, contrasts, library, top_n, region_filter, method, pthw_fdr_or_top_n,
                rank_metric='p.value', effect_size_filter=None, directions=[1, -1], transpose=False,
                annot_fdr_thr=0.05, annot=None, padj_method='fdr_bh', global_fdr=False, force_bonferroni=None, pval_asterisks=None, sort_by_n_column=None, sort_ascending=False, force_index=None,
                fmt='.0e', cbar_label='Signed {}log10{}P-value'.format(MINUS, NBSPC), annot_size=10, rename_pthw=False, return_just_enrichments=False,
                leukotypes_regex='_treat_FC2_FDR1.bed$',
                rename_leukotypes=True, swap_to_non_resp=False,
                fig_width=2, cell_height=0.3, save_fig=False, show_fig=True, fig_format='svg',
                enr_df=None, terms=None, display_metric='pval', annot_metric='padj',
                heatmap_kws={}, cbar_kws={}, annot_kws={}, results_dir='results'):
    assert not (terms is None and enr_df is not None)
    if isinstance(library, str):
        library = [library]

    if enr_df is None:
        _terms = {-1: set([]), 1: set([])}
        enr_df, hits_df = {}, {}
        for direction in directions:
            enr_df[direction], hits_df[direction] = {}, {}
            for coef in contrasts:
                df = read_enrichr_or_lola(celltype, model, coef=coef, library=library, effect_size_filter=effect_size_filter,
                                          rank_metric=rank_metric, top_n=top_n, region_filter=region_filter,
                                          method=method, direction=direction, leukotypes_regex=leukotypes_regex,
                                          rename_leukotypes=rename_leukotypes, swap_to_non_resp=swap_to_non_resp,
                                          results_dir=results_dir)
                if df is not None:
                    df.loc[df['P-value'] == 0, 'P-value'] = sys.float_info.min
                    if len(library) == 1 and 'Adjusted P-value' in df and not global_fdr and not force_bonferroni:
                        pd.testing.assert_series_equal(
                            df['Adjusted P-value'], adjusted_pvals(df['P-value'], method=padj_method), check_names=False)
                    else:
                        if force_bonferroni:
                            df['Adjusted P-value'] = df['P-value'] * force_bonferroni
                            df['Adjusted P-value'] = df['Adjusted P-value'].where(df['Adjusted P-value'] < 1, other=1)
                        elif global_fdr:
                            _pvals = []
                            for _coef in contrasts:
                                if _coef != coef:
                                    _df = read_enrichr_or_lola(
                                        celltype, model, coef=_coef, library=library, effect_size_filter=effect_size_filter,
                                        rank_metric=rank_metric, top_n=top_n, region_filter=region_filter, method=method,
                                        direction=direction, results_dir=results_dir)
                                    _df.loc[_df['P-value'] == 0, 'P-value'] = sys.float_info.min
                                    _pvals.extend(_df['P-value'].tolist())
                            df['Adjusted P-value'] = adjusted_pvals(pd.concat([df['P-value'], pd.Series(_pvals)]), method=padj_method).loc[df['P-value'].index]
                            del _pvals, _df
                        else:
                            df['Adjusted P-value'] = adjusted_pvals(df['P-value'], method=padj_method)

                    if isinstance(pthw_fdr_or_top_n, dict):
                        if isinstance(pthw_fdr_or_top_n[coef], tuple):
                            _fdr, _top_n_down, _top_n_up = pthw_fdr_or_top_n[coef]
                            top_hits = df.loc[df['Adjusted P-value'] < _fdr]['P-value'].sort_values().head(_top_n_down if direction < 0 else _top_n_up).index
                        elif pthw_fdr_or_top_n[coef] < 1:
                            top_hits = df.loc[df['Adjusted P-value'] < pthw_fdr_or_top_n[coef]].index
                        else:
                            top_hits = df['P-value'].sort_values().head(pthw_fdr_or_top_n[coef]).index
                    elif isinstance(pthw_fdr_or_top_n, tuple):
                        if len(pthw_fdr_or_top_n) == 2:
                            _fdr, _top_n = pthw_fdr_or_top_n
                            top_hits = df.loc[df['Adjusted P-value'] < _fdr]['P-value'].sort_values().head(_top_n).index
                        else:
                            _fdr, _top_n_down, _top_n_up, _coef = pthw_fdr_or_top_n
                            if coef == _coef:
                                top_hits = df['P-value'].sort_values().head(_top_n_down if direction < 0 else _top_n_up).index
                            else:
                                top_hits = df.loc[df['Adjusted P-value'] < _fdr].index
                    elif pthw_fdr_or_top_n < 1:
                        top_hits = df.loc[df['Adjusted P-value'] < pthw_fdr_or_top_n].index
                    else:
                        top_hits = df['P-value'].sort_values().head(pthw_fdr_or_top_n).index
                    _terms[direction] = _terms[direction].union(top_hits)
                    enr_df[direction][coef] = df
                    hits_df[direction][coef] = df.loc[top_hits]
        if return_just_enrichments:
            return enr_df

        # if len(directions) > 1:
        #     for coef in contrasts:
        #         assert len(hits_df[1][coef].index.intersection(hits_df[-1][coef].index)) == 0

    if terms is None:
        if len(_terms[1].intersection(_terms[-1])) != 0:
            print('Some terms were enriched both positively and negatively!!!')
            print('Only the stronger direction will be displayed:', _terms[1].intersection(_terms[-1]))
        terms = _terms[1].union(_terms[-1])

    results = {}
    if np.sum([len(enr_df[d]) for d in directions]) > 0:
        _max_abs = lambda x: max(x.min(), x.max(), key=abs)
        for metric in ['pval', 'padj', 'oddsratio']:
            df = pd.DataFrame(index=terms)
            for coef in contrasts:
                _df = pd.DataFrame(index=terms)
                for direction in directions:
                    if len(enr_df[direction]) > 0:
                        x = enr_df[direction][coef].loc[terms.intersection(enr_df[direction][coef].index)]
                        if metric == 'pval':
                            _df[direction] = direction * (-np.log10(x['P-value']))
                        elif metric == 'padj':
                            _df[direction] = x['Adjusted P-value']
                        elif metric == 'oddsratio':
                            _df[direction] = direction * x['Odds Ratio']
                        else:
                            raise ValueError
                if metric == 'pval':
                    _df = _df.fillna(0).apply(_max_abs, axis=1).rename(coef)
                elif metric == 'padj':
                    _df = _df.fillna(1).apply(min, axis=1).rename(coef)
                elif metric == 'oddsratio':
                    _df = _df.fillna(0).apply(_max_abs, axis=1).rename(coef)
                df = pd.concat([df, _df], axis=1)
            df.index = df.index.str.replace(' \(GO:[0-9]*\)', '')
            df.index = df.index.str.replace(' \(MA[0-9\.]*\)', '')
            results[metric] = df

        if force_index is None and sort_by_n_column is None:
            results[display_metric] = results[display_metric].sort_index()
        else:
            if force_index is not None:
                if isinstance(force_index, str) and force_index == 'fix_max':
                    results[display_metric] = results[display_metric].loc[results[display_metric].idxmax(axis=1).sort_values().index]
                elif isinstance(force_index, set):
                    results[display_metric] = results[display_metric].loc[results[display_metric].index.isin(force_index)]
                else:
                    results[display_metric] = results[display_metric].reindex(force_index)
            if sort_by_n_column is not None:
                if isinstance(sort_by_n_column, int):
                    results[display_metric] = results[display_metric].sort_values(results[display_metric].columns[sort_by_n_column], ascending=sort_ascending)
                elif sort_by_n_column == 'special':
                    sort1 = results[display_metric][list(contrasts)[0]]
                    sort1 = -1 * sort1[(sort1 < 0) if swap_to_non_resp else (sort1 > 0)]
                    sort2 = results[display_metric][list(contrasts)[-1]]
                    sort2 = sort2[sort2 > 0]
                    sort = pd.concat([sort2.sort_values(ascending=False), sort1.sort_values(ascending=False)]).index
                    sort = sort[~sort.duplicated()]
                    assert sort.is_unique
                    if set(sort) != set(results[display_metric].index):
                        print('WARNING: loosing this by sorting:')
                        print(set(sort).symmetric_difference(results[display_metric].index))
                    results[display_metric] = results[display_metric].loc[sort]
                else:
                    raise ValueError

        # results[display_metric].index = results[display_metric].index.str.replace(' \(GO:[0-9]*\)', '')
        results['pval'] = results['pval'].loc[results[display_metric].index]
        results['padj'] = results['padj'].loc[results[display_metric].index]
        results['oddsratio'] = results['oddsratio'].loc[results[display_metric].index]

        _kws = dict(use_asterisk=True if pval_asterisks else False, fmt=fmt)
        if is_iterable(pval_asterisks):
            _kws['asterisks'] = pval_asterisks

        _annot = heatmap_pval_annot(results[annot_metric], thr=annot_fdr_thr, **_kws)
        if annot is None:
            annot = _annot
        else:
            assert annot.shape == _annot.shape, (annot.shape, _annot.shape)
            annot.index = _annot.index
            annot.columns = _annot.columns

        if 'square' not in heatmap_kws:
            heatmap_kws['square'] = True
        if 'fmt' not in heatmap_kws:
            heatmap_kws['fmt'] = ''
        if 'lw' not in heatmap_kws:
            heatmap_kws['lw'] = 1
        if 'center' not in heatmap_kws:
            heatmap_kws['center'] = 0
        if 'cmap' not in heatmap_kws:
            heatmap_kws['cmap'] = 'RdBu_r'
        if 'xticklabels' not in heatmap_kws:
            heatmap_kws['xticklabels'] = 1
        if 'yticklabels' not in heatmap_kws:
            heatmap_kws['yticklabels'] = 1
        if cbar_label is not None:
            cbar_kws['label'] = cbar_label
        if 'size' not in annot_kws:
            annot_kws['size'] = annot_size
        # print(annot_kws['size'])

        figsize = [fig_width, results['padj'].shape[0] * cell_height]
        if transpose:
            results['pval'] = results['pval'].T
            results['padj'] = results['padj'].T
            results['oddsratio'] = results['oddsratio'].T
            annot = annot.T
            figsize = figsize[::-1]

        fig, ax = plt.subplots(1, 1, figsize=figsize)

        if rename_pthw:
            for m in results:
                results[m].columns = rename_pathways(results[m].columns)
            annot.columns = rename_pathways(annot.columns)

        sns.heatmap(results[display_metric], annot=annot, cbar_kws=cbar_kws, annot_kws=annot_kws, **heatmap_kws, ax=ax)

        ax.set_title('{}\nTop {}, {}'.format(' + '.join([l[:-len('_min15_max500')] if l.endswith('_min15_max500') else l for l in library]), top_n, region_filter))

        if save_fig or show_fig:
            if save_fig:
                if method == 'enrichr':
                    fn = enrichr_fn(celltype, model, data='{data}', coef=None, regions=region_filter,
                                    effect_size_filter=effect_size_filter, rank_metric=rank_metric, top_n=top_n,
                                    direction=None, db=None, subdir='.', ext='{ext}')
                elif method == 'lola':
                    fn = lola_fn(celltype, model, data='{data}', coef=None,
                                 effect_size_filter=effect_size_filter, rank_metric=rank_metric, top_n=top_n,
                                 direction=None, db=None, subdir='.', ext='{ext}')
                savefig(fn.format(data='enrplot.{}'.format('.'.join(library)), ext=fig_format))
            if show_fig:
                plt.show()
            plt.close()
    return ax, results, enr_df


def _mean_ML(df, score_names=['mean_train_score', 'mean_cv_score', 'mean_test_score', 'union_test_score'], extra_cols=['classifier', 'scoring', 'binarize', 'y']):
    assert len(df) != 0
    assert df.reset_index().set_index(['model', 'seed']).index.is_unique

    std_df = df.groupby(['model'] + extra_cols)[score_names].std().reset_index().set_index('model')
    mean_df = df.groupby(['model'] + extra_cols)[score_names].mean().reset_index().set_index('model')
    assert std_df.index.is_unique and mean_df.index.is_unique

    # fix negative zeros
    for s in score_names:
        std_df.loc[std_df[s] == 0, s] = 0
        mean_df.loc[mean_df[s] == 0, s] = 0

    return mean_df, std_df


def read_ML(key, models=None, X_names=None, ys=None, X_visits=None, y_visits=None, binarizing=None,
             data=None, scale=None,
             results_fn=ML_RESULTS_FN, verbose=True,
             show_visits=True, show_data=True, show_estimator=True,
             show_X_name=True, show_y_name=True, show_binarizing=True, rename_y=None):

    if isinstance(models, str):
        models = [models]
    if isinstance(X_names, str):
        X_names = [X_names]
    if isinstance(ys, str):
        ys = [ys]
    if isinstance(X_visits, str):
        X_visits = [X_visits]
    if isinstance(y_visits, str):
        y_visits = [y_visits]
    if isinstance(binarizing, str):
        binarizing = [binarizing]
    if isinstance(data, str):
        data = [data]

    df = pd.read_hdf(results_fn, key=key)
    df = df.rename(dict(model='data', global_PCA='gPCA', global_KPCA='gKPCA', global_UMAP='gUMAP'), axis=1)
    df['classifier'] = ~df['class_ratio'].isnull()

    if scale is not None:
        if isinstance(scale, dict):
            for _X_name in scale:
                df = df.loc[(df['X_name'] != _X_name) | (df['scale'] == scale[_X_name])]
        else:
            df = df.loc[df['scale'] == scale]
    if X_names:
        df = df.loc[df['X_name'].isin(X_names)]
    if ys:
        df = df.loc[df['y'] .isin(ys)]
    if X_visits:
        df = df.loc[df['X_visits'].isin(X_visits)]
    if y_visits:
        df = df.loc[df['y_visits'].isin(y_visits)]
    if binarizing:
        df = df.loc[df['binarize'].isin(binarizing)]
    if data:
        df = df.loc[df['data'].isin(data)]
    if rename_y:
        df['y'] = df['y'].replace(rename_y)

    if len(df) != 0:
        _extra_cols = ['classifier', 'scoring', 'binarize', 'y']
        if not show_visits:
            assert np.unique(df['X_visits']).shape[0] == 1
            assert np.unique(df['y_visits']).shape[0] == 1
            df = df.drop(['X_visits', 'y_visits'], axis=1)
        if not show_data:
            assert np.unique(df['data']).shape[0] == 1
            df = df.drop('data', axis=1)
        if not show_estimator:
            assert np.unique(df['estimator']).shape[0] == 1
            df = df.drop('estimator', axis=1)
        if not show_X_name:
            assert np.unique(df['X_name']).shape[0] == 1
            df = df.drop('X_name', axis=1)
        if not show_y_name:
            assert np.unique(df['y']).shape[0] == 1
            df = df.drop('y', axis=1)
            _extra_cols.remove('y')
        if not show_binarizing:
            assert np.unique(df['binarize']).shape[0] == 1
            df = df.drop('binarize', axis=1)
            _extra_cols.remove('binarize')

        def _rename(x):
            model = '_'.join(x.loc[[col for col in ['estimator', 'data', 'X_visits', 'X_name', 'y_visits', 'y', 'binarize'] if col in x.index]].astype(str))
            for col in ['poly', 'gKPCA', 'gPCA', 'gUMAP', 'features', 'DE_filter', 'select', 'pca']:
                if col in x.index and x[col] and x[col] and not np.isnan(x[col]):
                    model += '_{}:{}'.format(col, x[col])
            return model

        df['model'] = df.apply(_rename, axis=1)
        if models:
            df = df.loc[df['model'].isin(models)]
        df = df.set_index(['model', 'seed'])
        assert df.index.is_unique
        mean_df, std_df = _mean_ML(df, extra_cols=_extra_cols)
    else:
        mean_df, std_df = None, None

    return df, mean_df, std_df


def get_preds_for_model(results_df, model, n_seeds=None, return_grids=False):
    df = results_df.reset_index()
    df = df.loc[df['model'] == model]
    assert n_seeds is None or len(df) == n_seeds, '{} != {}'.format(len(df), n_seeds)

    for idx in df.index:
        model_replicate = df.loc[idx]
        _return = [model_replicate, np.load(model_replicate['preds_fn'], allow_pickle=True)]
        if return_grids:
            _return.append(joblib.load(model_replicate['grids_fn'], allow_pickle=True))
        yield _return


def get_preds_for_X(results_df, X_name, n_seeds=None, return_grids=False, return_roc=False, return_prc=False):
    df = results_df.loc[results_df['X_name'] == X_name]
    for col in ['features', 'pca', 'downsampling', 'binarize', 'classifier', 'data', 'X_visits', 'y_visits',
                'estimator', 'select', 'DE_filter', 'scale']:
        assert df[col].isnull().all() or len(set(df[col])) == 1, col
    assert n_seeds is None or df.shape[0] == n_seeds, df.shape[0]
    assert n_seeds is None or sorted(df['seed'].tolist()) == [100 * (i + 1) for i in range(n_seeds)]

    for idx in df.index:
        preds = np.load(df.loc[idx, 'preds_fn'], allow_pickle=True)
        _return = preds if not return_roc and not return_prc and not return_grids else [preds]
        if return_grids:
            grids = joblib.load(df.loc[idx, 'grids_fn'])
            _return.append(grids)
        if return_roc:
            roc = df.loc[idx, 'FPR'], df.loc[idx, 'TPR']
            _return.append(roc)
        if return_prc:
            prc = df.loc[idx, 'precision'], df.loc[idx, 'recall']
            _return.append(prc)
        yield _return


def calc_delongs(reference, all_y_trues, all_y_preds):
    for model in all_y_trues:
        if model != reference:
            for seed in all_y_trues[model]:
                yt, yp = all_y_trues[model][seed], all_y_preds[model][seed]
                delong_p = np.power(10, delong_roc_test(yt, yp, all_y_preds[reference][seed])[0, 0])
                auroc1 = roc_auc_score(all_y_trues[reference][seed], all_y_preds[reference][seed])
                auroc2 = roc_auc_score(yt, yp)
                print('{} vs. {}: {:.2f} vs. {:.2f} [{:.2f}], p = {:.3f}{}'.format(reference, model, auroc1, auroc2, auroc1 - auroc2, delong_p, ' [**]' if delong_p < 0.01 else ' [*]' if delong_p< 0.05 else ' [~]' if delong_p < 0.1 else ''))


def plot_ML_curves(results_df, models, plot_type, merge_folds=True, n_seeds=1, colors=None, linestyles=None, figsize=None,
                    title=None, show_random=True, all_seeds_same_data=True, random_label='random', random_color=None, random_linestyle=None,
                    show_all_curves=True, show_mean=False, show_std=False, show_iqr=False, show_ci=False,
                   smoothness=11, alpha_all_curves=0.25, alpha_std=0.5, ax=None, **kwargs):
    if isinstance(models, str):
        models = [models]
    elif models is None:
        models = np.unique(results_df.reset_index()['model'])
    if isinstance(show_all_curves, bool):
        show_all_curves = [show_all_curves]
    if isinstance(colors, str) or colors is None:
        colors = [colors]
    if isinstance(linestyles, str) or linestyles is None:
        linestyles = [linestyles]
    assert np.sum([show_std, show_iqr, show_ci]) < 2
    assert plot_type in ['ROC', 'PRC']

    if ax is None:
        fig, ax = plt.subplots(1, 1, figsize=figsize)

    if show_random:
        plt.plot([0, 1], [0, 1] if plot_type == 'ROC' else [0.5, 0.5],
                 label=random_label, c=random_color, ls=random_linestyle, **kwargs)

    all_y_preds, all_y_trues = {}, {}
    for model, show_raw_curves, color, ls in zip(models,
                                np.tile(show_all_curves, int(np.ceil(len(models) / len(show_all_curves)))),
                                np.tile(colors, int(np.ceil(len(models) / len(colors)))),
                                np.tile(linestyles, int(np.ceil(len(models) / len(linestyles))))):
        base_x_axis = np.linspace(0, 1, smoothness)
        previous_samples, previous_y_true = None, None
        all_y_preds[model], all_y_trues[model] = {}, {}
        all_y_axis = []
        for result, preds in get_preds_for_model(results_df, model=model, n_seeds=n_seeds):

            seed = result['seed']
            # different seeds may differ in the order
            sort = np.argsort(np.concatenate(preds['samples']))
            samples = np.concatenate(preds['samples'])[sort]
            assert len(samples) == len(set(samples))
            y_true = np.concatenate(preds['y_true'])[sort]
            y_pred = np.concatenate(preds['y_pred'])[sort]

            if all_seeds_same_data:
                if previous_samples is None:
                    previous_samples = samples
                    previous_y_true = y_true
                else:
                    assert np.array_equal(samples, previous_samples)
                    assert np.array_equal(y_true, previous_y_true) or 'PERMUTED_' in model

            _first = True
            for yt, yp in [(y_true, y_pred)] if merge_folds else zip(preds['y_true'], preds['y_pred']):
                if plot_type == 'ROC':
                    x_axis, y_axis, _ = roc_curve(yt, yp)
                else:
                    y_axis, x_axis, _ = precision_recall_curve(yt, yp)

                if show_raw_curves:
                    plt.plot(x_axis, y_axis, label=model if not show_mean and _first else None, c=color, ls=ls, alpha=alpha_all_curves, **kwargs)

                y_axis = np.interp(base_x_axis, x_axis, y_axis)
                y_axis[0] = 0.0
                all_y_axis.append(y_axis)
                _first = False
                
            all_y_preds[model][seed] = y_pred
            all_y_trues[model][seed] = y_true

        all_y_axis = np.array(all_y_axis)
        mean_y_axis = all_y_axis.mean(axis=0)

        if show_std:
            std_y_axis = all_y_axis.std(axis=0)
            y_axis_upper = np.minimum(mean_y_axis + std_y_axis, 1)
            y_axis_lower = mean_y_axis - std_y_axis
            plt.fill_between(base_x_axis, y_axis_lower, y_axis_upper, color=color, alpha=alpha_std, linewidth=0)

        if show_iqr:
            y_axis_upper = np.percentile(all_y_axis, 75, axis=0)
            y_axis_lower = np.percentile(all_y_axis, 25, axis=0)
            plt.fill_between(base_x_axis, y_axis_lower, y_axis_upper, color=color, alpha=alpha_std, linewidth=0)

        if show_ci:
            bs_mean_y_axis = []
            for _ in range(1000):
                bs_all_y_axis = all_y_axis[
                    np.random.choice(np.arange(len(all_y_axis)), size=len(all_y_axis), replace=True)]
                bs_mean_y_axis.append(bs_all_y_axis.mean(axis=0))
            bs_mean_y_axis = np.array(bs_mean_y_axis)
            y_axis_upper = np.percentile(bs_mean_y_axis, 99, axis=0)
            y_axis_lower = np.percentile(bs_mean_y_axis, 1, axis=0)
            plt.fill_between(base_x_axis, y_axis_lower, y_axis_upper, color=color, alpha=alpha_std, linewidth=0)

        if show_mean:
            plt.plot(base_x_axis, mean_y_axis, label=model, c=color, ls=ls, **kwargs)

    ax.set_xlim([-0.05, 1.05])
    ax.set_ylim([-0.05, 1.05])
    ax.set_xticks(np.arange(0, 1.01, 0.2))
    ax.set_yticks(np.arange(0, 1.01, 0.2))
    ax.set_xlabel('False positive rate' if plot_type == 'ROC' else 'Recall')
    ax.set_ylabel('True positive rate' if plot_type == 'ROC' else 'Precision')
    ax.set_title(title)
    _handles, _labels = ax.get_legend_handles_labels()
    ax.legend(_handles[::-1], _labels[::-1], bbox_to_anchor=(1, 1))
    sns.despine()

    return ax, all_y_preds, all_y_trues


def plot_ML_summary(mean_df, score_names=dict(mean_train_score='Training set', mean_cv_score='Grid search CV', union_test_score='Blind test'),
                    fmt='.2f', vmin=0.5, vmax=0.8, fontsize=10, scoring_names=dict(roc_auc='AUROC', average_precision='AUPRC', r2='$R^2$'),
                    cell_width=0.5, cell_height=0.3, hspace=None, wspace=None):

    assert len(set(mean_df['scoring'])) == 1
    scoring = mean_df['scoring'].iloc[0]
    _is_classifier = mean_df['classifier'].drop_duplicates().values
    assert len(_is_classifier) == 1
    _is_classifier = _is_classifier[0]
    _n = mean_df.shape[0]
    _df = mean_df.copy()

    if 'y' in _df.columns and 'binarize' in _df.columns:
        for y, b in _df[['y', 'binarize']].drop_duplicates().values:
            _df.index = _df.index.str.replace(r'_{y}_{b}|{y}_{b}_'.format(y=y, b=b), '')
        _row_cols = ['y', 'binarize']
    elif 'y' in _df.columns:
        for y in _df[['y']].drop_duplicates().values:
            _df.index = _df.index.str.replace(r'_{y}|{y}_'.format(y=y), '')
        _row_cols = ['y']
    elif 'binarize' in _df.columns:
        for b in _df[['binarize']].drop_duplicates().values:
            _df.index = _df.index.str.replace(r'_{b}|{b}_'.format(b=b), '')
        _row_cols = ['binarize']

    fig, axs = plt.subplots(nrows=len(score_names), ncols=1,
                            figsize=(len(_df.index.drop_duplicates()) * cell_width,
                                     len(score_names) * len(_df[_row_cols].drop_duplicates()) * cell_height),
                            sharex=True, sharey=True, constrained_layout=False, squeeze=False)
    fig.subplots_adjust(hspace=hspace, wspace=wspace)

    for row, score_name in enumerate(score_names):
        df = pd.pivot_table(_df, values=score_name, index=_row_cols, columns=_df.index)
        assert _n == (~df.isnull()).sum().sum()

        # df = df.sort_index()[[_x for _x in _X_names if _x in df.columns]]

        ax = sns.heatmap(df,
                         vmin=vmin, vmax=vmax,
                         xticklabels=1, linewidths=1,
                         annot=True, annot_kws={'size': fontsize}, fmt=fmt,
                         cbar=True,
                         ax=axs[row, 0],
                         )
        ax.tick_params(labelleft=True, rotation=0)
        ax.collections[0].colorbar.ax.tick_params(labelsize=fontsize)
        ax.collections[0].colorbar.ax.set_ylabel('{} {}'.format(score_names[score_name], scoring_names.get(scoring, scoring)), fontsize=fontsize)
        ax.set_ylim(df.shape[0], 0)
        ax.set_xlabel(None)
        ax.set_ylabel(None)
        ax.set_xticklabels(ax.get_xticklabels(), rotation=90)  # , ha='right')
        ax.set_title('{} ({})'.format(score_names[score_name], 'classification' if _is_classifier else 'regression'))
    return fig, axs


def FVE_boxplot(df, x, y_label='Variance explained (%)', palette=None, figsize=None, alpha=None,
                size=1.5, jitter=True, rasterized=False, random_state=RANDOM_STATE):
    if x == 'contrast':
        order = df.groupby('contrast').median().sort_values('FVE', ascending=False).index.tolist()
        if 'Residuals' in order:
            order.remove('Residuals')
            order.append('Residuals')
        if 'Within-donor' in order:
            order.remove('Within-donor')
            order.append('Within-donor')
    else:
        assert len(df['contrast'].unique()) == 1
    fig, ax = plt.subplots(1, 1, figsize=figsize)
    ax = sns.boxplot(data=df, x=x, y='FVE', fliersize=0, order=order if x == 'contrast' else None, palette=palette)
    ax = stripplot(data=df, x=x, y='FVE', s=size, alpha=alpha, palette=['k'], jitter=jitter, rasterized=rasterized,
                   order=order if x == 'contrast' else None, random_state=random_state)
    ax.set_ylim((-5, 105))
    ax.set_xlim((ax.get_xlim()[0] - 0.25, ax.get_xlim()[1] + 0.25))
    plt.setp(ax.artists, edgecolor='k', lw=0.5)
    plt.setp(ax.lines, color='k', lw=0.5)
    ax.set_xticklabels(ax.get_xticklabels(), rotation=90) #, ha='right')
    ax.set_ylabel(y_label)
    ax.set_xlabel(None)
    sns.despine()


def make_corr_long_table(corr_df, pval_df, symmetrical,
                         col_name=None, idx_name=None, idx_name_suffix=None, col_name_suffix=None):
    assert corr_df.index.equals(pval_df.index)
    assert corr_df.columns.equals(pval_df.columns)
    corr_df = corr_df.copy()
    pval_df = pval_df.copy()
    if symmetrical:
        assert corr_df.index.equals(corr_df.columns)
    else:
        assert corr_df.isnull().sum().sum() == 0
        assert pval_df.isnull().sum().sum() == 0

    if col_name is None:
        if corr_df.columns.str.startswith('CYTO:').all():
            col_name = 'CYTO'
        elif corr_df.columns.str.startswith('CM:').all():
            col_name = 'CM'
        elif corr_df.columns.str.startswith('WB_PER_ML:').all():
            col_name = 'WB_PER_ML'
    if idx_name is None:
        if corr_df.index.str.startswith('CYTO:').all():
            idx_name = 'CYTO'
        elif corr_df.index.str.startswith('CM:').all():
            idx_name = 'CM'
        elif corr_df.index.str.startswith('WB_PER_ML:').all():
            idx_name = 'WB_PER_ML'

    index = rename_Y_vars(Y_name=idx_name, var=corr_df.index, with_time=True)
    columns = rename_Y_vars(Y_name=col_name, var=corr_df.columns, with_time=True)
    corr_df.index = pval_df.index = index
    corr_df.columns = pval_df.columns = columns

    corr_df = corr_df.iloc[case_insensitive_sort_index(corr_df.index), case_insensitive_sort_index(corr_df.columns)]
    pval_df = pval_df.iloc[case_insensitive_sort_index(pval_df.index), case_insensitive_sort_index(pval_df.columns)]
    if symmetrical:
        corr_df = upper_triangle(corr_df)
        pval_df = upper_triangle(pval_df)

    corrs = corr_df.stack()
    pvals = pval_df.stack()
    assert corrs.index.equals(corrs.index)
    assert pvals.index.equals(pvals.index)

    df = pd.concat([corrs, pvals], axis=1)
    idx_name = LONG_NAMES[idx_name] + (idx_name_suffix if idx_name_suffix else '')
    col_name = LONG_NAMES[col_name] + (col_name_suffix if col_name_suffix else '')
    if idx_name == col_name:
        idx_name += ' 1'
        col_name += ' 2'
    df.index.names = [idx_name, col_name]
    df.columns = [R_COL, PVAL_COL]
    return df


def case_insensitive_sort_index(index):
    index_df = index.to_frame().reset_index(drop=True)
    return index_df.apply(lambda col: col.astype(str).str.lower()).sort_values(index_df.columns.tolist()).index
#
#
# def case_insensitive_sort_index(df):
#     # use numeric index
#     _names = df.index.names
#     assert len(_names) != 0
#     df = df.reset_index()
#     # get sorted index with lower-case columns
#     _sort = df[_names].apply(lambda col: col.str.lower()).sort_values(_names).index
#     # sort the dataframe and set the MultiIndex
#     df = df.loc[_sort].set_index(_names)
#     return df


def get_legends_sheet_name(table_name):
    name_split = table_name.split("_")
    if name_split[0] == 'Ext':
        return f'Ext_Table{name_split[1]}_legends'
    else:
        return f'Table{name_split[0]}_legends'


def get_container_with_tables(tables, modes=None, supp_dir=os.path.join('results', 'tables'), include_legends_sheet=True):
    if isinstance(tables, str):
        tables = [tables]
    if isinstance(modes, str):
        modes = [modes]
    if modes is None:
        modes = ['w'] * len(tables)
    assert len(tables) == len(modes)
    suppl_tables = {}
    for table, mode in zip(tables, modes):
        fn = os.path.join(supp_dir, f'{table}.xlsx')
        assert mode == 'w' or os.path.exists(fn), f'If you are appending, {fn} must exist'
        if mode == 'w' and include_legends_sheet:
            # create file with the legends sheet empty
            save_excel(pd.DataFrame(index=pd.Index([], name='Sheet name'), columns=['Contents']), fn, get_legends_sheet_name(table))

        suppl_tables[table] = {
            'info': OrderedDict(pd.read_excel(fn, sheet_name=get_legends_sheet_name(table), index_col=0)['Contents'].to_dict() if include_legends_sheet else {}),
            'file': pd.ExcelWriter(fn, engine='openpyxl', mode=mode),
            'filename': fn,
            'index_cols': {},
            'header_rows': {},
        }
        if os.path.exists(fn):
            workbook = openpyxl.load_workbook(fn)
            suppl_tables[table]['file'].book = workbook
            suppl_tables[table]['file'].sheets = dict((ws.title, ws) for ws in workbook.worksheets)
    return suppl_tables


def close_tables_and_save(suppl_tables):
    infos, tables = [], []
    for table in suppl_tables:
        sheet_names_too_long = [sheet_name for sheet_name in suppl_tables[table]['info'] if len(sheet_name) > 31]
        if len(sheet_names_too_long) != 0:
            print('These sheet names are too long:', sheet_names_too_long)
        try:
            info_df = pd.DataFrame.from_dict(suppl_tables[table]['info'], orient='index', columns=['Contents'])
            if len(info_df) != 0:
                info_df.index.name = 'Sheet name'
                suppl_tables[table]['header_rows'][get_legends_sheet_name(table)] = len(info_df.columns.names)
                suppl_tables[table]['index_cols'][get_legends_sheet_name(table)] = len(info_df.index.names)
                save_excel(info_df, suppl_tables[table]['file'], get_legends_sheet_name(table))
                infos.append(info_df)
                tables.append(table)
            suppl_tables[table]['file'].close()
        except IndexError:
            pass
    return infos, tables


def save_excel(df, file, sheet_name, freeze_panes='auto', **kwargs):
    if freeze_panes == 'auto':
        freeze_panes = (len(df.columns.names), len(df.index.names))
    if 'merge_cells' not in kwargs:
        kwargs['merge_cells'] = False
    if 'merge_cells' not in kwargs:
        kwargs['merge_cells'] = False
    if 'na_rep' not in kwargs:
        kwargs['na_rep'] = ''

    df.to_excel(file, sheet_name=sheet_name, freeze_panes=freeze_panes, **kwargs)


def fix_font_for_suppl_tables(suppl_tables, base_col_width=20, special_width_for_sheet_names=30, special_width_for_legends=200, fix_legends_sheet=True):
    side = openpyxl.styles.Side(border_style=None)
    no_border = openpyxl.styles.Border(left=side, right=side, top=side, bottom=side)
    font = openpyxl.styles.Font(name='Calibri', bold=True)
    alignment = openpyxl.styles.Alignment(horizontal=None, vertical=None)
    for table in suppl_tables:
        wb = openpyxl.load_workbook(suppl_tables[table]['filename'])
        assert set(suppl_tables[table]['index_cols'].keys()) == set(suppl_tables[table]['header_rows'].keys())
        excel_column_widths(wb, base_col_width=base_col_width, special_width_for_legends=special_width_for_legends, special_width_for_sheet_names=special_width_for_sheet_names)
        for sheet in ([get_legends_sheet_name(table)] if fix_legends_sheet else []) + list(suppl_tables[table]['index_cols'].keys()):
            ws = wb[sheet]
            for col in range(1, suppl_tables[table]['index_cols'][sheet] + 1):
                for cell in ws[openpyxl.utils.get_column_letter(col)]:
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = no_border
            for row in range(1, suppl_tables[table]['header_rows'][sheet] + 1):
                for cell in ws[row]:
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = no_border
        wb.save(suppl_tables[table]['filename'])


def excel_column_widths(workbook, base_col_width, special_width_for_legends=None, special_width_for_sheet_names=None):
    for sheet_name in workbook.sheetnames:
        worksheet: openpyxl.worksheet.Worksheet = workbook[sheet_name]
        sheet_prop: openpyxl.worksheet.dimensions.SheetFormatProperties = worksheet.sheet_format
        sheet_prop.baseColWidth = base_col_width

        if sheet_name.endswith("legends"):
            if special_width_for_legends:
                assert worksheet['B1'].value == 'Contents'
                worksheet.column_dimensions['B'].width = special_width_for_legends
            if special_width_for_sheet_names:
                assert worksheet['A1'].value == 'Sheet name'
                worksheet.column_dimensions['A'].width = special_width_for_sheet_names


def get_factors_summary_names(factors, factors_kind):
    if factors_kind is None:
        if all(['scar' in f.lower() for f in factors]):
            factors_col_name, factors_tab_name, factors_human_name = SCAR_COL, 'scarSize', 'scar size'
        else:
            factors_col_name, factors_tab_name, factors_human_name = HOST_FACTOR_COL, 'hostFactors', 'host factors'
    elif factors_kind.startswith('include_blood'):
        factors_col_name, factors_tab_name, factors_human_name = CELL_TYPE_COL, 'cellFreq', 'immune cell frequencies'
    else:
        raise ValueError
    return factors_col_name, factors_tab_name, factors_human_name


def get_day_suffix(visit):
    return f'd{VISIT_TO_DAYS[visit]}FC' if visit != 'V1' else 'd0'


def related_to_fig_str(fig_n):
    return f' (related to Figure{"s" if "-" in fig_n or "and" in fig_n else ""} {fig_n})' if fig_n is not None else ''


def gsea_description_prefix(top_n=None, region_sets=False, promoter_mapping=False, distal_mapping=False):
    assert not (promoter_mapping and distal_mapping)
    assert region_sets or promoter_mapping or distal_mapping
    return f'Chromatin accessibility {"region" if region_sets else "gene"} set enrichment of {f"the top {top_n:,}" if top_n is not None else ""} {"putative promoters" if promoter_mapping else "gene-linked (distal mapping) genomic regions" if distal_mapping else "genomic regions"}'


def atac_seq_description_prefix():
    return f'Association of chromatin accessibility'


def quantity_description(quantity, visit):
    return f'{f"{quantity} at baseline" if visit == "V1" else f"BCG-induced changes in {quantity} {VISIT_TO_DAYS[visit]} days after vaccination"}'


def suppl_table_cohort(df, drop_excluded=True, suppl_tables=None, table_name='S1', fig_n=None):
    print('All donors:', len(set(df['SAMPLE:DONOR'])))
    suppl_df = df.loc[~df['DONOR:IC_DATE_REAL'].isnull()]
    print('Donors with BCG:', len(set(suppl_df['SAMPLE:DONOR'])))
    if drop_excluded:
        suppl_df = suppl_df.loc[suppl_df['SAMPLE:EXCLUSION'].isnull()]
        print('Donors without exclusions:', len(set(suppl_df['SAMPLE:DONOR'])))
    suppl_df = suppl_df.loc[suppl_df['SAMPLE:TISSUE'] == 'PBMC']
    print('Donors with PBMCs:', len(set(suppl_df['SAMPLE:DONOR'])))

    suppl_df = suppl_df[DONOR_COLS.keys()]
    suppl_df['SAMPLE:VISIT'] = suppl_df['SAMPLE:VISIT'].map(VISITS_TO_TIMEPOINTS)
    suppl_df['DONOR:AGE'] = np.floor(suppl_df['DONOR:AGE']).astype(pd.Int64Dtype())
    suppl_df[DONOR_INTEGER_COLS] = suppl_df[DONOR_INTEGER_COLS].astype(pd.Int64Dtype())
    suppl_df[DONOR_BOOLEAN_COLS] = suppl_df[DONOR_BOOLEAN_COLS].astype(
        'boolean').astype('str').applymap(lambda x: '' if x == '<NA>' else x)
    suppl_df = suppl_df.rename(DONOR_COLS, axis=1)
    suppl_df = suppl_df.set_index(['Donor', 'Time point'])
    suppl_df = suppl_df.iloc[case_insensitive_sort_index(suppl_df.index)]

    if drop_excluded:
        assert suppl_df['Exclusion'].isnull().all()
        suppl_df = suppl_df.drop('Exclusion', axis=1)

    if suppl_tables is not None:
        sheet_name = '300BCG_cohort_annotation'
        suppl_tables[table_name]['info'][sheet_name] = f'The 300BCG cohort sample-level annotation{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_immune_response_scores(df, suppl_tables=None, table_name='S1', fig_n='7'):
    suppl_df = []
    for score_col in [
        'innate_nonspecific_24h_wo_LAC_IL10_IL1ra',
        'adaptive_MTB_7d'
    ]:
        score_name = f'{IMMUNITY_TYPES[score_col][0].upper()}{IMMUNITY_TYPES[score_col][1:]} score'
        log2_scores = df.loc[
            (df['SAMPLE:TISSUE'] == 'PBMC') & ~df[f'thm.{score_col}_V3'].isnull()].set_index('SAMPLE:DONOR')
        log2_scores = log2_scores.loc[
            ~log2_scores.index.duplicated(), f'thm.{score_col}_V3']
        suppl_df.append(np.power(2, log2_scores.rename(score_name)))
    suppl_df = pd.concat(suppl_df, axis=1)
    suppl_df.index.name = 'Donor'
    suppl_df = suppl_df.loc[~suppl_df.isnull().all(axis=1)]

    if suppl_tables is not None:
        sheet_name = 'immune_response_scores'
        suppl_tables[table_name]['info'][sheet_name] = f'Trained and adaptive immunity response scores used to annotate donors as responders and non-responders{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_atac_seq_stats(df, drop_excluded=True, suppl_tables=None, table_name='S1', fig_n=None):
    print('All donors:', len(set(df['SAMPLE:DONOR'])))
    suppl_df = df.loc[~df['DONOR:IC_DATE_REAL'].isnull()]
    print('Donors with BCG:', len(set(suppl_df['SAMPLE:DONOR'])))
    suppl_df = suppl_df.loc[~suppl_df['QC:PASS'].isnull()]
    print('Donors with (any) ATAC-seq:', len(set(suppl_df['SAMPLE:DONOR'])))
    if drop_excluded:
        suppl_df = suppl_df.loc[suppl_df['SAMPLE:EXCLUSION'].isnull()]
        print('Donors without exclusions:', len(set(suppl_df['SAMPLE:DONOR'])))

    suppl_df = suppl_df.loc[((suppl_df['SAMPLE:TISSUE'] == 'PBMC') |
                 (suppl_df['SAMPLE:TISSUE'].isin(['cd8t', 'nkcell', 'monocyte']) & (suppl_df['SAMPLE:VISIT'] == 'V1')))]

    suppl_df = suppl_df[ATAC_QC_COLS.keys()]
    suppl_df[ATAC_QC_INTEGER_COLS] = suppl_df[ATAC_QC_INTEGER_COLS].astype(pd.Int64Dtype())
    assert suppl_df.index.tolist() == (suppl_df['SAMPLE:DONOR'] + '_' + suppl_df['SAMPLE:VISIT'] + '_' +suppl_df['SAMPLE:TISSUE']).tolist()
    suppl_df['SAMPLE:VISIT'] = suppl_df['SAMPLE:VISIT'].map(VISITS_TO_TIMEPOINTS)
    suppl_df.index = suppl_df['SAMPLE:DONOR'] + '_' + suppl_df['SAMPLE:VISIT'] + '_' +suppl_df['SAMPLE:TISSUE']

    for t in set(suppl_df['SAMPLE:TISSUE']):
        print(t, len(set(suppl_df.loc[(suppl_df['SAMPLE:TISSUE'] == t), 'SAMPLE:DONOR'])),
              'QC:', len(set(suppl_df.loc[(suppl_df['SAMPLE:TISSUE'] == t) & (suppl_df['QC:PASS'] == True), 'SAMPLE:DONOR'])))

    suppl_df['QC:PASS'] = suppl_df['QC:PASS'].astype(str)
    suppl_df = suppl_df.rename(ATAC_QC_COLS, axis=1)
    suppl_df.index.name = 'Sample ID'
    suppl_df = suppl_df.iloc[case_insensitive_sort_index(suppl_df.index)]

    if suppl_tables is not None:
        sheet_name = 'ATACseq_statistics'
        suppl_tables[table_name]['info'][sheet_name] = f'The ATAC-seq sample-level quality statistics{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_peaks(df, allow_nulls=False, suppl_tables=None, sheet_name=None, info=None, table_name='S1', fig_n=None):
    suppl_df = df.copy()
    suppl_df['region_id'] = suppl_df.index.values
    suppl_df = simplify_peak_annot(suppl_df, location_col='genomic_location',
                                   feature_col='genomic_feature', regulatory_col='reg_feature', allow_nulls=allow_nulls)
    suppl_df = suppl_df[list(PEAK_ANNOT_COLS.keys())]
    for region_filter in [TSS_PROXIMAL, GENE_AND_DISTAL_10kb]:
        suppl_df = pd.concat([suppl_df, get_region_to_gene_mapping(suppl_df, region_filter)], axis=1)
    suppl_df.index = rename_regions(suppl_df.index, peaks_df=suppl_df)
    assert suppl_df.index.is_unique
    suppl_df['distance'] = suppl_df['distance'].astype(int)
    suppl_df = suppl_df.sort_values('region_id')
    suppl_df = suppl_df.rename(PEAK_ANNOT_COLS, axis=1)

    if suppl_tables is not None:
        suppl_tables[table_name]['info'][sheet_name] = info
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_immune_corr(corr_df, pval_df, Y_name, suppl_tables=None, table_name='S2', fig_n='1'):
    if Y_name == 'WB_PER_ML':
        assert len(corr_df) == 10
        Y_name_human = 'frequencies of 10 immune cell types'
    elif Y_name == 'CM':
        assert len(corr_df) == 73
        Y_name_human = '73 circulating inflammatory marker concentrations'
    elif Y_name == 'CYTO':
        assert len(corr_df) == 30
        assert corr_df.index.str.lower().str.contains('lactate').sum() == 2
        Y_name_human = '28 ex vivo cytokine-stimulus and two lactate-stimulus production capacity measurements'
    else:
        raise ValueError
    suppl_df = make_corr_long_table(corr_df, pval_df, symmetrical=True, col_name=Y_name, idx_name=Y_name)
    suppl_df['Data'] = 'All time points'
    old_index_names = suppl_df.index.names
    suppl_df = suppl_df.reset_index().set_index(['Data'] + old_index_names)

    if suppl_tables is not None:
        sheet_name = f'{NAMES[Y_name]}_corr'
        suppl_tables[table_name]['info'][sheet_name] = f'Pairwise Spearman correlations of {Y_name_human}{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_immune_donor_var(df, Y_name, suppl_tables=None, table_name='S2', fig_n='S2A'):
    suppl_df = df.reset_index().set_index('target')
    suppl_df = suppl_df.rename({'FVE': FVE_COL}, axis=1).loc[suppl_df['contrast'] == 'Across-donor', [FVE_COL]]
    assert (suppl_df[FVE_COL] > 11).any()
    suppl_df.loc[:, FVE_COL] = suppl_df.loc[:, FVE_COL] / 100
    suppl_df.index = pd.Index(rename_Y_vars(Y_name, suppl_df.index, with_time=True), name=LONG_NAMES[Y_name])
    suppl_df = suppl_df.iloc[case_insensitive_sort_index(suppl_df.index)]

    if suppl_tables is not None:
        sheet_name = f'{NAMES[Y_name]}_donorVar'
        suppl_tables[table_name]['info'][sheet_name] = f'Fraction of variance explained ({FVE_COL}) by donor for {HUMAN_NAMES[Y_name]}{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_atac_seq_donor_var(var_df, dispersions, peaks_df, limit_top_n, enr_top_n, enr_region_filter=GENE_AND_DISTAL_10kb,
                                   suppl_tables=None, table_name='S3', fig_n='S2'):
    suppl_df = var_df.reset_index().set_index(REGION_COL).join(dispersions.rename(NORM_DISP_COL)).rename({'FVE': FVE_COL}, axis=1)
    suppl_df = suppl_df.loc[suppl_df['contrast'] == 'Across-donor', [FVE_COL, NORM_DISP_COL]]
    assert (suppl_df[FVE_COL] > 11).any()
    suppl_df.loc[:, FVE_COL] = suppl_df.loc[:, FVE_COL] / 100

    if limit_top_n:
        enr_top_n_regions = get_top_n_regions(suppl_df.loc[peaks_to_genes(peaks_df.loc[suppl_df.index], **PEAKS_TO_GENES[enr_region_filter])],
                                              coef=None, direction=None, top_n=enr_top_n, rank_metric='FVE')
        if limit_top_n >= 1:
            suppl_df = suppl_df.sort_values(FVE_COL, ascending=False).head(limit_top_n)
        else:
            suppl_df = suppl_df.loc[suppl_df[FVE_COL] >= limit_top_n]
        assert np.isin(enr_top_n_regions, suppl_df.index).all()
        print(f'INFO: The {limit_top_n} selected regions contain the top {enr_top_n} regions with {enr_region_filter} mapping.')

    for region_filter in [TSS_PROXIMAL, GENE_AND_DISTAL_10kb]:
        suppl_df = pd.concat([suppl_df, get_region_to_gene_mapping(peaks_df.loc[suppl_df.index], region_filter)], axis=1)
    suppl_df.index = rename_regions(suppl_df.index, peaks_df=peaks_df.loc[suppl_df.index])
    suppl_df = suppl_df.sort_values(FVE_COL, ascending=False)

    if suppl_tables is not None:
        sheet_name = f'donorVar'
        suppl_tables[table_name]['info'][sheet_name] = f'Fraction of variance explained ({FVE_COL}) by donor for chromatin accessibility{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_enrich_donor_var(suppl_dict, method, limit_FDR, promoters, suppl_tables=None, table_name='S4', fig_n='S2'):
    for library in suppl_dict:
        suppl_dict[library] = suppl_dict[library].reset_index()[get_ENR_suppl_cols()].rename(
            rename_ENR_suppl_cols(region_sets=method == "lola"), axis=1)
        suppl_dict[library][LIBRARY_COL] = library
    suppl_df = pd.concat(suppl_dict.values()).set_index([LIBRARY_COL, REGION_SET_COL if method == 'lola' else GENE_SET_COL])
    suppl_df = suppl_df.sort_values([PADJ_COL, PVAL_COL])

    if limit_FDR:
        if isinstance(limit_FDR, tuple):
            fdr_limit, top_n_limit = limit_FDR
            top_n_df = suppl_df.sort_values([PADJ_COL, PVAL_COL]).groupby([LIBRARY_COL]).head(top_n_limit)
            fdr_df = suppl_df.loc[suppl_df[PADJ_COL] <= fdr_limit]
            suppl_df = pd.concat([top_n_df, fdr_df])
            suppl_df = suppl_df.loc[~suppl_df.index.duplicated()].sort_values([PADJ_COL, PVAL_COL])
        else:
            suppl_df = suppl_df.loc[suppl_df[PADJ_COL] <= limit_FDR]
        enrichments_counts = suppl_df.groupby([LIBRARY_COL])[PVAL_COL].count().rename(table_name)
        print(enrichments_counts)

    if suppl_tables is not None:
        sheet_name = f'donorVar'
        suppl_tables[table_name]['info'][sheet_name] = f'{gsea_description_prefix(top_n=1000, region_sets=method == "lola", promoter_mapping=promoters, distal_mapping=not promoters and method != "lola")} with the highest fraction of variance explained (FVE) by donor{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_atac_seq_d0_host_factors(df, coef, limit_top_n, enr_top_n, enr_rank_metric, enr_region_filter=GENE_AND_DISTAL_10kb,
                                         drop_coef_name=False, plot_enr_top_n_pvals=True, suppl_tables=None, table_name='S3', fig_n='2'):
    factor = rename_factor(coef.replace('DONOR.', '').replace('SAMPLE.', '').replace(
        'SEXM', 'SEX').replace('IncludingMenTrue', 'IncludingMen').replace('Last24hTrue', 'Last24h'),
                           caps=False, use_sample_collect_time=True)

    suppl_df = make_suppl_table_atac_seq_for_coef(
        df, coef, human_coef=factor[0].upper() + factor[1:], human_coef_col='Host factor', l2fc=False, limit_top_n=limit_top_n,
        enr_top_n=enr_top_n, enr_rank_metric=enr_rank_metric, enr_region_filter=enr_region_filter, plot_enr_top_n_pvals=plot_enr_top_n_pvals)
    if drop_coef_name:
        suppl_df = suppl_df.droplevel('Host factor')

    if suppl_tables is not None:
        sheet_name = f'd0_{camel_case(factor.replace("-", " "))}'
        suppl_tables[table_name]['info'][sheet_name] = f'{atac_seq_description_prefix()} with {factor} at baseline{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_enrich_d0_host_factors(suppl_dict, coefs, method, limit_FDR, promoters, drop_coef_name=False, suppl_tables=None, table_name='S4', fig_n='2'):
    human_coef_func = lambda x: rename_factor(x.replace('DONOR.', '').replace('SAMPLE.', '').replace(
        'SEXM', 'SEX').replace('IncludingMenTrue', 'IncludingMen').replace('Last24hTrue', 'Last24h'),
                                              caps=True, use_sample_collect_time=True)
    human_coef_col = 'Host factor'
    suppl_df = make_suppl_table_enrich_for_coefs(suppl_dict, coefs, human_coef_func, human_coef_col, method)

    if limit_FDR:
        if isinstance(limit_FDR, tuple):
            fdr_limit, top_n_limit = limit_FDR
            top_n_df = suppl_df.sort_values([PADJ_COL, PVAL_COL]).groupby([LIBRARY_COL, human_coef_col, ASSOC_COL]).head(top_n_limit)
            fdr_df = suppl_df.loc[suppl_df[PADJ_COL] <= fdr_limit]
            suppl_df = pd.concat([top_n_df, fdr_df])
            suppl_df = suppl_df.loc[~suppl_df.index.duplicated()].sort_values([PADJ_COL, PVAL_COL])
        else:
            suppl_df = suppl_df.loc[suppl_df[PADJ_COL] <= limit_FDR]
        enrichments_counts = suppl_df.groupby([LIBRARY_COL, human_coef_col, ASSOC_COL])[PVAL_COL].count().rename(table_name)
        print(enrichments_counts)

    if suppl_tables is not None:
        human_coef_func_no_caps = lambda x: rename_factor(x.replace('DONOR.', '').replace('SAMPLE.', '').replace(
            'SEXM', 'SEX').replace('IncludingMenTrue', 'IncludingMen').replace('Last24hTrue', 'Last24h'),
                                                  caps=False, use_sample_collect_time=True)
        for coef in coefs:
            factor_df = suppl_df.loc[suppl_df.index.get_level_values(human_coef_col) == human_coef_func(coef)]
            if drop_coef_name:
                factor_df = factor_df.droplevel(human_coef_col)
            sheet_name = f'd0_{camel_case(human_coef_func_no_caps(coef).replace("-", " "))}'
            suppl_tables[table_name]['info'][sheet_name] = f'{gsea_description_prefix(top_n=1000, region_sets=method == "lola", promoter_mapping=promoters, distal_mapping=not promoters and method != "lola")} most associated with {human_coef_func_no_caps(coef)} at baseline{related_to_fig_str(fig_n)}.'
            suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
            suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
            save_excel(factor_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def get_immune_phenotype(cyto, phenotypes):
    cytogroups = get_fold_change_cytogroups_masks(pd.Index([cyto]))
    phenotype = np.asarray(phenotypes)[[cytogroups[p].all() for p in phenotypes]]
    if len(phenotype) == 0:
        return None
    elif len(phenotype) == 1:
        return phenotype[0]
    else:
        raise ValueError


def make_suppl_table_atac_seq_d0_immune(suppl_dict, Y_name, proteins, visit,
                                        peaks_df, region_filter, limit_top_n, rank_metric, phenotypes=EWAS_PHENOTYPES):
    assert isinstance(proteins, pd.Index)
    for protein in suppl_dict:
        coef = f'{f"LFC_{visit}_CORR_" if visit != "V1" else ""}{safe_R_name(protein)}'
        
        if region_filter is not None or limit_top_n is not None:
            selected_regions_mask = np.isin(
                suppl_dict[protein].index,
                get_top_n_regions_both_directions(suppl_dict[protein], coef, top_n=limit_top_n,
                                                  rank_metric=rank_metric, region_filter=region_filter,
                                                  peaks_df=peaks_df)
            )
        else:
            selected_regions_mask = np.ones((len(suppl_dict[protein]),), dtype=bool)
        
        suppl_dict[protein] = suppl_dict[protein][get_limma_suppl_cols(coef)].rename(rename_limma_suppl_cols(coef), axis=1)
        for region_filter in [TSS_PROXIMAL, GENE_AND_DISTAL_10kb]:
            suppl_dict[protein] = pd.concat([suppl_dict[protein], get_region_to_gene_mapping(peaks_df.loc[suppl_dict[protein].index], region_filter)], axis=1)
        suppl_dict[protein][LONG_NAMES[Y_name]] = rename_Y_vars(Y_name, protein, with_time=True)
        phenotype = get_immune_phenotype(protein, phenotypes=phenotypes)
        suppl_dict[protein][PHENO_COL] = IMMUNITY_TYPES[phenotype] if phenotype else None
        suppl_dict[protein]['selected'] = False
        suppl_dict[protein].loc[selected_regions_mask, 'selected'] = True
        suppl_dict[protein]['selected'] = suppl_dict[protein]['selected'].astype(bool)
        suppl_dict[protein].index = rename_regions(suppl_dict[protein].index, peaks_df=peaks_df.loc[suppl_dict[protein].index])
    suppl_df = pd.concat([suppl_dict[protein] for protein in suppl_dict])
    suppl_df = suppl_df.reset_index().set_index([REGION_COL, LONG_NAMES[Y_name], PHENO_COL, 'selected'])
    suppl_df = suppl_df.sort_values(PVAL_COL)
    return suppl_df


def make_suppl_table_enrich_d0_immune(suppl_dict, Y_name, proteins, method, limit_pval, phenotypes=EWAS_PHENOTYPES):
    assert isinstance(proteins, pd.Index)
    for library in suppl_dict:
        for direction in [-1, 1]:
            for protein in proteins:
                formatted_protein = rename_Y_vars(Y_name, protein, with_time=True)
                suppl_dict[library][direction][protein] = suppl_dict[library][direction][protein][get_ENR_suppl_cols()].rename(
                    rename_ENR_suppl_cols(region_sets=method == "lola"), axis=1)
                suppl_dict[library][direction][protein][LIBRARY_COL] = library
                suppl_dict[library][direction][protein][LONG_NAMES[Y_name]] = formatted_protein
                suppl_dict[library][direction][protein][ASSOC_COL] = NEG if direction < 0 else POS
                phenotype = get_immune_phenotype(protein, phenotypes=phenotypes)
                suppl_dict[library][direction][protein][PHENO_COL] = IMMUNITY_TYPES[phenotype] if phenotype else None
                suppl_dict[library][direction][protein]['selected'] = False
                if limit_pval:
                    selected_sets_mask = suppl_dict[library][direction][protein][PVAL_COL] <= limit_pval
                else:
                    selected_sets_mask = np.ones((len(suppl_dict[library][direction][protein]),), dtype=bool)
                suppl_dict[library][direction][protein].loc[selected_sets_mask, 'selected'] = True
                suppl_dict[library][direction][protein]['selected'] = suppl_dict[library][direction][protein]['selected'].astype(bool)
    suppl_df = pd.concat([suppl_dict[library][direction][protein] for library in suppl_dict for direction in [-1, 1] for protein in proteins])
    suppl_df = suppl_df.set_index([LIBRARY_COL, REGION_SET_COL if method == 'lola' else GENE_SET_COL, LONG_NAMES[Y_name], PHENO_COL, 'selected', ASSOC_COL])
    suppl_df = suppl_df.sort_values([PADJ_COL, PVAL_COL])
    return suppl_df


def suppl_table_atac_seq_d0_immune(suppl_df, Y_name, visit, split_by_phenotype, split_by_protein, only_selected,
                                   method=None, promoters=False, gene_sets=False, drop_padj=False, force_sheet_name=None,
                                   suppl_tables=None, table_name=None, fig_n=None, phenotypes=EWAS_PHENOTYPES):
    assert sum([split_by_phenotype, split_by_protein]) < 2
    suppl_df = suppl_df.sort_values(PVAL_COL)
    protein_index = suppl_df.index.get_level_values(LONG_NAMES[Y_name])
    selected_index = suppl_df.index.get_level_values('selected') if only_selected else np.ones((len(suppl_df),), dtype=bool)
    if 'selected' in suppl_df.index.names:
        suppl_df = suppl_df.droplevel('selected')
    phenotype_index = suppl_df.index.get_level_values(PHENO_COL) if split_by_phenotype else nans((len(suppl_df),))
    if PHENO_COL in suppl_df.index.names:
        suppl_df = suppl_df.droplevel(PHENO_COL)
    atac_descr = f'{gsea_description_prefix(top_n=1000, region_sets=method == "lola", promoter_mapping=promoters, distal_mapping=not promoters and method != "lola")} most associated' if gene_sets else atac_seq_description_prefix()

    if drop_padj:
        suppl_df = suppl_df.drop(PADJ_COL, axis=1)

    if split_by_phenotype:
        assert not phenotype_index.isnull().any()
        for phenotype in phenotypes:
            phenotype_mask = phenotype_index == IMMUNITY_TYPES[phenotype]
            if not drop_padj:
                print('adjusting per phenotype')
                suppl_df.loc[phenotype_mask, PADJ_COL] = adjusted_pvals(suppl_df.loc[phenotype_mask, PVAL_COL])
            if force_sheet_name:
                sheet_name = force_sheet_name
            else:
                sheet_name = f'd0_{NAMES[Y_name]}_{get_day_suffix(visit)}_{SHORT_IMMUNITY_TYPES[phenotype]}'
            suppl_tables[table_name]['info'][sheet_name] = f'{atac_descr} with {quantity_description(f"{HUMAN_NAMES[Y_name]} with regards to {IMMUNITY_TYPES[phenotype]} phenotype", visit)}{related_to_fig_str(fig_n)}.'
            suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
            suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
            save_excel(suppl_df.loc[phenotype_mask & (selected_index == True)], suppl_tables[table_name]['file'], sheet_name=sheet_name)
    else:
        if split_by_protein:
            sorted_proteins_index = protein_index[case_insensitive_sort_index(protein_index)]
            sorted_proteins_index = sorted_proteins_index[~sorted_proteins_index.duplicated()]
            for protein in sorted_proteins_index:
                protein_mask = protein_index == protein
                if not drop_padj:
                    print('adjusting per protein')
                    suppl_df.loc[protein_mask, PADJ_COL] = adjusted_pvals(suppl_df.loc[protein_mask, PVAL_COL])
                protein_wo_spaces = protein.replace('(', '').replace(')', '').replace(',', '').replace(' ', '.')
                if force_sheet_name:
                    sheet_name = force_sheet_name
                else:
                    sheet_name = f'd0_{protein_wo_spaces}_{get_day_suffix(visit)}'
                suppl_tables[table_name]['info'][sheet_name] = f'{atac_descr} with {quantity_description(protein, visit)}{related_to_fig_str(fig_n)}.'
                suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
                suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
                save_excel(suppl_df.loc[protein_mask & (selected_index == True)], suppl_tables[table_name]['file'], sheet_name=sheet_name)
        else:
            if not drop_padj:
                print('adjusting alltogeter')
                suppl_df[PADJ_COL] = adjusted_pvals(suppl_df[PVAL_COL])
            if force_sheet_name:
                sheet_name = force_sheet_name
            else:
                sheet_name = f'd0_{NAMES[Y_name]}_{get_day_suffix(visit)}'
            suppl_tables[table_name]['info'][sheet_name] = f'{atac_descr} with {quantity_description(HUMAN_NAMES[Y_name], visit)}{related_to_fig_str(fig_n)}.'
            suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
            suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
            save_excel(suppl_df[(selected_index == True)], suppl_tables[table_name]['file'], sheet_name=sheet_name)


def suppl_table_immune_LFC(df, Y_name, visit, suppl_tables=None, table_name='S2', fig_n='5'):
    assert len(set(df.index.get_level_values('contrast'))) == 1
    suppl_df = df.droplevel('contrast')[get_LR_suppl_cols()].rename(rename_LR_suppl_cols(l2fc=True), axis=1)
    suppl_df.index = pd.Index(rename_Y_vars(Y_name, suppl_df.index, with_time=True), name=LONG_NAMES[Y_name])
    suppl_df = suppl_df.iloc[case_insensitive_sort_index(suppl_df.index)]

    if suppl_tables is not None:
        sheet_name = f'{NAMES[Y_name]}_{get_day_suffix(visit)}'
        suppl_tables[table_name]['info'][sheet_name] = f'BCG-induced changes in {HUMAN_NAMES[Y_name]} {VISIT_TO_DAYS[visit]} days after vaccination{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_immune_LFC_corr(corr_df, pval_df, Y_name, visit_x, visit_y, symmetrical, suppl_tables=None, table_name='S2', fig_n='5'):
    # For simplicity of producing a meaningful table description:
    assert visit_y != 'V1' and (visit_x == 'V1' or visit_x == visit_y), (visit_x, visit_y)
    visit_x = get_day_suffix(visit_x)
    visit_y = get_day_suffix(visit_y)
    # transpose so that first index predicts the second (x (columns) predicts y (index))
    suppl_df = make_corr_long_table(corr_df.T, pval_df.T, symmetrical=symmetrical,
                                    idx_name_suffix=f'_{visit_x}', col_name_suffix=f'_{visit_y}')

    if suppl_tables is not None:
        sheet_name = f'{NAMES[Y_name]}_{visit_y}_corr{f"_{visit_x}" if visit_x != visit_y else ""}'
        suppl_tables[table_name]['info'][sheet_name] = f'Pairwise Spearman correlations of BCG-induced changes in {HUMAN_NAMES[Y_name]} {visit_y[1:-2]} days after vaccination{f" with baseline {HUMAN_NAMES[Y_name]}" if visit_x == "d0" else ""}{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def make_suppl_table_immune_d0_host_factors(df, Y_name, factors_kind):
    suppl_df = df[get_LR_suppl_cols()].rename(rename_LR_suppl_cols(), axis=1)
    targets = rename_Y_vars(Y_name, suppl_df.index.get_level_values('target'), with_time=True)
    factors = rename_factors(suppl_df.index.get_level_values('contrast').str.replace(
        'DONOR_', '').str.replace('SAMPLE_', '').str.replace('\[T\.M\]$', ''),
                             caps=True, use_sample_collect_time=True, include_V1_prefix=False)
    factors_col_name, factors_tab_name, factors_human_name = get_factors_summary_names(factors, factors_kind)
    suppl_df.index = pd.MultiIndex.from_arrays([targets, factors], names=[LONG_NAMES[Y_name], factors_col_name])
    suppl_df = suppl_df.iloc[case_insensitive_sort_index(suppl_df.index)]
    return suppl_df, factors_tab_name, factors_human_name


def suppl_table_immune_d0_host_factors(df, Y_name, factors_kind, suppl_tables=None, table_name='S2', fig_n='2'):
    suppl_df, factors_tab_name, factors_human_name = make_suppl_table_immune_d0_host_factors(df, Y_name, factors_kind)

    if suppl_tables is not None:
        sheet_name = f'{NAMES[Y_name]}_d0_{factors_tab_name}'
        suppl_tables[table_name]['info'][sheet_name] = f'Association of {HUMAN_NAMES[Y_name]} with {factors_human_name} at baseline{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_immune_LFC_host_factors(df, Y_name, visit, factors_kind, suppl_tables=None, table_name='S2', fig_n='5'):
    suppl_df, factors_tab_name, factors_human_name = make_suppl_table_immune_d0_host_factors(df, Y_name, factors_kind)
    
    if suppl_tables is not None:
        sheet_name = f'{NAMES[Y_name]}_{get_day_suffix(visit)}_{factors_tab_name}'
        suppl_tables[table_name]['info'][sheet_name] = f'Association of BCG-induced changes in {HUMAN_NAMES[Y_name]} {VISIT_TO_DAYS[visit]} days after vaccination with {factors_human_name}{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_immune_LFC_season(df, cos_coefs_df, sin_coefs_df, Y_name, visit, suppl_tables=None, table_name='S2', fig_n='5'):
    suppl_df = df[get_LR_suppl_cols()].rename(rename_LR_suppl_cols(F_test=True), axis=1).drop(COEF_COL, axis=1)
    cos_coefs = cos_coefs_df.droplevel('contrast')['Coef'].rename(f'{COEF_COL} (cosine)')
    sin_coefs = sin_coefs_df.droplevel('contrast')['Coef'].rename(f'{COEF_COL} (sine)')
    suppl_df = pd.DataFrame(cos_coefs).join(sin_coefs, how='outer').join(suppl_df, how='right')
    suppl_df.index = pd.Index(rename_Y_vars(Y_name, suppl_df.index, with_time=True), name=LONG_NAMES[Y_name])
    suppl_df = suppl_df.iloc[case_insensitive_sort_index(suppl_df.index)]

    if suppl_tables is not None:
        sheet_name = f'{NAMES[Y_name]}_{get_day_suffix(visit)}_season'
        suppl_tables[table_name]['info'][sheet_name] = f'Seasonal variability of BCG-induced changes in {HUMAN_NAMES[Y_name]} {VISIT_TO_DAYS[visit]} days after vaccination{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_atac_seq_LFC_season(df, cos_coefs, sin_coefs, visit, limit_FDR, suppl_tables=None, table_name='S7', fig_n='5'):
    coef = 'SEASON.{}'.format(visit)
    suppl_df = df[get_limma_suppl_cols(coef, F_test=True, include_coef=False) + [CLUSTER_COL]].rename(rename_limma_suppl_cols(coef), axis=1)
    suppl_df = pd.DataFrame(cos_coefs).join(sin_coefs, how='outer').join(suppl_df, how='right')
    for region_filter in [TSS_PROXIMAL, GENE_AND_DISTAL_10kb]:
        suppl_df = pd.concat([suppl_df, get_region_to_gene_mapping(df.loc[suppl_df.index], region_filter)], axis=1)
    if limit_FDR:
        suppl_df = suppl_df.loc[suppl_df[PADJ_COL] < limit_FDR]
    suppl_df.index = rename_regions(suppl_df.index, peaks_df=df.loc[suppl_df.index])
    suppl_df = suppl_df.sort_values(PVAL_COL)

    if suppl_tables is not None:
        sheet_name = f'd{VISIT_TO_DAYS[visit]}FC_season'
        suppl_tables[table_name]['info'][sheet_name] = f'Seasonal variability of BCG-induced chromatin remodeling {VISIT_TO_DAYS[visit]} days after vaccination{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_enrich_LFC_season(suppl_dict, visit, method, limit_FDR, promoters, suppl_tables=None, table_name='S8', fig_n='5'):
    for library in suppl_dict:
        for cluster in suppl_dict[library]:
            suppl_dict[library][cluster] = suppl_dict[library][cluster].reset_index()[get_ENR_suppl_cols()].rename(rename_ENR_suppl_cols(region_sets=method == 'lola'), axis=1)
            suppl_dict[library][cluster][LIBRARY_COL] = library
            suppl_dict[library][cluster][CLUSTER_COL] = cluster
    suppl_df = pd.concat([suppl_dict[library][cluster] for library in suppl_dict for cluster in suppl_dict[library]])
    suppl_df = suppl_df.set_index([LIBRARY_COL, REGION_SET_COL if method == 'lola' else GENE_SET_COL, CLUSTER_COL])
    suppl_df = suppl_df.sort_values([PADJ_COL, PVAL_COL])

    if limit_FDR:
        if isinstance(limit_FDR, tuple):
            fdr_limit, top_n_limit = limit_FDR
            top_n_df = suppl_df.sort_values([PADJ_COL, PVAL_COL]).groupby([LIBRARY_COL, CLUSTER_COL]).head(top_n_limit)
            fdr_df = suppl_df.loc[suppl_df[PADJ_COL] <= fdr_limit]
            suppl_df = pd.concat([top_n_df, fdr_df])
            suppl_df = suppl_df.loc[~suppl_df.index.duplicated()].sort_values([PADJ_COL, PVAL_COL])
        else:
            suppl_df = suppl_df.loc[suppl_df[PADJ_COL] <= limit_FDR]
        enrichments_counts = suppl_df.groupby([LIBRARY_COL, CLUSTER_COL])[PVAL_COL].count().rename(table_name)
        print(enrichments_counts)

    if suppl_tables is not None:
        sheet_name = f'd{VISIT_TO_DAYS[visit]}FC_season'
        suppl_tables[table_name]['info'][sheet_name] = f'{gsea_description_prefix(top_n=None, region_sets=method == "lola", promoter_mapping=promoters, distal_mapping=not promoters and method != "lola")} associated with seasonal variability of BCG-induced chromatin remodeling {VISIT_TO_DAYS[visit]} days after vaccination{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_atac_seq_LFC(df, coef, comparison, comparison_text, limit_top_n, enr_top_n, enr_rank_metric,
                             enr_region_filter=GENE_AND_DISTAL_10kb, drop_coef_name=False, plot_enr_top_n_pvals=True,
                             gene_columns=None, rename_regions_to_coord=True, suppl_tables=None, sheet_name=None, table_name='S7', fig_n='7'):
    suppl_df = make_suppl_table_atac_seq_for_coef(
        df, coef, human_coef=comparison, human_coef_col=COMPARISON_COL, l2fc=True, limit_top_n=limit_top_n,
        enr_top_n=enr_top_n, enr_rank_metric=enr_rank_metric, enr_region_filter=enr_region_filter,
        plot_enr_top_n_pvals=plot_enr_top_n_pvals, gene_columns=gene_columns, rename_regions_to_coord=rename_regions_to_coord)
    if drop_coef_name:
        suppl_df = suppl_df.droplevel(COMPARISON_COL)

    if suppl_tables is not None:
        suppl_tables[table_name]['info'][sheet_name] = f'{atac_seq_description_prefix()} with {comparison_text}{related_to_fig_str(fig_n)}.'
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_enrich_LFC(suppl_dict, contrasts, phenotype, method, limit_FDR, promoters, drop_coef_name=False, suppl_tables=None, table_name='S8', fig_n='7', info=None, sheet_name=None):
    human_coef_func = lambda x: contrasts[x][0]
    human_coef_col = COMPARISON_COL
    suppl_df = make_suppl_table_enrich_for_coefs(suppl_dict, contrasts, human_coef_func, human_coef_col, method)

    if limit_FDR:
        if isinstance(limit_FDR, tuple):
            fdr_limit, top_n_limit = limit_FDR
            top_n_df = suppl_df.sort_values([PADJ_COL, PVAL_COL]).groupby([LIBRARY_COL, human_coef_col, ASSOC_COL]).head(top_n_limit)
            fdr_df = suppl_df.loc[suppl_df[PADJ_COL] <= fdr_limit]
            suppl_df = pd.concat([top_n_df, fdr_df])
            suppl_df = suppl_df.loc[~suppl_df.index.duplicated()].sort_values([PADJ_COL, PVAL_COL])
        else:
            suppl_df = suppl_df.loc[suppl_df[PADJ_COL] <= limit_FDR]
        enrichments_counts = suppl_df.groupby([LIBRARY_COL, human_coef_col, ASSOC_COL])[PVAL_COL].count().rename(table_name)
        print(enrichments_counts)

    if suppl_tables is not None:
        for coef in contrasts:
            if coef is not None:
                comparison_df = suppl_df.loc[suppl_df.index.get_level_values(human_coef_col) == human_coef_func(coef)]
                if drop_coef_name:
                    comparison_df = comparison_df.droplevel(human_coef_col)
            sheet_name = sheet_name if sheet_name is not None else f'{contrasts[coef][1]}_{SHORT_IMMUNITY_TYPES[phenotype]}'
            suppl_tables[table_name]['info'][sheet_name] = info if info is not None else \
                f'{gsea_description_prefix(top_n=1000, region_sets=method == "lola", promoter_mapping=promoters, distal_mapping=not promoters and method != "lola")} most associated with {contrasts[coef][2]}{related_to_fig_str(fig_n)}.'
            suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
            suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
            save_excel(comparison_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_enrich_LFC_extras(suppl_df, phenotype, suppl_tables=None, table_name='S10', fig_n='7',
                                  info=None, sheet_name=None):
    if info is None:
        info = f'{gsea_description_prefix(top_n=1000, region_sets=True)} most associated with {IMMUNITY_TYPES[phenotype]} ({SHORT_IMMUNITY_TYPES[phenotype]}) responders at baseline and BCG-induced chromatin remodeling {VISIT_TO_DAYS["V2"]} and {VISIT_TO_DAYS["V3"]} days after vaccination{related_to_fig_str(fig_n)}.'
    if sheet_name is None:
        sheet_name = f'extras_{SHORT_IMMUNITY_TYPES[phenotype]}'

    if suppl_tables is not None:
        sheet_name = sheet_name
        suppl_tables[table_name]['info'][sheet_name] = info
        suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
        suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
        save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def suppl_table_rna_seq_LFC(de_RNA_dict, background, comparison, cmp_description, suppl_tables, sheet_name, table_name, fig_n):
    # def quick_rename_rna_group(x):
    #     return x.replace('(', '').replace(')', '').replace('control ', '').replace(
    #         'Day', 'Unstim. day').replace('vs. day', 'vs. unstim. day')

    suppl_df = []
    for stat, stat_name in [
        ('coef', L2FC_COL),
        ('tval', T_COL),
        ('pval', PVAL_COL),
    ]:
        suppl_df.append(de_RNA_dict[stat].loc[background, comparison].rename(stat_name))
        if stat == 'pval':
            suppl_df.append(adjusted_pvals(suppl_df[-1]).rename(PADJ_COL))
    suppl_df = pd.concat(suppl_df, axis=1)

    suppl_df[COMPARISON_COL] = comparison
    suppl_df.index.name = GENE_COL
    suppl_df = suppl_df.set_index([suppl_df.index, COMPARISON_COL])
    suppl_df = suppl_df.iloc[suppl_df[T_COL].abs().argsort()[::-1]]
    assert suppl_df[PVAL_COL].tolist() == suppl_df[PVAL_COL].sort_values().tolist()

    suppl_tables[table_name]['info'][sheet_name] = \
        f'{cmp_description} 90 days after {"BCG " if "BCG" not in cmp_description else ""}vaccination{related_to_fig_str(fig_n)}.'
    suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
    suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
    save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)

    return suppl_df


def suppl_table_rna_seq_epi_potent(suppl_df, description, suppl_tables, sheet_name, table_name, fig_n):
    suppl_tables[table_name]['info'][sheet_name] = \
        f'{description}{related_to_fig_str(fig_n)}.'
    suppl_tables[table_name]['header_rows'][sheet_name] = len(suppl_df.columns.names)
    suppl_tables[table_name]['index_cols'][sheet_name] = len(suppl_df.index.names)
    save_excel(suppl_df, suppl_tables[table_name]['file'], sheet_name=sheet_name)
    return suppl_df


def get_phenotype_contrasts(phenotype, swap_to_non_resp):
    r1 = 'Resp' if not swap_to_non_resp else 'Non-R'
    r2 = 'Non-R' if not swap_to_non_resp else 'Resp'
    contrasts = {
        f'thm.{phenotype}_V3_FC1.2_R':    (f'{r1}. vs. {r2}. (d0)', f'd0_{r1}_vs_{r2.replace("-", "")}', f'{IMMUNITY_TYPES[phenotype]} ({SHORT_IMMUNITY_TYPES[phenotype]}) {"non-" if swap_to_non_resp else ""}responders at baseline'),
        f'thm.{phenotype}_V3_FC1.2_R.V2': ('Day 14 vs. day 0 (Resp.)', 'd14FC_Resp', f'BCG-induced chromatin remodeling 14 days after vaccination with regards to {IMMUNITY_TYPES[phenotype]} ({SHORT_IMMUNITY_TYPES[phenotype]}) responders'),
        f'thm.{phenotype}_V3_FC1.2_R.V3': ('Day 90 vs. day 0 (Resp.)', 'd90FC_Resp', f'BCG-induced chromatin remodeling 90 days after vaccination with regards to {IMMUNITY_TYPES[phenotype]} ({SHORT_IMMUNITY_TYPES[phenotype]}) responders')
    }
    return contrasts


def make_suppl_table_enrich_for_coefs(suppl_dict, coefs, human_coef_func, human_coef_col, method):
    for library in suppl_dict:
        for direction in suppl_dict[library]:
            for coef in coefs:
                human_coef = human_coef_func(coef)
                suppl_dict[library][direction][coef] = suppl_dict[library][direction][coef].reset_index()
                suppl_dict[library][direction][coef] = suppl_dict[library][direction][coef][get_ENR_suppl_cols()].rename(rename_ENR_suppl_cols(region_sets=method == "lola"), axis=1)
                suppl_dict[library][direction][coef][LIBRARY_COL] = library
                suppl_dict[library][direction][coef][human_coef_col] = human_coef
                suppl_dict[library][direction][coef][ASSOC_COL] = NEG if direction < 0 else POS
    suppl_df = pd.concat([suppl_dict[library][direction][coef] for library in suppl_dict for direction in  suppl_dict[library] for coef in coefs])
    suppl_df = suppl_df.set_index([LIBRARY_COL, rename_ENR_suppl_cols(region_sets=method == "lola")['Term'], human_coef_col, ASSOC_COL])
    suppl_df = suppl_df.sort_values([PADJ_COL, PVAL_COL])
    return suppl_df


def make_suppl_table_atac_seq_for_coef(df, coef, human_coef, human_coef_col, l2fc, limit_top_n,
                                       enr_top_n, enr_rank_metric, enr_region_filter=GENE_AND_DISTAL_10kb,
                                       plot_enr_top_n_pvals=False, gene_columns=None, rename_regions_to_coord=True):
    suppl_df = df[get_limma_suppl_cols(coef)].rename(rename_limma_suppl_cols(coef, l2fc=l2fc), axis=1)

    if limit_top_n:
        enr_top_n_regions = get_top_n_regions_both_directions(
            df, coef, top_n=enr_top_n, rank_metric=enr_rank_metric, region_filter=enr_region_filter)
        if limit_top_n >= 1:
            suppl_df = suppl_df.sort_values(PVAL_COL).head(limit_top_n)
        else:
            assert enr_rank_metric == 'p.value'
            suppl_df = suppl_df.loc[suppl_df[PVAL_COL] <= limit_top_n]
        assert np.isin(enr_top_n_regions, suppl_df.index).all(), len(enr_top_n_regions.difference(suppl_df.index))
        print(f'INFO: The {limit_top_n} selected regions contain the top {enr_top_n} regions with {enr_region_filter} mapping.')

        if plot_enr_top_n_pvals:
            enr_top_n_df = suppl_df.loc[enr_top_n_regions]
            _, ax = plt.subplots(figsize=(3, 1))
            sns.distplot(enr_top_n_df.loc[enr_top_n_df[L2FC_COL if l2fc else COEF_COL] < 0, PVAL_COL], kde=False, label='neg')
            sns.distplot(enr_top_n_df.loc[enr_top_n_df[L2FC_COL if l2fc else COEF_COL] > 0, PVAL_COL], kde=False, label='pos')
            ax.set_title(f'{human_coef}: top {enr_top_n} regions for enrichment')
            ax.legend(bbox_to_anchor=(1, 1))
            ax.axvline(0.05, c='k')
            sns.despine()
            plt.show()

    for region_filter in [TSS_PROXIMAL, GENE_AND_DISTAL_10kb]:
        suppl_df = pd.concat([suppl_df, get_region_to_gene_mapping(
            df.loc[suppl_df.index], region_filter, columns='gene_name' if gene_columns is None else gene_columns,
            name=None if gene_columns is None else region_filter
        )], axis=1)
    suppl_df[human_coef_col] = human_coef
    if rename_regions_to_coord:
        suppl_df.index = rename_regions(suppl_df.index, peaks_df=df.loc[suppl_df.index])
    suppl_df.index.name = REGION_COL
    suppl_df = suppl_df.reset_index().set_index([REGION_COL, human_coef_col])
    suppl_df = suppl_df.sort_values(PVAL_COL)

    return suppl_df


def cross_RNA_and_EPI(rna_results, epi_results_df, celltypes, rna_coefs, epi_coefs, exp_frac_thr, rna_FDR, epi_P, region_filter,
                      invert_responders=True, verbose=False):
    OR = {}
    P = {}
    for epi_coef in epi_coefs:
        short_epi_coef = (epi_coef.replace("_R", "_NonResp") if invert_responders and epi_coef.endswith("_R") else epi_coef).split("FC1.2_")[-1]
        OR[short_epi_coef] = {}
        P[short_epi_coef] = {}

        for rna_direction in [1, -1]:
            ors = []
            ps = []

            for celltype in celltypes:
                if verbose:
                    print(celltype)
                ors.append([np.nan] * len(rna_coefs))
                ps.append([np.nan] * len(rna_coefs))

                for i, rna_coef in enumerate(rna_coefs):
                    if rna_coef in rna_results[celltype]['tval'] and not rna_results[celltype]['tval'][rna_coef].isnull().all():
                        rna_tval = rna_results[celltype]['tval'][rna_coef]
                        rna_padj = rna_results[celltype]['padj' if rna_FDR < 1 else 'pval'][rna_coef]
                        assert rna_padj.index.is_unique

                        peaks_mask = peaks_to_genes(epi_results_df, **PEAKS_TO_GENES[region_filter])
                        peaks_and_genes = epi_results_df.loc[peaks_mask, 'gene_name'].str.upper()
                        peaks_and_genes = peaks_and_genes.str.split(', ', expand=True).stack().rename('gene_name')
                        peaks_and_genes.index = peaks_and_genes.index.get_level_values(0)

                        background = rna_results[celltype]["expFrac"][rna_coef].index[
                            rna_results[celltype]["expFrac"][rna_coef] > exp_frac_thr]
                        background = background.intersection(set(peaks_and_genes))
                        # print('background:', len(background))

                        for epi_direction in [1, -1]:
                            assert rna_padj.index.equals(rna_tval.index)

                            epi_tval = epi_results_df.loc[peaks_mask, stat_col(epi_coef)]
                            if invert_responders and epi_coef.endswith('_R'):
                                epi_tval = -1 * epi_tval
                            epi_tval = epi_tval.loc[epi_tval * epi_direction > 0]
                            idx = pd.concat([epi_tval, peaks_and_genes.loc[epi_tval.index]], axis=1).groupby(
                                'gene_name')[stat_col(epi_coef)].apply(lambda x: x.abs().idxmax()).values
                            assert len(idx) == len(set(idx))
                            assert (epi_tval.loc[idx] > 0).all() if epi_direction > 0 else (epi_tval.loc[idx] < 0).all()

                            epi_pval = epi_results_df.loc[idx, pval_col(epi_coef)]
                            epi_pval.index = peaks_and_genes.loc[epi_pval.index].values

                            if rna_FDR < 1:
                                rna_genes = rna_padj.loc[(rna_tval * rna_direction > 0) & (rna_padj <= rna_FDR)].index.intersection(background)
                            else:
                                rna_genes = rna_padj.loc[(rna_tval * rna_direction > 0) & rna_padj.index.isin(background)].sort_values().head(rna_FDR).index
                            epi_genes = epi_pval.loc[epi_pval <= epi_P].index.intersection(background)

                            odds_ratio, pval, overlap = fisher_enrichment(epi_genes, rna_genes, background)
                            odds_ratio = (1 if epi_direction == rna_direction else -1) * odds_ratio

                            if np.isnan(ors[-1][i]) or (abs(ors[-1][i]) < abs(odds_ratio) and (ps[-1][i] > 0.05 or pval <= 0.05)):
                                if not np.isnan(ors[-1][i]):
                                    print(f'WARNING: overwriting the other direction ({ors[-1][i]:.2f}, {ps[-1][i]:.2e}) with {odds_ratio:.2f}, {pval:.2e}')
                                ors[-1][i] = odds_ratio
                                ps[-1][i] = pval
                            if verbose:
                                print(f'{short_epi_coef:>7} {rna_coef:>23} {region_filter} {rna_direction:>2} (RNA) {epi_direction:>2} (EPI) {overlap:>9} (RNA) {str(len(epi_genes)):>4} (EPI) OR={abs(odds_ratio):.2f} P={pval:.2e} {"*" if pval < 0.05 else ""} {"(+)" if pval < 0.05 and epi_direction == rna_direction else "(-)" if pval < 0.05 and epi_direction != rna_direction else ""}')
            OR[short_epi_coef][rna_direction] = np.asarray(ors)
            P[short_epi_coef][rna_direction] = np.asarray(ps)

    return OR, P


def plot_RNA_and_EPI(OR, P, celltypes, test_name, exp_frac_thr, rna_FDR, epi_P, region_filter, rename_macrophages=True):
    fig, axs = plt.subplots(1, 2 * len(OR.keys()), figsize=(len(OR.keys()) * 10, 5), sharey=False, sharex=False)
    plt.subplots_adjust(wspace=1, hspace=1)
    for i, short_epi_coef in enumerate(OR.keys()):
        for j, rna_direction in enumerate([1, -1]):
            ax = axs[i * 2 + j]
            ax = sns.heatmap(OR[short_epi_coef][rna_direction], center=0, vmin=-3.1, vmax=3.1,
                             cbar_kws={'label': 'Odds ratio\nRNA & epigen. disagree        RNA & epigen. agree'},
                             annot=pvals_to_asterisks(P[short_epi_coef][rna_direction]),
                             fmt='', cmap='RdBu_r', ax=ax)
            ax.set_xticks(np.arange(len(OR[short_epi_coef][rna_direction].keys())) + 0.5)
            ax.set_xticklabels(list(OR[short_epi_coef][rna_direction].keys()), rotation=90)
            ax.set_yticks(np.arange(len(celltypes)) + 0.5)
            ax.set_yticklabels(pd.Series(celltypes).str.replace('Intermediate macrophages', 'Monocytes').values if rename_macrophages else celltypes, rotation=0)
            ax.set_title(
                f'RNA: {"UP" if rna_direction > 0 else "DOWN"}-regulated\nEpigen.: {short_epi_coef.replace(".", "_")} {region_filter}\n{test_name if test_name else ""} FDR < {rna_FDR} P < {epi_P} exp > {exp_frac_thr}')
    savefig(f'RNA_{f"{test_name}_" if test_name else ""}fdr{rna_FDR}_epigen_pval{epi_P}_exp{exp_frac_thr}.pdf', fig)
    plt.show()

def pvals_to_asterisks(pvals_matrix):
    return np.asarray(['***' if p <= 0.001 else '**' if p <= 0.01 else '*' if p <= 0.05 else '' \
                       for row in pvals_matrix for p in row]).reshape(pvals_matrix.shape)


def cell_frac_analysis():
    raise ValueError
# The code below is from a notebook for cell_frac_analysis
#     NOT_PBMC = [
#         'MAIT cells', 'Megakaryocytes/platelets',
#     ]
#     adata_LF = adata_LF[~adata_LF.obs['low_level_majority_voting'].isin(NOT_PBMC)]
#
#     REMAP_CELLS = {
#         # 'Megakaryocytes/platelets': 'Other',
#         # 'MAIT cells': 'Other',
#
#         'Tem/Trm cytotoxic T cells': 'Cytotoxic T cells',
#         'Tem/Temra cytotoxic T cells': 'Cytotoxic T cells',
#         'Tcm/Naive cytotoxic T cells': 'Cytotoxic T cells',
#
#         'Tcm/Naive helper T cells': 'Helper T cells',
#         'Tem/Effector helper T cells': 'Helper T cells',
#
#         'DC2': 'DC',
#         'pDC2': 'DC',
#
#         'Plasma cells': 'B cells',
#         'Age-associated B cells': 'B cells',
#         'Naive B cells': 'B cells',
#         'Memory B cells': 'B cells',
#
#         'Intermediate macrophages': 'Monocytes',
#         'Classical monocytes': 'Monocytes',
#         'Non-classical monocytes': 'Non-classical monocytes',
#
#         'CD16+ NK cells': 'NK cells'
#     }
#     celltypes = adata_LF.obs.set_index('low_level_majority_voting').rename(REMAP_CELLS, axis=0).reset_index()[
#         'low_level_majority_voting']
#     celltypes.index = adata_LF.obs.index
#     celltypes
#
#     for ts in adata_LF.obs['ts'].unique():
#         c = celltypes.loc[adata_LF.obs['ts'] == ts]
#         display((c.groupby(c).count() / len(c) * 100).sort_values(ascending=False).rename(ts))
#
#     df = misc.get_sample_annot()
#     for visit in ['V1', 'V2', 'V3']:
#         display(df.loc[df['SAMPLE:VISIT'] == visit, misc.BLOOD].mean().sort_values(ascending=False).rename(visit))


def genes_to_regions(gene_ids, peaks_and_genes, make_upper=True, bag_of_regions=False):
    if make_upper:
        gene_ids = gene_ids.str.upper()
        peaks_and_genes = peaks_and_genes.str.upper()
    regions = []
    for gene in gene_ids:
        if bag_of_regions:
            regions.extend(list(np.unique(peaks_and_genes.index[peaks_and_genes == gene])))
        else:
            regions.append(';'.join(np.unique(peaks_and_genes.index[peaks_and_genes == gene])))
    return regions


def get_scRNAseq_results(results_dir, model='scLM', celltype_cols=['PBMC', 'celltypist'], interaction=False,
                         subsample_sorted_cells=False, results=None):
    if isinstance(celltype_cols, str):
        celltype_cols = [celltype_cols]
    assert model in ['scLM', 'scLM_interaction', 't_test']
    if model == 'scLM_interaction':
        assert interaction
        model = 'scLM'
    assert set(celltype_cols).issubset(['PBMC', 'celltypist'])

    RENAME = {
        'time[T.T3m]': RNA_D90,
        'stim[T.LPS]': RNA_LPS,
        'time[T.T3m]:stim[T.LPS]': RNA_TRIM,

        'ts[T.T3m_RPMI]': RNA_D90,
        'ts[T.T0_LPS]': RNA_LPS,
        'ts[T.T3m_LPS]': RNA_TRIM,
        'ts[T.T3m_LPS_vs_T3m_RPMI]': RNA_TRIM_ABOVE_BCG,
    }

    if results is None:
        results = {}
    else:
        # if you want read additional results into an existing directory
        assert isinstance(results, dict)

    for celltype_col in celltype_cols:
        if model == 't_test':
            model_name = f'{model}.{celltype_col}{".subsampled" if subsample_sorted_cells and celltype_col != "PBMC" else ""}'
            results[model_name] = joblib.load(os.path.join(results_dir,
                                                           f'DE_with_{model}.{celltype_col}{".subsampled" if subsample_sorted_cells and celltype_col != "PBMC" else ""}.results.pckl'))

        else:
            assert model == 'scLM'
            if interaction:
                model_name = f'{model}_interaction.{celltype_col}{".subsampled" if subsample_sorted_cells and celltype_col != "PBMC" else ""}'
                results[model_name] = joblib.load(os.path.join(results_dir,
                                                               f'DE_{model}_pearson_residuals.interaction.{celltype_col}{".subsampled" if subsample_sorted_cells and celltype_col != "PBMC" else ""}.results.pckl'))

            else:
                assert model == 'scLM'
                model_name = f'{model}.{celltype_col}{".subsampled" if subsample_sorted_cells and celltype_col != "PBMC" else ""}'
                rpmi_results = joblib.load(os.path.join(results_dir,
                                                        f'DE_{model}_pearson_residuals.T0_RPMI.{celltype_col}{".subsampled" if subsample_sorted_cells and celltype_col != "PBMC" else ""}.results.pckl'))
                lps_results = joblib.load(os.path.join(results_dir,
                                                       f'DE_{model}_pearson_residuals.T0_LPS.{celltype_col}{".subsampled" if subsample_sorted_cells and celltype_col != "PBMC" else ""}.results.pckl'))
                if celltype_col == "PBMC":
                    lps_over_rpmi_results = joblib.load(os.path.join(results_dir,
                                                           f'DE_{model}_pearson_residuals.T3m_RPMI.{celltype_col}{".subsampled" if subsample_sorted_cells and celltype_col != "PBMC" else ""}.results.pckl'))
                results[model_name] = defaultdict(lambda: {})
                for celltype in rpmi_results:
                    for stat in rpmi_results[celltype]:
                        if 'ts[T.T3m_LPS]' not in lps_results[celltype][stat]:
                            lps_results[celltype][stat].loc[:, 'ts[T.T3m_LPS]'] = np.nan
                        results[model_name][celltype][stat] = pd.concat(
                            [rpmi_results[celltype][stat][['ts[T.T3m_RPMI]', 'ts[T.T0_LPS]']],
                             lps_results[celltype][stat]['ts[T.T3m_LPS]']], axis=1)
                        if celltype_col == "PBMC":
                            results[model_name][celltype][stat] = pd.concat(
                                [results[model_name][celltype][stat],
                                 lps_over_rpmi_results[celltype][stat]['ts[T.T3m_LPS]'].rename('ts[T.T3m_LPS_vs_T3m_RPMI]')], axis=1
                            )

        for celltype in results[model_name]:
            for stat in results[model_name][celltype]:
                results[model_name][celltype][stat] = results[model_name][celltype][stat].rename(RENAME, axis=1)

        if 'Non-classical monocytes' in results[model_name]:
            results[model_name]['Non-classical\nmonocytes'] = results[model_name]['Non-classical monocytes']
            del results[model_name]['Non-classical monocytes']

            for stat in results[model_name]['Non-classical\nmonocytes']:
                if RNA_TRIM not in results[model_name]['Non-classical\nmonocytes'][stat].columns:
                    results[model_name]['Non-classical\nmonocytes'][stat][RNA_TRIM] = np.nan

        if 'Regulatory T cells' in results[model_name]:
            results[model_name]['Tregs'] = results[model_name]['Regulatory T cells']
            del results[model_name]['Regulatory T cells']

        for stat in ['expFrac', 'minCells', 'meanExp']:
            for celltype in results[model_name]:
                idx = results[model_name][celltype]['pval'].index
                assert results[model_name][celltype]['coef'].index.equals(idx)
                if stat not in results[model_name][celltype]:
                    other_model_name = f't_test.{celltype_col}'
                else:
                    other_model_name = model_name
                if other_model_name in results:
                    results[model_name][celltype][stat] = results[other_model_name][celltype][stat].loc[idx]

    if model == 't_test':
        model_name = 't_test'
    elif model == 'scLM':
        if interaction:
            model_name = 'scLM_interaction'
        else:
            model_name = 'scLM'
    else:
        raise ValueError

    results[f'{model_name}{".subsampled" if subsample_sorted_cells else ""}'] = {}
    if 'PBMC' in celltype_cols:
        results[f'{model_name}{".subsampled" if subsample_sorted_cells else ""}']['PBMC'] = results[f'{model_name}.PBMC']['PBMC']
        # del results[f'{model_name}.PBMC']
    if 'celltypist' in celltype_cols:
        for c in results[f'{model_name}.celltypist{".subsampled" if subsample_sorted_cells else ""}']:
            results[f'{model_name}{".subsampled" if subsample_sorted_cells else ""}'][c] = results[f'{model_name}.celltypist{".subsampled" if subsample_sorted_cells else ""}'][c]
        # del results[f'{model_name}.celltypist{".subsampled" if subsample_sorted_cells else ""}']

    for celltype in results[f'{model_name}{".subsampled" if subsample_sorted_cells else ""}']:
        idx = results[f'{model_name}{".subsampled" if subsample_sorted_cells else ""}'][celltype]['pval'].index
        for stat in results[f'{model_name}{".subsampled" if subsample_sorted_cells else ""}'][celltype]:
            assert results[f'{model_name}{".subsampled" if subsample_sorted_cells else ""}'][celltype][stat].index.equals(idx)

    return results

#
# def gene_set_enrichment_test(
#         top_genes, gene_set_libraries, background, padj_method='fdr_bh', fisher_hypergeom_tol=1e-4,
#         check_hypergeom=False, check_background=False, adjust_gene_sets_to_background=False, padj_union=False,
#         min_gs=None, max_gs=None):
#     from statsmodels.stats.multitest import multipletests
#     assert len(top_genes) == len(set(top_genes))
#     assert utils.is_iterable(background) or not adjust_gene_sets_to_background
#     assert utils.is_iterable(background) or not check_background
#
#     if check_background:
#         assert all([g in background for g in top_genes])
#
#     background_size = len(background) if utils.is_iterable(background) else background
#
#     res_df = pd.DataFrame(columns=utils.ENRICHR_COLS)
#     for lib_name, library in gene_set_libraries:
#         results = []
#         for term in library:
#             gs = np.asarray(library[term]).copy()
#
#             if adjust_gene_sets_to_background:
#                 gs = set(background).intersection(gs)
#
#             if check_background:
#                 assert all([g in background for g in gs])
#
#             if min_gs is not None and len(gs) < min_gs:
#                 continue
#
#             if max_gs is not None and len(gs) > max_gs:
#                 continue
#
#             # hits are sorted in the order of top_genes
#             hits = [g for g in top_genes if g in gs]
#
#             overlap = '{}/{}'.format(len(hits), len(gs))
#             if len(hits) != 0:
#                 term_not_top = set(gs).difference(top_genes)
#                 top_not_term = set(top_genes).difference(gs)
#                 term_or_top = set(gs).union(top_genes)
#                 oddsratio, f_p = scipy.stats.fisher_exact(
#                     [[len(hits), len(top_not_term)], [len(term_not_top), background_size - len(term_or_top)]],
#                     alternative='greater'
#                 )
#                 results.append((lib_name, term, f_p, np.nan, np.nan, oddsratio, overlap, ';'.join(hits)))
#
#                 if check_hypergeom:
#                     k = len(hits)
#                     M = background_size
#                     n = len(gs)
#                     N = len(top_genes)
#                     hg_p = scipy.stats.hypergeom.sf(k - 1, M, n, N)
#                     assert f_p - hg_p < fisher_hypergeom_tol
#
#         if len(results) != 0:
#             df = pd.DataFrame(results, columns=utils.ENRICHR_COLS)
#             df['Adjusted P-value'] = multipletests(df['P-value'].values, method=padj_method)[1]
#             res_df = pd.concat([res_df, df.sort_values('P-value')])
#         else:
#             print('WARNING: No overlap for any term:', lib_name)
#
#     if padj_union:
#         res_df['Adjusted P-value'] = multipletests(res_df['P-value'].values, method=padj_method)[1]
#
#     return res_df.reset_index(drop=True)

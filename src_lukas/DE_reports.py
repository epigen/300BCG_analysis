#!/usr/bin/env python
# coding: utf-8

import os
import pandas as pd
from statsmodels.stats.multitest import multipletests
import re
import argparse

from misc import *
from bcg_utils import *

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles.numbers import BUILTIN_FORMATS, FORMAT_NUMBER_00
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK, BORDER_MEDIUM
from openpyxl.formatting import Rule
from openpyxl.styles import Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import Alignment


def set_column_width(columns=None, width='auto', extra=0, worksheet=None, filename=None, sheet_name=None):
    assert sum([worksheet is None, filename is None]) == 1

    if filename is not None:
        workbook = load_workbook(filename)
        worksheet = workbook[sheet_name]

    for column_cells in worksheet.columns:
        if columns is None or column_cells[0].column in columns:
            if isinstance(columns, dict) and columns[column_cells[0].column] != 'auto':
                length = columns[column_cells[0].column]
            elif width != 'auto':
                length = width
            else:
                length = max(len(str(cell.value or '')) for cell in column_cells) + extra
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    if filename is not None:
        workbook.save(filename)


def set_number_format(columns, offset=1, fmt=FORMAT_NUMBER_00, worksheet=None, filename=None, sheet_name=None):
    assert sum([worksheet is None, filename is None]) == 1

    if filename is not None:
        workbook = load_workbook(filename)
        worksheet = workbook[sheet_name]

    for column_cells in worksheet.columns:
        if column_cells[0].column in columns:
            for cell in column_cells[offset:]:
                cell.number_format = fmt

    if filename is not None:
        workbook.save(filename)


def set_background(columns, fill=None, color=None, worksheet=None, filename=None, sheet_name=None):
    assert sum([worksheet is None, filename is None]) == 1

    if fill is None:
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    if filename is not None:
        workbook = load_workbook(filename)
        worksheet = workbook[sheet_name]

    for column_cells in worksheet.columns:
        if column_cells[0].column in columns:
            for cell in column_cells:
                cell.fill = fill

    if filename is not None:
        workbook.save(filename)


def set_borders(columns, border=None, color=None, thickness=BORDER_THIN, n_rows=None, worksheet=None, filename=None, sheet_name=None):
    assert sum([worksheet is None, filename is None]) == 1

    if border is None:
        _border = Side(border_style=thickness, color=color)
        border = Border(left=_border, right=_border, top=_border, bottom=_border)

    if filename is not None:
        workbook = load_workbook(filename)
        worksheet = workbook[sheet_name]

    for column_cells in worksheet.columns:
        if columns is None or column_cells[0].column in columns:
            for cell in column_cells if n_rows is None else column_cells[:n_rows]:
                cell.border = border

    if filename is not None:
        workbook.save(filename)


def set_color_scale(columns, rule=None, string=None, font=None, border=None, fill=None, stopIfTrue=None,
                    offset=1, worksheet=None, filename=None, sheet_name=None):
    assert sum([worksheet is None, filename is None]) == 1
    assert sum([rule is None, string is None]) == 1

    if filename is not None:
        workbook = load_workbook(filename)
        worksheet = workbook[sheet_name]

    for c in columns:
        column_letter = get_column_letter(c)
        _rule = rule if rule is not None else FormulaRule(
            formula=['NOT(ISERROR(SEARCH("{}",{}{})))'.format(string, column_letter, 1 + offset)],
            stopIfTrue=stopIfTrue, font=font, border=border, fill=fill)
        worksheet.conditional_formatting.add(
            '{col}{row1}:{col}{row2}'.format(col=column_letter, row1=1 + offset, row2=len(worksheet[column_letter])),
            _rule)

    if filename is not None:
        workbook.save(filename)


def set_alignment(columns=None, v_align=None, h_align=None, wrap_text=None, offset=1, worksheet=None, filename=None, sheet_name=None):
    assert sum([worksheet is None, filename is None]) == 1

    if filename is not None:
        workbook = load_workbook(filename)
        worksheet = workbook[sheet_name]

    for column_cells in worksheet.columns:
        if columns is None or column_cells[0].column in columns:
            for cell in column_cells[offset:]:
                cell.alignment = Alignment(vertical=v_align, horizontal=h_align, wrapText=wrap_text)

    if filename is not None:
        workbook.save(filename)


def _get_sheet_name(celltype, model_name, pathway, coef=None):
    if coef is None:
        return '{} ({})'.format(pathway, model_name if model_name is not None else rename_celltype(celltype)) if pathway is not None \
                else '{}{}'.format(rename_celltype(celltype, capitalize=True), ' ({})'.format(model_name) if model_name is not None else '')
    else:
        assert pathway is not None
        return '{} ({}, {})'.format(pathway, coef.replace('vsV1', '').replace('_TIME', ''), model_name if model_name is not None else rename_celltype(celltype))


def save_DE_excel_file(df, writer, sheet_name='Sheet1'):
    assert not hasattr(df.index, 'levels')
    _n_col_levels = len(df.columns.levels) if hasattr(df.columns, 'levels') else 1
    df.to_excel(writer, sheet_name=sheet_name, merge_cells=True, freeze_panes=(_n_col_levels, 1))
    
    
def format_DE_excel_file(df, excel_fn, sheet_name='Sheet1', fdr='$A$1', numeric_cols=[], hyperlinks=None, color_scales=[], alignments=[], fdr_cols=[], col_widths=None):
    assert not hasattr(df.index, 'levels')
    _col_idx = lambda name, level: 2 + np.arange(df.shape[1])[(df.columns.get_level_values(level) if level is not None else df.columns).values == name]
    _n_col_levels = len(df.columns.levels) if hasattr(df.columns, 'levels') else 1
    
    wb = load_workbook(excel_fn)

    # general maintenance
    wb[sheet_name].cell(row=1, column=1).fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
    if col_widths is not None:
        set_column_width(worksheet=wb[sheet_name], columns=col_widths)
    if _n_col_levels != 1:
        wb[sheet_name].delete_rows(_n_col_levels + 1)
     
    # number format
    for name, level, fmt in numeric_cols:
        set_number_format(columns=_col_idx(name, level), offset=_n_col_levels, worksheet=wb[sheet_name], fmt=fmt)
    
    # sort and filter pop-downs
    wb[sheet_name].auto_filter.ref = 'A{}:{}{}'.format(_n_col_levels,
                                                       get_column_letter(df.shape[1] + 1 - (len(hyperlinks) if hyperlinks is not None else 0)),
                                                       df.shape[0] + _n_col_levels)
    
    # blank fields should be always without fill
    for col in 2 + np.arange(df.shape[1]):
        rule = FormulaRule(formula=['ISBLANK({}{})'.format(get_column_letter(col), _n_col_levels + 1)], fill=PatternFill(fill_type=None), stopIfTrue=True)
        set_color_scale(columns=[col], rule=rule, worksheet=wb[sheet_name], offset=_n_col_levels)
    
    # diverging colour scales
    for _name, _level, _start, _mid, _end in color_scales:
        for _col in _col_idx(_name, _level):
            _col_letter = get_column_letter(_col)
            wb[sheet_name].conditional_formatting.add(
                '{}{}:{}{}'.format(_col_letter, _n_col_levels, _col_letter, _n_col_levels + df.shape[0]),
                ColorScaleRule(start_type=_start[0], start_value=_start[1], start_color=_start[2],
                               mid_type=_mid[0], mid_value=_mid[1], mid_color=_mid[2],
                               end_type=_end[0], end_value=_end[1], end_color=_end[2]))
    
    for _name, _level, _algn in alignments:
        if _name == df.index.name:
            set_alignment(columns=[1], h_align=_algn, v_align='center', wrap_text=False, offset=_n_col_levels, worksheet=wb[sheet_name])
        else:
            for _col in _col_idx(_name, _level):
                for row in range(_n_col_levels + 1, df.shape[0] + _n_col_levels + 1):
                    alignment = copy.copy(wb[sheet_name].cell(row=row, column=_col).alignment)
                    alignment.horizontal = _algn
                    wb[sheet_name].cell(row=row, column=_col).alignment = alignment

    # bold italics for FDR threshold
    if fdr is not None:
        for padj_col_names in fdr_cols:
            padj_col = set(2 + np.arange(df.shape[1]))
            for i, _col in enumerate(padj_col_names) if is_iterable(padj_col_names) else [(None, padj_col_names)]:
                padj_col = padj_col.intersection(_col_idx(_col, i))
            assert len(padj_col) == 1
            padj_col = padj_col.pop()
            for target_col_names in fdr_cols[padj_col_names]:
                target_col = set(2 + np.arange(df.shape[1]))
                for i, _col in enumerate(target_col_names) if is_iterable(target_col_names) else [(None, target_col_names)]:
                    target_col = target_col.intersection(_col_idx(_col, i))
                assert len(target_col) == 1
                target_col = target_col.pop()
                set_color_scale(columns=[target_col],
                                rule=FormulaRule(formula=['{}{}<={}'.format(get_column_letter(padj_col), _n_col_levels + 1, fdr)],
                                                 font=Font(bold=True)),
                                worksheet=wb[sheet_name], offset=_n_col_levels)

    # setup hyperlinks
    if hyperlinks is not None:
        for col_name, col_level in hyperlinks:
            _col = [1] if col_name == '__index__'  and col_level is None else _col_idx(col_name, col_level)
            assert len(_col) == 1
            _col = _col[0]
            h_name, h_level, _bold, _center = hyperlinks[(col_name, col_level)]
            _hyperlink = _col_idx(h_name, h_level)
            assert len(_hyperlink) == 1
            _hyperlink = _hyperlink[0]
            for row in range(_n_col_levels + 1, df.shape[0] + _n_col_levels + 1):
                if wb[sheet_name].cell(row=row, column=_col).value is not None:
                    wb[sheet_name].cell(row=row, column=_col).hyperlink = str(wb[sheet_name].cell(row=row, column=_hyperlink).value)
                    wb[sheet_name].cell(row=row, column=_col).style = 'Hyperlink'
                    if _bold:
                        font = copy.copy(wb[sheet_name].cell(row=row, column=_col).font)
                        font.bold = True
                        wb[sheet_name].cell(row=row, column=_col).font = font
                    if _center:
                        alignment = copy.copy(wb[sheet_name].cell(row=row, column=_col).alignment)
                        alignment.horizontal = 'center'
                        wb[sheet_name].cell(row=row, column=_col).alignment = alignment
        
    # to do this at the very end
    if hyperlinks is not None:
        to_remove = [_col_idx(hyperlinks[(n, l)][0], hyperlinks[(n, l)][1])[0] for n, l in hyperlinks]
        for i, idx in enumerate(sorted(to_remove)):
            wb[sheet_name].delete_cols(idx - i)
    
    set_borders(columns=None, color=GRAY, worksheet=wb[sheet_name])
    set_borders(columns=None, color=BLACK, thickness=BORDER_MEDIUM, n_rows=_n_col_levels, worksheet=wb[sheet_name])
    set_borders(columns=[1], color=BLACK, thickness=BORDER_MEDIUM, worksheet=wb[sheet_name])
    
    wb.save(excel_fn)


def _rename_columns_dir(columns):
    return {c: c.replace(S_RANK, 'signed_rank').
               replace(PADJ, 'adjusted_P').
               replace(PVAL, 'P_value').
               replace(LFC, 'log2_FC').
               replace('vsV1', '').
               replace('_TIME', '').
               replace('TIME', 'BC_time').
               replace('DATE', 'BC_date').
               replace('perc.A', 'mean_signal.perc.').
               replace('A', 'mean_signal.log2_count') for c in columns}

PEAK_COLUMNS = ['position', 'length', 'reg_feature', 'reg_feature_id', 'chr', 'gene_name', 'distance', 'location', 'characterization', 'region_hyperlink', 'gene_hyperlink', 'reg_feature_hyperlink']
# PEAK_COLUMNS_RENAME = {'gene_name': 'gene', 'reg_feature': 'regulatory_feature'}

ENSEMBL_REGION = 'http://www.ensembl.org/Homo_sapiens/Location/View?r=' # + {chr}%3A{start}-{end}
ENSEMBL_GENE = 'http://www.ensembl.org/Homo_sapiens/Gene/Summary?g=' # + gene
ENSEMBL_REG_FEATURE = 'http://www.ensembl.org/Homo_sapiens/Regulation/Summary?rf=' # + reg_feature_id
# GENECARDS_GENE = 'http://www.genecards.org/cgi-bin/carddisp.pl?gene=' # + gene

ORANGE = 'FF8C00'
GRAY = '808080'
LIGHT_GRAY = 'DCDCDC'
WHITE = 'ffffff'
BLACK = '000000'
RED = 'e41a1c'
BLUE = '377eb8'
LIGHT_BLUE = '89CFF0'
LIGHT_RED = 'FF7F7F'
GREEN = '4daf4a'
BROWN = 'a18459'

LFC, PADJ, PVAL, S_RANK = 'Coef', 'p.adj', 'p.value', 's_neg_log10_p.value'
COLUMN = '{metric}.{coef}'


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--celltype', required=True, choices=['PBMC', 'monocyte', 'nkcell', 'cd8t'])
    parser.add_argument('--model', required=True)
    parser.add_argument('--coefs', nargs='+', required=True, choices=['V2vsV1', 'V3vsV1', 'V2vsV1_TIME', 'V3vsV1_TIME'])
    parser.add_argument('--n_perms', type=int, default=5000)
    parser.add_argument('--model_name', help='This will be included in the sheet name')
    parser.add_argument('--libraries', nargs='+')
    parser.add_argument('--regions', choices=['TSS_ONLY', 'TSS_PROXIMAL', 'TSS_AND_GENES', 'TSS_10000', 'ALL_REGIONS'])
    parser.add_argument('--pathways', choices=[BCG_PATHWAYS, KEGG, WIKI_PATHWAYS])
    parser.add_argument('--snp_hits', help='CSV file with two columns: SNP_id and region_id')
    parser.add_argument('--my_genes', nargs='+', help='list of genes to filter out from all regions')
    parser.add_argument('--fdr', type=float, default=0.05)
    parser.add_argument('--results_dir', default='results')
    parser.add_argument('--reports_dir', default='results/reports')
    parser.add_argument('--peak_annot_fn', default=PEAK_ANNOT_ALL_FN)
    args = parser.parse_args()
    assert args.libraries is not None or args.regions is None

    print('celltype', args.celltype)
    print('model', args.model)
    print('coefs', args.coefs)
    print('model_name', args.model_name)
    print('pathways', args.pathways)
    print('regions', args.regions)
    print('pathways', args.pathways)
    print('snp_hits', args.snp_hits)
    print('my_genes', args.my_genes)
    print('fdr', args.fdr)
    print('results_dir', args.results_dir)
    print('reports_dir', args.reports_dir)
    print('peak_annot_fn', args.peak_annot_fn)

    if args.libraries is None:
        make_diffCA_report(celltype=args.celltype, model=args.model, coefs=args.coefs, peak_annot_fn=args.peak_annot_fn,
                     results_dir=args.results_dir, reports_dir=args.reports_dir, model_name=args.model_name,
                     pathways=args.pathways, snp_hits=args.snp_hits, my_genes=args.my_genes, fdr=args.fdr)
    else:
        make_gsea_report(celltype=args.celltype, model=args.model, coefs=args.coefs, n_perms=args.n_perms,
                     results_dir=args.results_dir, reports_dir=args.reports_dir, model_name=args.model_name,
                     libraries=args.libraries, regions=args.regions, fdr=args.fdr)


def make_gsea_report(celltype, model, coefs, n_perms, results_dir, reports_dir, model_name, libraries, regions, fdr):
    excel_fn = os.path.join(reports_dir, '{}.GSEA.{}.{}.{}.xlsx'.format(PROJECT, celltype, model, regions))
    data = {}
    with pd.ExcelWriter(excel_fn) as writer:
        for library in libraries:
            data[library] = {}
            for coef in coefs:
                _gsea_fn = gsea_fn(library=library, celltype=celltype, model=model, regions=regions, coef=coef, n_perms=n_perms, results_dir=results_dir) + '2'
                print(_gsea_fn)
                gsea_df = pd.read_csv(_gsea_fn, index_col=0)
                if library == 'BCG_pathways_KEGG_2019_Human':
                    _bcg_pthws = gene_set_library(os.path.join(METADATA, 'gene_set_libraries', BCG_PATHWAYS)).keys()
                    gsea_df = gsea_df.loc[gsea_df['description'].isin(_bcg_pthws)]
                gsea_df.index.name = 'rank'
                if library == WIKI_PATHWAYS:
                    gsea_df['description'] = gsea_df['description'].str.replace(' WP[0-9]+$', '')
                gsea_df = gsea_df.reset_index().set_index('description')
                gsea_df = gsea_df.rename({'p_value': 'P_value', 'adjusted_p_value': 'adjusted_P'}, axis=1)
                gsea_df['ledge_genes'] = gsea_df['ledge_genes'].apply(lambda x: eval(x))
                gsea_df.loc[gsea_df['ES'] < 0, 'ledge_genes'] = gsea_df.loc[gsea_df['ES'] < 0, 'ledge_genes'].apply(lambda x: x[::-1])

                try:
                    ranked_genes_df = get_gsea_ranked_genes(library, celltype, model, regions, coef, n_perms=n_perms, results_dir=results_dir)
                    gsea_df['ledge_regions'] = [', '.join(ledge_regions) for ledge_regions in get_ledge_regions(gsea_df['ledge_genes'], ranked_genes_df)]
                except FileNotFoundError:
                    print('WARNING: missing ranked regions file, removing column "ledge_regions"')

                gsea_df['ledge_genes'] = gsea_df['ledge_genes'].apply(lambda x: ', '.join(x))
                gsea_df = gsea_df.drop(['gene_set_library', 'comparison'], axis=1)
                gsea_df = gsea_df[['rank', 'ES', 'NES', 'P_value', 'adjusted_P', 'gene_set_size', 'ledge_genes'] + \
                                  (['ledge_regions'] if 'ledge_regions' in gsea_df.columns else [])]

                sheet_name = _get_sheet_name(celltype, model_name, library.split('_')[0].replace('Pathways', 'Pthws') if library != 'BCG_pathways_KEGG_2019_Human' else 'BCG_2', coef)
                data[library][coef] = gsea_df
                save_DE_excel_file(data[library][coef], writer, sheet_name=sheet_name)
                print(sheet_name, 'saved')

    for library in libraries:
        for coef in coefs:
            sheet_name = _get_sheet_name(celltype, model_name, library.split('_')[0].replace('Pathways', 'Pthws') if library != 'BCG_pathways_KEGG_2019_Human' else 'BCG_2', coef)
            format_DE_excel_file(data[library][coef], excel_fn, sheet_name=sheet_name, fdr=0.05,
                numeric_cols=[('ES', 0, FORMAT_NUMBER_00), ('NES', 0, FORMAT_NUMBER_00),
                              ('P_value', 0, BUILTIN_FORMATS[11]), ('adjusted_P', 0, BUILTIN_FORMATS[11])],
                color_scales=[('NES', 0, ('min', None, LIGHT_BLUE), ('num', 0, WHITE), ('max', None, LIGHT_RED)),
                              ('gene_set_size', 0, ('num', 5, ORANGE), (None, None, None), ('num', 15, WHITE))],
                alignments=[('gene_set_size', 0, 'center'), ('rank', 0, 'center'), ('description', 0, 'left')],
                fdr_cols={'adjusted_P': ['adjusted_P', 'P_value', 'ES', 'NES']},
                col_widths={i + 1: w for i, w in
                            enumerate([50, 10, 12, 12, 16, 16, 19, 100] + ([100] if 'ledge_regions' in data[library][coef].columns else []))}
            )
            print(sheet_name, 'formatted')

def make_diffCA_report(celltype, model, coefs, peak_annot_fn, results_dir, reports_dir, model_name=None, pathways=None, snp_hits=None, my_genes=None, fdr=0.05):
    peaks_info_df = get_peak_annot(fn=peak_annot_fn)
    peaks_info_df['position'] = peaks_info_df['chr'].str.replace('^chr', '') + ':' + peaks_info_df['start'].astype(str) + '-' + peaks_info_df['end'].astype(str)
    peaks_info_df['distance'] = peaks_info_df['distance'].astype(pd.Int64Dtype())
    peaks_info_df['gene_hyperlink'] = ENSEMBL_GENE + peaks_info_df['gene_id']
    peaks_info_df['region_hyperlink'] = ENSEMBL_REGION + peaks_info_df['position'].replace(':', '%3A')
    peaks_info_df['reg_feature_hyperlink'] = ENSEMBL_REG_FEATURE + peaks_info_df['reg_feature_id']

    excel_template = os.path.join(reports_dir, '{}.GRCh38.diffCA{}.{}.{}.xlsx'.format(PROJECT, '{specials}', celltype, model))
    if snp_hits is not None:
        excel_fn = excel_template.format(specials='.{}'.format(re.sub('.csv$', '', snp_hits.split(os.path.sep)[-1])))
    elif my_genes is not None:
        excel_fn = excel_template.format(specials='.{}'.format('_'.join(my_genes)))
    elif pathways is not None:
        excel_fn = excel_template.format(specials='.{}'.format(pathways))
    else:
        excel_fn = excel_template.format(specials='')

    if pathways is not None:
        pathways = gene_set_library(os.path.join(METADATA, 'gene_set_libraries', pathways))
    else:
        pathways = None

    _de_fn = de_fn(data='results_p5', celltype=celltype, model=model, results_dir=results_dir)
    print(_de_fn)
    de_df = pd.read_csv(_de_fn, index_col=0)
    assert de_df.index.tolist() == de_df['genes'].tolist()
    de_df['perc.A'] = round((de_df['A'] - np.log2(peaks_info_df.loc[de_df.index, 'length'] / 500)).rank(method='max', pct=True) * 100).astype(int)
    # adjust pvals and rank with signed negative log10 pvals
    for coef in coefs:
        _padj_col = COLUMN.format(metric=PADJ, coef=coef)
        _pval_col = COLUMN.format(metric=PVAL, coef=coef)
        _s_rank_col = COLUMN.format(metric=S_RANK, coef=coef)
        _coef_col = COLUMN.format(metric=LFC, coef=coef)

        de_df[_padj_col] = multipletests(de_df[_pval_col], method='fdr_bh')[1]
        de_df[_s_rank_col] = -np.log10(de_df[_pval_col])
        de_df.loc[de_df[_coef_col] < 0, _s_rank_col] *= -1

    columns = ['perc.A'] \
              + [COLUMN.format(metric=LFC, coef=coef) for coef in coefs] \
              + [COLUMN.format(metric=PADJ, coef=coef) for coef in coefs] \
              + [COLUMN.format(metric=PVAL, coef=coef) for coef in coefs] \
              + [COLUMN.format(metric=S_RANK, coef=coef) for coef in coefs]
    de_df = de_df[columns].rename(_rename_columns_dir(columns), axis=1)
    _peaks_df = peaks_info_df.loc[de_df.index, PEAK_COLUMNS].copy()
    _peaks_df.loc[_peaks_df['characterization'] == 'NONE', 'characterization'] = ''
    _peaks_df.loc[_peaks_df['reg_feature'] == 'reg_NONE', 'reg_feature'] = ''
    _peaks_df.loc[_peaks_df['reg_feature'] == 'reg_NONE', 'reg_feature_id'] = ''
    _peaks_df.loc[_peaks_df['reg_feature'] == 'reg_NONE', 'reg_feature_hyperlink'] = ''
    _peaks_df.columns = ['{}.{}'.format('hyperlink' if c.endswith('_hyperlink') else 'peak_annotation', c) for c in _peaks_df.columns]
    de_df = pd.concat([_peaks_df.loc[:, _peaks_df.columns.str.startswith('peak_annotation')],
                       de_df,
                       _peaks_df.loc[:, _peaks_df.columns.str.startswith('hyperlink')]], axis=1)

    de_df.index.name = None
    assert de_df.index.is_unique
    de_df.columns = de_df.columns.str.split('.', n=1, expand=True)
    de_df.columns.names = (fdr, 'region')

    if snp_hits is not None:
        snp_regions_df = pd.read_csv(snp_hits).set_index('region')
        assert snp_regions_df.index.is_unique
        snp_regions_df.columns = pd.MultiIndex.from_tuples([('', 'SNP')])
        de_df = de_df.loc[de_df.index.isin(snp_regions_df.index)]
        de_df = pd.concat([de_df, snp_regions_df.loc[de_df.index, ('', 'SNP')]], axis=1)
    elif my_genes is not None:
        de_df = de_df.loc[de_df[('peak_annotation', 'gene_name')].isin(my_genes)]

    # debug
    # de_df = de_df.head(1000)

    data = {}
    with pd.ExcelWriter(excel_fn) as writer:
        for pathway in [None] if pathways is None else pathways.keys():
            sheet_name = _get_sheet_name(celltype, model_name, pathway)
            data[pathway] = de_df if pathway is None else de_df.loc[
                de_df[('peak_annotation', 'gene_name')].isin(pathways[pathway])]
            save_DE_excel_file(data[pathway], writer, sheet_name=sheet_name)
            print(sheet_name, 'saved')

    for pathway in [None] if pathways is None else pathways.keys():
        sheet_name = _get_sheet_name(celltype, model_name, pathway)
        format_DE_excel_file(data[pathway], excel_fn, sheet_name=sheet_name,
            fdr='$A$1',
            numeric_cols=[('log2_FC', 0, FORMAT_NUMBER_00),
                          ('adjusted_P', 0, BUILTIN_FORMATS[11]),
                          ('P_value', 0, BUILTIN_FORMATS[11]),
                          ('signed_rank', 0, FORMAT_NUMBER_00)],
            hyperlinks={('position', 1): ('region_hyperlink', 1, False, True),
                        ('gene_name', 1): ('gene_hyperlink', 1, False, True),
                        ('reg_feature_id', 1): ('reg_feature_hyperlink', 1, False, True)},
            color_scales=[('signed_rank', 0, ('num', -np.log2(6), LIGHT_BLUE), ('num', 0, WHITE), ('num', np.log2(6), LIGHT_RED)),
                          ('log2_FC', 0, ('num', -np.log2(2), LIGHT_BLUE), ('num', 0, WHITE), ('num', np.log2(2), LIGHT_RED)),
                          ('perc.', 1, ('min', None, ORANGE), (None, None, None), ('num', 25, WHITE)),
                          ('distance', 1, ('num', 10000, WHITE), (None, None, None), ('num', 100000, BROWN))],
            alignments=[('chr', 1, 'center')],
            fdr_cols={('adjusted_P', coef): [('adjusted_P', coef), ('P_value', coef), ('signed_rank', coef), ('log2_FC', coef)] for coef in _rename_columns_dir(coefs).values()},
            col_widths={i + 1: w for i, w in enumerate([20, 25, 12, 24, 21, 8.5, 16, 13, 20, 20, 13, 8, 8, 10, 10, 10, 10, 8, 8, 15])}
        )
        print(sheet_name, 'formatted')


if __name__ == '__main__':
    main()
